## cd C:\Moraes\Pesquisa\LabCim\Reservas\LabCim_Manager_Fase1_3\labcim_manager
## python -m venv .venv
## Set-ExecutionPolicy -ExecutionPolicy RemoteSigned -Scope CurrentUser
## .venv\Scripts\activate
## pip install -r requirements.txt
## streamlit run app.py

from __future__ import annotations

from datetime import date, datetime, time, timedelta
from email.message import EmailMessage
import hashlib
from html import escape
from io import BytesIO
from numbers import Integral, Real
import os
import secrets as py_secrets
import smtplib
from time import perf_counter
import zipfile
from pathlib import Path
import re

import pandas as pd
import plotly.express as px
import qrcode
import streamlit as st
import streamlit.components.v1 as components
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from labcim_manager.db import (
    change_booking_status,
    change_maintenance_status,
    connect,
    create_attachment,
    create_booking,
    create_corrective_ticket,
    create_equipment,
    create_access_code_record,
    create_preventive_activity,
    create_project,
    create_project_service,
    create_supply,
    create_supply_lot,
    create_supply_movement,
    create_user,
    deactivate_attachment,
    get_active_user_by_email,
    get_attachment,
    get_latest_attachment_for_entity,
    import_base_xlsx,
    inactivate_supply_lot,
    init_db,
    is_operational_database_empty,
    list_attachments,
    list_booking_status_history,
    list_equipment_for_spare_part,
    list_maintenance_status_history,
    list_project_services,
    list_supply_lots,
    list_spare_parts_for_equipment,
    list_upcoming_preventive_maintenance,
    log_notification,
    query_df,
    seed_default_pops,
    set_spare_part_equipment_links,
    table_counts,
    update_corrective_ticket,
    update_equipment_master,
    update_equipment_operational_info,
    update_legacy_attachment_path,
    update_preventive_activity,
    update_project,
    update_project_service,
    update_supply_lot,
    update_supply,
    update_user,
    inactivate_project_service,
    inactivate_maintenance_record,
    verify_access_code_record,
)
from labcim_manager.storage import (
    LocalStorageBackend,
    R2StorageBackend,
    R2_REQUIRED_KEYS,
    StorageConfigurationError,
    get_active_storage_backend,
    get_storage_backend_for_name,
    resolve_config_value,
)

APP_TITLE = "LabCim Manager"
APP_SUBTITLE = "Gestão integrada, rastreabilidade e governança operacional do LabCim"
DB_PATH = Path("data/labcim_manager.db")
BASE_XLSX = Path("data/LabCim_Base.xlsx")
LOGO_PATH = Path("assets/logo_labcim.png")
POP_DIR = Path("assets/pops")
ACCESS_CODE_TTL_MINUTES = 10
CACHE_TTL_SECONDS = 120
DB_CONNECTION_KEY = "labcim_db_connection"
DB_CONNECTION_FINGERPRINT_KEY = "labcim_db_connection_fingerprint"
PERF_EVENTS_KEY = "labcim_perf_events"
APP_CACHE_GENERATION_KEY = "labcim_app_cache_generation"
REPORT_EXCEL_BYTES_KEY = "labcim_report_excel_bytes"
REPORT_EXCEL_SIGNATURE_KEY = "labcim_report_excel_signature"
REPORT_CACHE_TABLES = (
    "bookings",
    "maintenance_preventive",
    "maintenance_corrective",
    "supplies",
    "supply_lots",
    "supply_movements",
    "projects",
    "project_services",
    "equipment",
    "attachments",
)

LAB_BLUE = "#0033A0"
LAB_CYAN = "#00AEEF"
LAB_DARK = "#102A43"
LAB_BG = "#F7FAFC"


STATUS_LABELS = {
    "scheduled": "Agendada",
    "done": "Concluída",
    "cancelled": "Cancelada",
    "no_show": "Não compareceu",
}
STATUS_REVERSE = {v: k for k, v in STATUS_LABELS.items()}
BOOKING_FINAL_STATUSES = {"done", "cancelled", "no_show"}

EQUIPMENT_STATUS_LABELS = {
    "available": "Apto",
    "restricted": "Uso restrito",
    "maintenance": "Em manutenção",
    "inactive": "Inativo",
}
EQUIPMENT_STATUS_REVERSE = {v: k for k, v in EQUIPMENT_STATUS_LABELS.items()}
EQUIPMENT_DOCUMENT_ROLE_LABELS = {
    "pop": "POP / procedimento operacional",
    "manual": "Manual",
    "certificate": "Certificado",
    "checklist": "Checklist",
    "technical_document": "Documento técnico",
    "other": "Outro",
}
EQUIPMENT_DOCUMENT_ROLE_REVERSE = {v: k for k, v in EQUIPMENT_DOCUMENT_ROLE_LABELS.items()}

ROLE_LABELS = {
    "member": "Membro",
    "manager": "Gerente",
    "admin": "Administrador",
}
ROLE_REVERSE = {v: k for k, v in ROLE_LABELS.items()}

BOOLEAN_LABELS = {0: "Não", 1: "Sim", False: "Não", True: "Sim"}
SUPPLY_TYPES = ["Insumo", "Peça de reposição"]
PROJECT_STATUSES = ["em andamento", "concluído", "pausado", "cancelado", "arquivado"]
SERVICE_STATUSES = ["em andamento", "aguardando amostras", "em análise", "concluído", "cancelado", "arquivado"]
SERVICE_TYPES = ["Análise", "Ensaio", "Caracterização", "Preparação", "Relatório", "Outro"]
PREVENTIVE_STATUSES = ["pendente", "realizado", "reprovado", "reagendado", "cancelado"]
CORRECTIVE_STATUSES = ["aberto", "em análise", "aguardando peça", "enviado para fornecedor", "concluído", "cancelado"]
MAINTENANCE_JUSTIFICATION_STATUSES = {"reprovado", "reagendado", "cancelado"}
SUPPLY_CONSUMPTION_MOVEMENT_TYPES = {"saída", "saida", "descarte", "ajuste negativo"}
LOT_EXPIRATION_ALERT_STATUSES = {"Vencido", "Vence em até 60 dias"}

PAGE_LABELS = [
    "Painel inicial",
    "Reservas",
    "Equipamentos",
    "Insumos",
    "Usuários",
    "Projetos",
    "Manutenção",
    "QR Codes",
    "Relatórios",
    "Importar base",
]
SIDEBAR_PAGE_KEY = "labcim_active_sidebar_page"
SIDEBAR_URL_PAGE_KEY = "labcim_sidebar_url_page"
SCROLL_TO_TOP_PAGE_KEY = "labcim_scroll_to_top_page"
PAGE_ICONS = {
    "Painel inicial": "🏠",
    "Reservas": "📅",
    "Equipamentos": "🔬",
    "Insumos": "📦",
    "Manutenção": "🛠",
    "QR Codes": "🔳",
    "Projetos": "📁",
    "Relatórios": "📊",
    "Usuários": "👥",
    "Importar base": "⚙️",
}
NAVIGATION_SECTIONS = (
    ("Painel", ("Painel inicial",)),
    ("Operação", ("Reservas", "Equipamentos", "Insumos", "Manutenção", "QR Codes")),
    ("Gestão", ("Projetos", "Relatórios")),
    ("Administração", ("Usuários", "Importar base")),
)

COLUMN_LABELS = {
    "id": "ID",
    "equipment_code": "Código",
    "equipment_name": "Equipamento",
    "lab_unit": "Unidade",
    "location": "Localização",
    "requires_operator": "Requer operador?",
    "operational_status": "Status operacional",
    "unavailable_functions": "Funcionalidades indisponíveis",
    "max_sample_capacity": "Capacidade máxima",
    "capacity_unit": "Unidade da capacidade",
    "capacity_enforced": "Bloqueia acima da capacidade?",
    "technical_manager": "Gestor técnico",
    "pop_title": "POP",
    "pop_path": "Arquivo POP",
    "pop_version": "Versão do POP",
    "pop_updated_at": "Atualização do POP",
    "pop_responsible": "Responsável pelo POP",
    "document_notes": "Observações documentais",
    "responsible_name": "Responsável",
    "responsible_phone": "Telefone do responsável",
    "supply_type": "Tipo de item",
    "supply_name": "Insumo",
    "supply_code": "Código interno",
    "commercial_name": "Nome comercial",
    "manufacturer": "Fabricante",
    "manufacturer_code": "Código do fabricante",
    "category": "Categoria",
    "physical_state": "Estado físico",
    "application_function": "Função/aplicação",
    "addition_mode": "Modo de adição",
    "compatible_model_family": "Modelo/família compatível",
    "unit": "Unidade",
    "current_quantity": "Saldo atual",
    "minimum_quantity": "Estoque mínimo",
    "lot": "Lote",
    "lot_code": "Lote",
    "supply_lot_id": "ID do lote",
    "supply_lot_code": "Lote",
    "supply_lot_expiration_date": "Validade do lote",
    "supplier_name": "Fornecedor",
    "initial_quantity": "Quantidade inicial",
    "lot_status": "Status do lote",
    "expiration_date": "Validade",
    "received_date": "Recebido em",
    "certificate_path": "Certificado de análise",
    "safety_doc_path": "FDS/FISPQ",
    "technical_doc_path": "Ficha técnica/caracterização",
    "density": "Massa específica",
    "recommended_concentration": "Faixa de concentração",
    "recommended_temperature": "Faixa de temperatura",
    "characterization_summary": "Caracterização",
    "movement_type": "Movimentação",
    "movement_date": "Data",
    "quantity": "Quantidade",
    "document_path": "Documento",
    "stock_status": "Status de estoque",
    "association_notes": "Observação da associação",
    "alerta": "Alerta",
    "quantity_to_minimum": "Falta para o mínimo",
    "days_until_expiration": "Dias até validade",
    "certificate_status": "Status do certificado",
    "movement_document_status": "Documento/anexo da movimentação",
    "movement_count": "Movimentações",
    "consumed_quantity": "Quantidade consumida",
    "project_label": "Projeto",
    "service_label": "Serviço/análise",
    "supply_lot_label": "Lote",
    "lot_supplier_name": "Fornecedor do lote",
    "lot_location": "Local do lote",
    "active": "Ativo?",
    "is_active": "Ativo?",
    "association_active": "Associação ativa?",
    "notes": "Observações",
    "created_at": "Criado em",
    "full_name": "Nome completo",
    "phone_e164": "Celular",
    "role": "Perfil",
    "department": "Departamento/Programa",
    "advisor_name": "Orientador(a)",
    "email": "E-mail",
    "training_completed": "Treinamento concluído?",
    "project_code": "Código do projeto",
    "project_name": "Projeto",
    "objective": "Objetivo",
    "funding_source": "Fonte de financiamento",
    "requester_name": "Solicitante",
    "coordinator_name": "Coordenador/responsável",
    "title": "Título",
    "service_code": "Código do serviço/análise",
    "service_title": "Serviço/análise",
    "service_type": "Tipo de serviço/análise",
    "requested_date": "Data solicitada",
    "expected_date": "Data prevista",
    "completed_date": "Data concluída",
    "responsible_name": "Responsável",
    "start_date": "Início do projeto",
    "end_date": "Fim do projeto",
    "start_datetime": "Início",
    "end_datetime": "Fim",
    "sample_count": "Amostras",
    "purpose": "Finalidade/observações",
    "status": "Status",
    "previous_status": "Status anterior",
    "new_status": "Novo status",
    "justification": "Justificativa",
    "changed_by_name": "Usuário",
    "changed_by_email": "E-mail do usuário",
    "changed_at": "Alterado em",
    "inactive_reason": "Motivo de inativação",
    "inactive_at": "Inativado em",
    "scheduled_reference": "Referência de agenda",
    "reference_label": "Próxima data",
    "operator": "Operador",
    "executante": "Executante",
    "solicitante": "Solicitante",
    "performed_by": "Executante",
    "blocks_booking": "Bloqueia reservas?",
    "planned_end_date": "Fim previsto",
}


def setup_page() -> None:
    st.set_page_config(
        page_title=APP_TITLE,
        page_icon="🔬",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    st.markdown(
        f"""
        <style>
        .stApp {{ background: {LAB_BG}; }}
        .stApp, .stMarkdown, .stText, p, label, span, div {{
            color: {LAB_DARK};
        }}
        .main .block-container {{ padding-top: 1.5rem; padding-bottom: 3rem; }}
        h1, h2, h3 {{ color: {LAB_DARK}; }}
        [data-testid="stSidebar"] {{
            background: linear-gradient(180deg, #FFFFFF 0%, #EEF7FF 100%);
            border-right: 1px solid #E6EEF8;
        }}
        [data-testid="stSidebar"] * {{
            color: {LAB_DARK} !important;
        }}
        [data-testid="stSidebar"] [role="radiogroup"] label {{
            color: {LAB_DARK} !important;
            opacity: 1 !important;
        }}
        .lab-hero {{
            background: linear-gradient(135deg, {LAB_BLUE} 0%, #0B4FD4 45%, {LAB_CYAN} 100%);
            color: white; padding: 1.25rem 1.5rem; border-radius: 22px;
            box-shadow: 0 12px 30px rgba(0, 51, 160, 0.18);
            margin-bottom: 1rem;
        }}
        .lab-hero, .lab-hero * {{ color: #FFFFFF !important; }}
        .lab-hero h1 {{ margin: 0; font-size: 2.1rem; font-weight: 850; }}
        .lab-hero p {{ margin: .35rem 0 0 0; opacity: .98; font-size: 1rem; }}
        .metric-card {{
            background: white; border: 1px solid #E6EEF8; border-radius: 18px; padding: 1rem;
            box-shadow: 0 8px 20px rgba(16, 42, 67, .06);
        }}
        .soft-card {{
            background: white; border: 1px solid #E6EEF8; border-radius: 18px; padding: 1rem;
            box-shadow: 0 8px 20px rgba(16, 42, 67, .05);
            color: {LAB_DARK};
        }}
        .soft-card, .soft-card * {{ color: {LAB_DARK} !important; opacity: 1 !important; }}
        .success-card {{
            background: linear-gradient(135deg, #ECFDF3 0%, #F7FFF9 100%);
            border: 1px solid #86EFAC;
            border-left: 7px solid #16A34A;
            border-radius: 18px;
            padding: 1rem 1.15rem;
            margin: .75rem 0 1rem 0;
            box-shadow: 0 10px 24px rgba(22, 163, 74, .10);
        }}
        .success-card, .success-card * {{ color: #14532D !important; opacity: 1 !important; }}
        .success-card-title {{ font-size: 1.08rem; font-weight: 850; margin-bottom: .25rem; }}
        .calendar-shell {{
            background: white; border: 1px solid #D9EAFB; border-radius: 20px;
            padding: .75rem; box-shadow: 0 8px 20px rgba(16, 42, 67, .05);
        }}
        .calendar-grid-week {{
            display: grid; grid-template-columns: repeat(7, minmax(120px, 1fr)); gap: .55rem;
        }}
        .calendar-grid-month {{
            display: grid; grid-template-columns: repeat(7, minmax(105px, 1fr)); gap: .35rem;
        }}
        .calendar-day {{
            min-height: 155px; border: 1px solid #E6EEF8; border-radius: 16px; padding: .55rem;
            background: #FBFDFF;
        }}
        .calendar-day-muted {{ background: #F3F7FB; opacity: .72; }}
        .calendar-today {{ border: 2px solid {LAB_CYAN}; background: #F2FBFF; }}
        .calendar-head {{
            font-weight: 800; color: {LAB_DARK}; font-size: .92rem; margin-bottom: .35rem;
        }}
        .calendar-date {{
            color: #627D98; font-size: .78rem; font-weight: 600;
        }}
        .calendar-pill {{
            display: block; border-radius: 10px; padding: .38rem .45rem; margin-top: .35rem;
            font-size: .76rem; line-height: 1.15rem; border-left: 4px solid #0033A0;
            background: #EEF7FF; color: #102A43;
        }}
        .calendar-pill-done {{ background: #ECFDF3; border-left-color: #2E7D32; }}
        .calendar-pill-cancelled {{ background: #F3F4F6; border-left-color: #9CA3AF; color: #6B7280; text-decoration: line-through; }}
        .calendar-pill-maintenance {{ background: #FFF7ED; border-left-color: #F97316; }}
        .calendar-pill-restricted {{ background: #FEFCE8; border-left-color: #CA8A04; }}
        .calendar-more {{ color: #627D98; font-size: .75rem; margin-top: .35rem; }}
        div[data-testid="stMetricValue"] {{ color: {LAB_BLUE}; }}
        div[data-testid="stMetricLabel"] {{ color: {LAB_DARK}; opacity: 1; }}
        .stButton>button {{ border-radius: 12px; border: 1px solid {LAB_CYAN}; }}
        .stDownloadButton>button {{ border-radius: 12px; }}
        .stButton>button,
        .stDownloadButton>button {{
            font-weight: 750 !important;
            color: {LAB_DARK} !important;
        }}
        div[data-testid="stFormSubmitButton"] button,
        .stFormSubmitButton button,
        button[data-testid="stBaseButton-primary"],
        div[data-testid="stBaseButton-primary"] button,
        .stButton>button[kind="primary"] {{
            background: {LAB_BLUE} !important;
            color: #FFFFFF !important;
            border: 1px solid {LAB_BLUE} !important;
            box-shadow: 0 8px 18px rgba(0, 51, 160, .18);
            font-weight: 800 !important;
        }}
        div[data-testid="stFormSubmitButton"] button *,
        .stFormSubmitButton button *,
        button[data-testid="stBaseButton-primary"] *,
        div[data-testid="stBaseButton-primary"] button *,
        .stButton>button[kind="primary"] * {{
            color: #FFFFFF !important;
            opacity: 1 !important;
        }}
        div[data-testid="stFormSubmitButton"] button:hover,
        .stFormSubmitButton button:hover,
        .stButton>button:hover {{
            border-color: {LAB_BLUE} !important;
            filter: brightness(.98);
        }}
        [data-testid="stSidebar"] .sidebar-brand {{
            font-weight: 900;
            font-size: 1.05rem;
            letter-spacing: 0;
            color: {LAB_BLUE} !important;
            margin: .25rem 0 .35rem 0;
        }}
        [data-testid="stSidebar"] .sidebar-nav-section {{
            margin: 1rem 0 .3rem 0;
            font-size: .72rem;
            font-weight: 900;
            text-transform: uppercase;
            letter-spacing: .08em;
            color: #627D98 !important;
        }}
        [data-testid="stSidebar"] .sidebar-user-card {{
            background: #FFFFFF;
            border: 1px solid #D9EAFB;
            border-radius: 14px;
            padding: .75rem .8rem;
            margin: .7rem 0 .55rem 0;
            box-shadow: 0 6px 16px rgba(16, 42, 67, .05);
        }}
        [data-testid="stSidebar"] .sidebar-user-name {{
            font-weight: 850;
            color: {LAB_DARK} !important;
            margin-bottom: .15rem;
        }}
        [data-testid="stSidebar"] .sidebar-user-meta,
        [data-testid="stSidebar"] .sidebar-user-role {{
            font-size: .78rem;
            color: #486581 !important;
            line-height: 1.2rem;
        }}
        [data-testid="stSidebar"] .stButton > button {{
            width: 100%;
            justify-content: flex-start;
            text-align: left;
            border-radius: 12px !important;
            padding: .5rem .7rem !important;
            margin-bottom: .18rem;
            border: 1px solid transparent !important;
            background: #FFFFFF !important;
            box-shadow: none !important;
        }}
        [data-testid="stSidebar"] .stButton > button:hover {{
            border-color: #B9D7F2 !important;
            background: #EEF7FF !important;
        }}
        [data-testid="stSidebar"] button[data-testid="stBaseButton-primary"],
        [data-testid="stSidebar"] .stButton > button[kind="primary"] {{
            background: {LAB_BLUE} !important;
            color: #FFFFFF !important;
            border-color: {LAB_BLUE} !important;
            box-shadow: 0 8px 18px rgba(0, 51, 160, .18) !important;
        }}
        [data-testid="stSidebar"] button[data-testid="stBaseButton-primary"] *,
        [data-testid="stSidebar"] .stButton > button[kind="primary"] * {{
            color: #FFFFFF !important;
            opacity: 1 !important;
        }}

        /* Limpeza do "chrome" nativo do Streamlit/Streamlit Cloud */
        #MainMenu {{
            visibility: hidden !important;
        }}
        footer {{
            visibility: hidden !important;
            display: none !important;
        }}
        header[data-testid="stHeader"],
        [data-testid="stHeader"],
        [data-testid="stToolbar"],
        [data-testid="stDecoration"],
        [data-testid="stStatusWidget"],
        [data-testid="manage-app-button"],
        [data-testid="stDeployButton"] {{
            visibility: hidden !important;
            display: none !important;
            height: 0 !important;
        }}

        /* O menu mobile existe no código, mas fica invisível no desktop. */
        .st-key-labcim_mobile_navigation {{
            display: none !important;
        }}

        @media (max-width: 768px) {{
            section[data-testid="stSidebar"] {{
                display: none !important;
            }}

            div[data-testid="collapsedControl"] {{
                display: none !important;
            }}

            .st-key-labcim_mobile_navigation {{
                display: block !important;
            }}

            .block-container {{
                padding-top: .75rem !important;
            }}
        }}

        /* Tema claro robusto para widgets do Streamlit/BaseWeb */
        [data-testid="stSelectbox"] div[data-baseweb="select"],
        [data-testid="stMultiSelect"] div[data-baseweb="select"],
        [data-testid="stSelectbox"] div[data-baseweb="select"] > div,
        [data-testid="stMultiSelect"] div[data-baseweb="select"] > div {{
            background-color: #FFFFFF !important;
            color: {LAB_DARK} !important;
            border-color: #CBD5E1 !important;
        }}
        [data-testid="stSelectbox"] div[data-baseweb="select"] *,
        [data-testid="stMultiSelect"] div[data-baseweb="select"] *,
        [data-testid="stSelectbox"] span,
        [data-testid="stMultiSelect"] span {{
            color: {LAB_DARK} !important;
            fill: {LAB_DARK} !important;
            opacity: 1 !important;
        }}
        .stTextInput input,
        .stTextArea textarea,
        .stNumberInput input,
        .stDateInput input,
        .stTimeInput input {{
            color: {LAB_DARK} !important;
            background-color: #FFFFFF !important;
            border-color: #CBD5E1 !important;
        }}
        [data-baseweb="popover"],
        [data-baseweb="popover"] > div,
        [data-baseweb="menu"],
        [role="listbox"],
        [data-baseweb="calendar"] {{
            background-color: #FFFFFF !important;
            color: {LAB_DARK} !important;
            border-color: #CBD5E1 !important;
        }}
        [data-baseweb="popover"] *,
        [data-baseweb="menu"] *,
        [role="listbox"] *,
        [data-baseweb="calendar"] *,
        div[role="option"] {{
            color: {LAB_DARK} !important;
            background-color: transparent !important;
            opacity: 1 !important;
        }}
        div[role="option"]:hover {{
            background-color: #EEF7FF !important;
        }}
        .stTabs [data-baseweb="tab-list"] {{
            gap: .45rem;
            border-bottom: 1px solid #D9E2EC;
            padding-bottom: .45rem;
            margin-bottom: .75rem;
        }}
        .stTabs [data-baseweb="tab"] {{
            color: {LAB_DARK} !important;
            background: #FFFFFF !important;
            border: 1px solid #CBD5E1 !important;
            border-radius: 999px !important;
            padding: .45rem .9rem !important;
            font-weight: 750 !important;
            box-shadow: 0 4px 10px rgba(16, 42, 67, .04);
        }}
        .stTabs [data-baseweb="tab"] * {{
            color: {LAB_DARK} !important;
            opacity: 1 !important;
        }}
        .stTabs [aria-selected="true"] {{
            background: {LAB_BLUE} !important;
            border-color: {LAB_BLUE} !important;
            color: #FFFFFF !important;
            font-weight: 850 !important;
            box-shadow: 0 8px 18px rgba(0, 51, 160, .18);
        }}
        .stTabs [aria-selected="true"] * {{ color: #FFFFFF !important; }}
        .stTabs [data-baseweb="tab-highlight"] {{ background: transparent !important; }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def _database_url() -> str | None:
    for key in ("DATABASE_URL", "database_url"):
        try:
            if hasattr(st, "secrets") and key in st.secrets:
                value = str(st.secrets[key]).strip()
                if value:
                    return value
        except Exception:
            pass
    try:
        if hasattr(st, "secrets") and "database" in st.secrets:
            database_secrets = st.secrets["database"]
            for key in ("url", "DATABASE_URL", "database_url"):
                if key in database_secrets:
                    value = str(database_secrets[key]).strip()
                    if value:
                        return value
    except Exception:
        pass
    return os.environ.get("DATABASE_URL") or None


def _database_fingerprint(database_url: str | None) -> str:
    if database_url:
        digest = hashlib.sha256(database_url.encode("utf-8")).hexdigest()[:16]
        return f"postgres:{digest}"
    return f"sqlite:{DB_PATH.as_posix()}"


def _base_xlsx_marker() -> str:
    if not BASE_XLSX.exists():
        return "missing"
    stat = BASE_XLSX.stat()
    return f"{stat.st_mtime_ns}:{stat.st_size}"


def _debug_perf_enabled() -> bool:
    value = os.environ.get("LABCIM_DEBUG_PERF")
    try:
        if hasattr(st, "secrets") and "LABCIM_DEBUG_PERF" in st.secrets:
            value = st.secrets["LABCIM_DEBUG_PERF"]
    except Exception:
        pass
    try:
        if hasattr(st, "secrets") and "debug" in st.secrets and "LABCIM_DEBUG_PERF" in st.secrets["debug"]:
            value = st.secrets["debug"]["LABCIM_DEBUG_PERF"]
    except Exception:
        pass
    return str(value or "").strip().lower() in {"1", "true", "yes", "sim", "on"}


class _PerfTimer:
    def __init__(self, label: str):
        self.label = label
        self.enabled = _debug_perf_enabled()
        self.start = 0.0

    def __enter__(self):
        if self.enabled:
            self.start = perf_counter()
        return self

    def __exit__(self, exc_type, exc, tb):
        if self.enabled:
            elapsed_ms = (perf_counter() - self.start) * 1000
            st.session_state.setdefault(PERF_EVENTS_KEY, []).append((self.label, elapsed_ms))
        return False


def perf_timer(label: str) -> _PerfTimer:
    return _PerfTimer(label)


def _reset_perf_events() -> None:
    if _debug_perf_enabled():
        st.session_state[PERF_EVENTS_KEY] = []


def _render_perf_debug() -> None:
    if not _debug_perf_enabled():
        return
    events = st.session_state.get(PERF_EVENTS_KEY, [])
    if not events:
        return
    with st.sidebar.expander("Performance debug", expanded=False):
        for label, elapsed_ms in events:
            st.caption(f"{label}: {elapsed_ms:.1f} ms")


@st.cache_resource(show_spinner=False)
def ensure_database_initialized(
    db_path: str,
    database_fingerprint: str,
    base_xlsx_marker: str,
    _database_url_value: str | None = None,
) -> dict[str, str]:
    conn = connect(Path(db_path), database_url=_database_url_value)
    try:
        init_db(conn)
        if BASE_XLSX.exists() and is_operational_database_empty(conn):
            import_base_xlsx(conn, BASE_XLSX)
        seed_default_pops(conn)
        return {
            "database": database_fingerprint,
            "base_xlsx": base_xlsx_marker,
            "initialized_at": datetime.now().isoformat(timespec="seconds"),
        }
    finally:
        conn.close()


def _connection_is_healthy(conn) -> bool:
    try:
        conn.execute("SELECT 1").fetchone()
        return True
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        return False


def _close_connection(conn) -> None:
    try:
        conn.close()
    except Exception:
        pass


def get_conn():
    database_url = _database_url()
    fingerprint = _database_fingerprint(database_url)
    with perf_timer("Inicialização do banco"):
        ensure_database_initialized(
            str(DB_PATH),
            fingerprint,
            _base_xlsx_marker(),
            _database_url_value=database_url,
        )

    conn = st.session_state.get(DB_CONNECTION_KEY)
    cached_fingerprint = st.session_state.get(DB_CONNECTION_FINGERPRINT_KEY)
    if conn is None or cached_fingerprint != fingerprint or not _connection_is_healthy(conn):
        if conn is not None:
            _close_connection(conn)
        with perf_timer("Conexão com banco"):
            conn = connect(DB_PATH, database_url=database_url)
        st.session_state[DB_CONNECTION_KEY] = conn
        st.session_state[DB_CONNECTION_FINGERPRINT_KEY] = fingerprint
    return conn


def clear_app_caches() -> None:
    try:
        st.cache_data.clear()
    except Exception:
        pass
    try:
        st.session_state[APP_CACHE_GENERATION_KEY] = int(st.session_state.get(APP_CACHE_GENERATION_KEY, 0) or 0) + 1
    except Exception:
        pass


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def _cached_table_counts(database_fingerprint: str, _database_url_value: str | None = None) -> dict[str, int]:
    conn = connect(DB_PATH, database_url=_database_url_value)
    try:
        return table_counts(conn)
    finally:
        conn.close()


def cached_table_counts(conn) -> dict[str, int]:
    database_url = _database_url()
    with perf_timer("table_counts"):
        return _cached_table_counts(
            _database_fingerprint(database_url),
            _database_url_value=database_url,
        )


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def _cached_reference_data(database_fingerprint: str, _database_url_value: str | None = None):
    conn = connect(DB_PATH, database_url=_database_url_value)
    try:
        equipment = query_df(conn, "SELECT * FROM equipment ORDER BY active DESC, equipment_code")
        users = query_df(conn, "SELECT * FROM users ORDER BY active DESC, full_name")
        projects = query_df(conn, "SELECT * FROM projects ORDER BY active DESC, project_name")
        return equipment, users, projects
    finally:
        conn.close()


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def _cached_supply_page_data(database_fingerprint: str, _database_url_value: str | None = None):
    conn = connect(DB_PATH, database_url=_database_url_value)
    try:
        supplies = query_df(conn, "SELECT * FROM supplies ORDER BY active DESC, supply_name")
        supply_lots = list_supply_lots(conn)
        users = query_df(conn, "SELECT * FROM users WHERE active=1 ORDER BY full_name")
        projects = query_df(conn, "SELECT * FROM projects WHERE active=1 ORDER BY project_name")
        equipment = query_df(conn, "SELECT * FROM equipment ORDER BY active DESC, equipment_code")
        return supplies, supply_lots, users, projects, equipment
    finally:
        conn.close()


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def _cached_project_services_data(
    database_fingerprint: str,
    active_only: bool = True,
    _database_url_value: str | None = None,
) -> pd.DataFrame:
    conn = connect(DB_PATH, database_url=_database_url_value)
    try:
        return list_project_services(conn, active_only=active_only)
    finally:
        conn.close()


def cached_project_services(*, active_only: bool = True) -> pd.DataFrame:
    database_url = _database_url()
    with perf_timer("Serviços/análises"):
        return _cached_project_services_data(
            _database_fingerprint(database_url),
            active_only=active_only,
            _database_url_value=database_url,
        )


def _storage_config_fingerprint() -> str:
    keys = (*R2_REQUIRED_KEYS, "R2_ACCOUNT_ID")
    values = [resolve_config_value(key) or "" for key in keys]
    digest = hashlib.sha256("\0".join(values).encode("utf-8")).hexdigest()[:16]
    return f"storage:{digest}"


@st.cache_resource(show_spinner=False)
def _cached_active_storage_backend(
    database_fingerprint: str,
    storage_fingerprint: str,
    _database_url_value: str | None = None,
):
    return get_active_storage_backend(database_url=_database_url_value)


def active_storage_backend():
    database_url = _database_url()
    return _cached_active_storage_backend(
        _database_fingerprint(database_url),
        _storage_config_fingerprint(),
        _database_url_value=database_url,
    )


@st.cache_resource(show_spinner=False)
def _cached_storage_backend_for_name(storage_backend: str, storage_fingerprint: str):
    return get_storage_backend_for_name(storage_backend)


def storage_backend_for_name(storage_backend: str):
    return _cached_storage_backend_for_name(storage_backend, _storage_config_fingerprint())


def _secret_value(*keys: str, default: str | None = None) -> str | None:
    for key in keys:
        try:
            if hasattr(st, "secrets") and key in st.secrets:
                return st.secrets[key]
        except Exception:
            pass
        try:
            if hasattr(st, "secrets") and "email" in st.secrets and key in st.secrets["email"]:
                return st.secrets["email"][key]
        except Exception:
            pass
        env_value = os.environ.get(key)
        if env_value:
            return env_value
    return default


def _email_config() -> dict[str, str | int | bool | None]:
    host = _secret_value("LABCIM_SMTP_HOST", "smtp_host")
    port = int(_secret_value("LABCIM_SMTP_PORT", "smtp_port", default="587") or "587")
    user = _secret_value("LABCIM_SMTP_USER", "smtp_user")
    password = _secret_value("LABCIM_SMTP_PASSWORD", "smtp_password")
    sender = _secret_value("LABCIM_SMTP_FROM", "smtp_from", default=user or "LabCim Manager <no-reply@labcim.local>")
    tls_raw = str(_secret_value("LABCIM_SMTP_TLS", "smtp_tls", default="true")).lower()
    return {
        "host": host,
        "port": port,
        "user": user,
        "password": password,
        "sender": sender,
        "use_tls": tls_raw not in {"0", "false", "nao", "não", "no"},
    }


def email_is_configured() -> bool:
    cfg = _email_config()
    return bool(cfg["host"] and cfg["user"] and cfg["password"])


def send_email(to_email: str, subject: str, body: str) -> tuple[bool, str]:
    cfg = _email_config()
    if not email_is_configured():
        return False, "SMTP não configurado."
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = str(cfg["sender"])
    msg["To"] = to_email
    msg.set_content(body)
    try:
        with smtplib.SMTP(str(cfg["host"]), int(cfg["port"]), timeout=20) as smtp:
            if cfg["use_tls"]:
                smtp.starttls()
            smtp.login(str(cfg["user"]), str(cfg["password"]))
            smtp.send_message(msg)
        return True, "E-mail enviado."
    except Exception as exc:
        return False, str(exc)


def _auth_debug_codes_enabled() -> bool:
    keys = ("LABCIM_AUTH_DEBUG_CODES", "auth_debug_codes")
    for key in keys:
        env_value = os.environ.get(key)
        if clean_input(env_value):
            return clean_input(env_value).lower() in {"1", "true", "sim", "yes", "y", "on"}
        try:
            if hasattr(st, "secrets") and key in st.secrets:
                return clean_input(st.secrets[key]).lower() in {"1", "true", "sim", "yes", "y", "on"}
        except Exception:
            pass
        try:
            if hasattr(st, "secrets") and "auth" in st.secrets and key in st.secrets["auth"]:
                return clean_input(st.secrets["auth"][key]).lower() in {"1", "true", "sim", "yes", "y", "on"}
        except Exception:
            pass
    return False


def _unique_emails(values) -> list[str]:
    seen = set()
    emails: list[str] = []
    for value in values:
        email = clean_input(value).lower()
        if email and "@" in email and email not in seen:
            emails.append(email)
            seen.add(email)
    return emails


def maintenance_notification_recipients(conn, equipment_id: int, include_future_users: bool = True) -> list[str]:
    manager_rows = conn.execute(
        """
        SELECT email
        FROM users
        WHERE active = 1
          AND email IS NOT NULL
          AND TRIM(email) != ''
          AND role IN ('admin', 'manager')
        """
    ).fetchall()
    specific_rows = conn.execute(
        """
        SELECT DISTINCT u.email
        FROM equipment e
        JOIN users u
          ON u.active = 1
         AND u.email IS NOT NULL
         AND TRIM(u.email) != ''
         AND (
             LOWER(u.full_name) = LOWER(COALESCE(e.responsible_name, ''))
             OR LOWER(u.full_name) = LOWER(COALESCE(e.technical_manager, ''))
         )
        WHERE e.id = ?
        """,
        [equipment_id],
    ).fetchall()
    emails = [r["email"] for r in manager_rows] + [r["email"] for r in specific_rows]
    if include_future_users:
        now_iso = datetime.now().isoformat(timespec="minutes")
        future_rows = conn.execute(
            """
            SELECT DISTINCT u.email
            FROM bookings b
            JOIN users u ON u.id = b.user_id
            WHERE b.equipment_id = ?
              AND b.status = 'scheduled'
              AND b.start_datetime >= ?
              AND u.active = 1
              AND u.email IS NOT NULL
              AND TRIM(u.email) != ''
            """,
            [equipment_id, now_iso],
        ).fetchall()
        emails.extend([r["email"] for r in future_rows])
    return _unique_emails(emails)


def notify_equipment_maintenance(
    conn,
    *,
    equipment_id: int,
    title: str,
    message: str,
    related_table: str,
    related_id: int | None = None,
    include_future_users: bool = True,
) -> tuple[int, int]:
    equipment = conn.execute("SELECT * FROM equipment WHERE id = ?", [equipment_id]).fetchone()
    if not equipment:
        return 0, 0
    recipients = maintenance_notification_recipients(conn, equipment_id, include_future_users=include_future_users)
    subject = f"LabCim Manager - {title}: {equipment['equipment_code']}"
    body = (
        f"Equipamento: {equipment['equipment_code']} — {equipment['equipment_name']}\n"
        f"Localização: {clean_value(equipment['location'])}\n"
        f"Responsável: {clean_value(equipment['responsible_name'])}\n\n"
        f"{message}\n\n"
        "Esta é uma notificação automática do LabCim Manager."
    )
    sent = 0
    total = 0
    for email in recipients:
        total += 1
        ok, error = send_email(email, subject, body)
        if ok:
            sent += 1
        log_notification(
            conn,
            event_type="equipment_maintenance",
            recipient_email=email,
            subject=subject,
            body=body,
            status="sent" if ok else "skipped_no_smtp" if not email_is_configured() else "error",
            error_message=None if ok else error,
            related_table=related_table,
            related_id=related_id or equipment_id,
        )
    return sent, total


def _hash_access_code(code: str) -> str:
    return hashlib.sha256(code.encode("utf-8")).hexdigest()


SESSION_ROLE_ALIASES = {
    "admin": "admin",
    "administrador": "admin",
    "manager": "manager",
    "gerente": "manager",
    "operator": "manager",
    "operador": "manager",
    "member": "member",
    "membro": "member",
    "usuario": "member",
    "usuário": "member",
    "user": "member",
}


def _validated_session_role(role: str | None) -> str | None:
    return SESSION_ROLE_ALIASES.get(clean_input(role).lower())


def _normalize_session_role(role: str | None) -> str:
    return _validated_session_role(role) or "member"


def is_authenticated() -> bool:
    return bool(st.session_state.get("auth_user"))


def current_user() -> dict:
    return st.session_state.get("auth_user", {})


def _set_authenticated_user(row) -> bool:
    role = _validated_session_role(row["role"])
    if role is None:
        logout()
        return False
    st.session_state["auth_user"] = {
        "id": int(row["user_id"] if "user_id" in row.keys() else row["id"]),
        "full_name": str(row["full_name"]),
        "email": str(row["user_email"] if "user_email" in row.keys() else row["email"]),
        "role": role,
    }
    st.session_state.pop("access_role", None)
    return True


def revalidate_authenticated_user(conn) -> dict | None:
    user = current_user()
    if not user:
        logout()
        return None

    row = None
    try:
        user_id = int(user.get("id"))
    except Exception:
        user_id = None
    if user_id is not None:
        row = conn.execute("SELECT * FROM users WHERE id = ? LIMIT 1", [user_id]).fetchone()

    if row is None:
        email = clean_input(user.get("email")).lower()
        if email:
            row = get_active_user_by_email(conn, email)

    if row is None or not truthy(row["active"]):
        logout()
        return None

    if not _set_authenticated_user(row):
        return None
    return current_user()


def logout() -> None:
    for key in ["auth_user", "access_role", "pending_login_email", "last_access_code"]:
        if key in st.session_state:
            del st.session_state[key]


def request_access_code(conn, email: str) -> tuple[bool, str]:
    user = get_active_user_by_email(conn, email)
    if not user:
        return False, "E-mail não encontrado entre usuários ativos do LabCim Manager."
    code = f"{py_secrets.randbelow(1_000_000):06d}"
    expires_at = (datetime.now() + timedelta(minutes=ACCESS_CODE_TTL_MINUTES)).isoformat(timespec="seconds")
    create_access_code_record(
        conn,
        user_id=int(user["id"]),
        email=str(user["email"]).strip().lower(),
        code_hash=_hash_access_code(code),
        expires_at=expires_at,
    )
    subject = "Código de acesso - LabCim Manager"
    body = (
        f"Olá, {user['full_name']}.\n\n"
        f"Seu código de acesso ao LabCim Manager é: {code}\n\n"
        f"Ele expira em {ACCESS_CODE_TTL_MINUTES} minutos.\n"
        "Se você não solicitou este acesso, ignore esta mensagem.\n\n"
        "LabCim Manager"
    )
    ok, msg = send_email(str(user["email"]), subject, body)
    debug_codes_enabled = _auth_debug_codes_enabled()
    log_notification(
        conn,
        event_type="access_code",
        recipient_email=str(user["email"]),
        subject=subject,
        body=(
            "Código de acesso enviado por e-mail."
            if ok
            else "Falha ao enviar código de acesso; código não exibido."
            if not debug_codes_enabled
            else "Código de acesso disponibilizado em modo debug de desenvolvimento."
        ),
        status="sent" if ok else "failed" if not debug_codes_enabled else "debug_code_available",
        error_message=None if ok else msg,
        related_table="users",
        related_id=int(user["id"]),
    )
    st.session_state["pending_login_email"] = str(user["email"]).strip().lower()
    if not ok:
        if debug_codes_enabled:
            st.session_state["last_access_code"] = code
            return True, "Modo debug de desenvolvimento: o código foi exibido na tela."
        conn.execute(
            "UPDATE access_codes SET used_at = CURRENT_TIMESTAMP WHERE user_id = ? AND code_hash = ? AND used_at IS NULL",
            [int(user["id"]), _hash_access_code(code)],
        )
        conn.commit()
        st.session_state.pop("last_access_code", None)
        return False, "Não foi possível enviar o código de acesso. Verifique a configuração de e-mail ou contate o administrador."
    st.session_state.pop("last_access_code", None)
    return True, f"Código enviado para {user['email']}."


def page_login(conn) -> None:
    hero()
    st.subheader("Acesso ao sistema")
    st.caption("Digite seu e-mail cadastrado. O sistema enviará uma senha volátil com validade curta.")

    with st.container(border=True):
        st.markdown("### Solicitar senha volátil")
        email = st.text_input(
            "E-mail cadastrado",
            value=clean_input(st.session_state.get("pending_login_email", "")),
            placeholder="seu.email@ufrn.br",
            key="login_email",
        ).strip().lower()
        if st.button("Enviar senha volátil", type="primary", key="send_access_code"):
            ok, msg = request_access_code(conn, email)
            (st.success if ok else st.error)(msg)

        if st.session_state.get("last_access_code"):
            st.warning(
                f"Modo debug de desenvolvimento: código de teste **{st.session_state['last_access_code']}**. "
                "Desative LABCIM_AUTH_DEBUG_CODES em produção."
            )

    with st.container(border=True):
        st.markdown("### Validar código")
        pending_email = clean_input(st.session_state.get("pending_login_email", "")).lower()
        if pending_email:
            st.info(f"Validando acesso para: **{pending_email}**")
        else:
            st.warning("Primeiro solicite uma senha volátil para o seu e-mail.")
        code = st.text_input(
            "Código recebido",
            max_chars=6,
            placeholder="000000",
            key="verify_code",
        ).strip()
        if st.button("Entrar", type="primary", key="verify_access_code"):
            if not pending_email:
                st.error("Solicite uma senha volátil antes de validar o código.")
                return
            if not code:
                st.error("Informe o código recebido por e-mail.")
                return
            ok, msg, row = verify_access_code_record(conn, email=pending_email, code_hash=_hash_access_code(code))
            if ok and row is not None:
                if _set_authenticated_user(row):
                    st.success("Acesso liberado.")
                    st.rerun()
                else:
                    st.error("Perfil de acesso inválido. Contate o administrador.")
            else:
                st.error(msg)

    if not email_is_configured():
        with st.expander("Configuração de e-mail SMTP para produção"):
            st.code(
                """
# .streamlit/secrets.toml
[email]
smtp_host = "smtp.seudominio.br"
smtp_port = 587
smtp_user = "usuario@seudominio.br"
smtp_password = "COLE_AQUI_A_SENHA_DE_APP"
smtp_from = "LabCim Manager <usuario@seudominio.br>"
smtp_tls = true
                """.strip(),
                language="toml",
            )


def hero():
    cols = st.columns([1, 4])
    with cols[0]:
        if LOGO_PATH.exists():
            st.image(str(LOGO_PATH), use_container_width=True)
        else:
            st.markdown("# 🔬")
    with cols[1]:
        st.markdown(
            f"""
            <div class="lab-hero">
                <h1>{APP_TITLE}</h1>
                <p>{APP_SUBTITLE}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )


def _initial_page_from_url() -> str:
    view = st.query_params.get("view", "")
    if view == "reserva":
        return "Reservas"
    if view == "manutencao":
        return "Manutenção"
    if view == "insumo":
        return "Insumos"
    if view == "pop":
        return "Reservas"
    return "Painel inicial"


def _allowed_sidebar_pages() -> list[str]:
    page_labels = list(PAGE_LABELS)
    if not can_view_users_directory():
        page_labels = [label for label in page_labels if label != "Usuários"]
    if not can_import_base():
        page_labels = [label for label in page_labels if label != "Importar base"]
    if not can_view_reports():
        page_labels = [label for label in page_labels if label != "Relatórios"]
    return page_labels


def _sidebar_button_key(page_label: str) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", page_label.lower(), flags=re.IGNORECASE).strip("_")
    return f"nav_{key or 'page'}"


def _sidebar_button_label(page_label: str) -> str:
    return f"{PAGE_ICONS.get(page_label, '•')} {page_label}"


def _sidebar_active_page(page_labels: list[str], default_page: str | None = None) -> str:
    url_page = default_page or _initial_page_from_url()
    previous_url_page = st.session_state.get(SIDEBAR_URL_PAGE_KEY)
    if previous_url_page != url_page and url_page in page_labels:
        st.session_state[SIDEBAR_PAGE_KEY] = url_page
    st.session_state[SIDEBAR_URL_PAGE_KEY] = url_page

    if SIDEBAR_PAGE_KEY not in st.session_state:
        st.session_state[SIDEBAR_PAGE_KEY] = url_page if url_page in page_labels else "Painel inicial"

    active_page = st.session_state.get(SIDEBAR_PAGE_KEY, "Painel inicial")
    if active_page not in page_labels:
        active_page = "Painel inicial"
        st.session_state[SIDEBAR_PAGE_KEY] = active_page
    return active_page


def _render_sidebar_user(user: dict) -> None:
    role = current_access_role()
    st.sidebar.markdown(
        f"""
        <div class="sidebar-user-card">
            <div class="sidebar-user-name">{escape(clean_value(user.get('full_name'), 'Usuário'))}</div>
            <div class="sidebar-user-meta">{escape(clean_value(user.get('email'), 'E-mail não informado'))}</div>
            <div class="sidebar-user-role">{escape(role_badge(role))}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def sidebar(default_page: str | None = None):
    if LOGO_PATH.exists():
        st.sidebar.image(str(LOGO_PATH), use_container_width=True)
    st.sidebar.markdown('<div class="sidebar-brand">LabCim Manager</div>', unsafe_allow_html=True)
    page_labels = _allowed_sidebar_pages()
    page = _sidebar_active_page(page_labels, default_page)

    for section, section_pages in NAVIGATION_SECTIONS:
        visible_pages = [label for label in section_pages if label in page_labels]
        if not visible_pages:
            continue
        st.sidebar.markdown(f'<div class="sidebar-nav-section">{escape(section)}</div>', unsafe_allow_html=True)
        for page_label in visible_pages:
            is_active = page_label == page
            clicked = st.sidebar.button(
                _sidebar_button_label(page_label),
                key=_sidebar_button_key(page_label),
                type="primary" if is_active else "secondary",
                use_container_width=True,
            )
            if clicked and not is_active:
                st.session_state[SIDEBAR_PAGE_KEY] = page_label
                st.rerun()

    st.sidebar.markdown("---")
    user = current_user()
    _render_sidebar_user(user)
    if st.sidebar.button("Sair", key="sidebar_logout"):
        logout()
        st.rerun()
    return page


def render_mobile_menu_navigation(selected_page: str) -> str:
    page_labels = _allowed_sidebar_pages()
    if selected_page not in page_labels:
        selected_page = "Painel inicial"
        st.session_state[SIDEBAR_PAGE_KEY] = selected_page

    with st.container(key="labcim_mobile_navigation"):
        mobile_page = st.selectbox(
            "☰ Menu",
            page_labels,
            index=page_labels.index(selected_page),
            format_func=lambda label: f"{PAGE_ICONS.get(label, '📄')} {label}",
            key="mobile_menu_navigation_page",
            help="Menu principal para uso em celular.",
        )

        if mobile_page != selected_page:
            st.session_state[SIDEBAR_PAGE_KEY] = mobile_page
            st.rerun()

        if st.button("Sair", key="mobile_menu_logout", use_container_width=True):
            logout()
            st.rerun()

    return mobile_page


def scroll_to_top_on_page_change(page: str) -> None:
    if st.session_state.get(SCROLL_TO_TOP_PAGE_KEY) == page:
        return

    st.session_state[SCROLL_TO_TOP_PAGE_KEY] = page
    components.html(
        """
        <script>
        const scrollToTop = () => {
            const target = window.parent || window;
            target.scrollTo({ top: 0, left: 0, behavior: "instant" });
        };
        requestAnimationFrame(scrollToTop);
        setTimeout(scrollToTop, 80);
        </script>
        """,
        height=0,
        width=0,
    )

def current_access_role() -> str:
    user = current_user()
    return _normalize_session_role(user.get("role")) if user else "member"


def is_admin() -> bool:
    return current_access_role() == "admin"


def can_manage_master_data() -> bool:
    return current_access_role() in {"manager", "admin"}


def can_edit_operational_data() -> bool:
    return current_access_role() in {"manager", "admin"}


def can_view_reports() -> bool:
    return current_access_role() in {"manager", "admin"}


def can_export_reports() -> bool:
    return current_access_role() in {"manager", "admin"}


def can_view_users_directory() -> bool:
    return current_access_role() in {"manager", "admin"}


def can_manage_inventory_adjustments() -> bool:
    return current_access_role() in {"manager", "admin"}


def can_export_inventory() -> bool:
    return current_access_role() in {"manager", "admin"}


def can_export_qr_bulk() -> bool:
    return current_access_role() in {"manager", "admin"}


def can_manage_users() -> bool:
    return is_admin()


def can_import_base() -> bool:
    return is_admin()


def admin_required_message(action: str = "alterar cadastros estruturais") -> None:
    st.info(f"Para {action}, use perfil Gerente ou Administrador.")


def status_badge(value: str) -> str:
    if is_blank(value):
        return "-"
    return STATUS_LABELS.get(value, value)


def equipment_status_badge(value: str) -> str:
    if is_blank(value):
        return "Apto"
    return EQUIPMENT_STATUS_LABELS.get(str(value), str(value))


def role_badge(value: str) -> str:
    if is_blank(value):
        return "-"
    if str(value).lower() in {"operator", "operador", "gerente"}:
        return "Gerente"
    return ROLE_LABELS.get(value, value)


def is_blank(value) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    text = str(value).strip()
    return text == "" or text.lower() in {"nan", "none", "nat"}


def clean_value(value, default: str = "-") -> str:
    if is_blank(value):
        return default
    return str(value).strip()


def clean_input(value) -> str:
    return "" if is_blank(value) else str(value).strip()


def truthy(value) -> bool:
    if is_blank(value):
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, Integral):
        return int(value) == 1
    if isinstance(value, Real):
        return float(value) == 1.0
    text = str(value).strip().lower()
    return text in {"1", "true", "sim", "yes", "y", "ativo"}


def yes_no(value) -> str:
    if is_blank(value):
        return "-"
    return "Sim" if truthy(value) else "Não"


def _format_datetime(value: str | None) -> str:
    if not value:
        return "-"
    try:
        text = str(value)
        parsed = datetime.fromisoformat(text)
        if len(text) == 10:
            return parsed.strftime("%d/%m/%Y")
        return parsed.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(value)


def _booking_status_reason_required(current_status: str, new_status: str, *, is_manager: bool) -> bool:
    current_status = clean_input(current_status)
    new_status = clean_input(new_status)
    if current_status == new_status:
        return False
    if new_status in {"cancelled", "no_show"}:
        return True
    return is_manager and current_status in BOOKING_FINAL_STATUSES and new_status != current_status


def render_booking_status_history(conn, booking_id: int) -> None:
    history = list_booking_status_history(conn, int(booking_id))
    if history.empty:
        st.info("Sem histórico registrado para esta reserva.")
        return

    display = history.copy()
    display["changed_at"] = display["changed_at"].map(_format_datetime)
    display["previous_status"] = display["previous_status"].map(lambda value: "Criação" if is_blank(value) else status_badge(str(value)))
    display["new_status"] = display["new_status"].map(lambda value: status_badge(str(value)))
    display["changed_by_name"] = display["changed_by_name"].map(lambda value: clean_value(value, "Usuário não informado"))
    display["reason"] = display["reason"].map(lambda value: clean_value(value, "-"))
    display["source"] = display["source"].map(lambda value: clean_value(value, "-"))
    display = display.rename(
        columns={
            "changed_at": "Data/hora",
            "previous_status": "Status anterior",
            "new_status": "Novo status",
            "changed_by_name": "Usuário",
            "reason": "Observação/justificativa",
            "source": "Origem",
        }
    )
    st.dataframe(
        display[["Data/hora", "Status anterior", "Novo status", "Usuário", "Observação/justificativa", "Origem"]],
        use_container_width=True,
        hide_index=True,
    )


def _date_input_value(value, default=None):
    if is_blank(value):
        return default
    try:
        return datetime.fromisoformat(str(value)).date()
    except Exception:
        try:
            return pd.to_datetime(value).date()
        except Exception:
            return default


def _resolve_local_doc(path_value) -> Path | None:
    path_text = clean_input(path_value)
    if not path_text:
        return None
    path = Path(path_text)
    if not path.is_absolute():
        path = Path.cwd() / path
    try:
        path = path.resolve()
        cwd = Path.cwd().resolve()
        if cwd not in path.parents and path != cwd:
            return None
    except Exception:
        return None
    return path if path.exists() and path.is_file() else None


def pop_download_button(equipment_row, key_prefix: str = "pop") -> None:
    raw_path = clean_input(equipment_row.get("pop_path"))
    if raw_path.lower().startswith(("http://", "https://")):
        st.link_button(f"📄 Abrir {clean_value(equipment_row.get('pop_title'), 'POP do equipamento')}", raw_path)
        return
    doc_path = _resolve_local_doc(raw_path)
    title = clean_value(equipment_row.get("pop_title"), "POP do equipamento")
    version = clean_value(equipment_row.get("pop_version"), "sem versão")
    responsible = clean_value(equipment_row.get("pop_responsible"))
    if doc_path:
        st.download_button(
            f"📄 Baixar {title}",
            data=doc_path.read_bytes(),
            file_name=doc_path.name,
            mime="application/pdf",
            key=f"{key_prefix}_{clean_value(equipment_row.get('equipment_code'), 'eq')}",
            help=f"{title} · {version} · Responsável: {responsible}",
        )
    else:
        st.info("Nenhum POP/documento operacional cadastrado para este equipamento.")


def _display_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in ["status"]:
        if c in out.columns:
            out[c] = out[c].map(status_badge).fillna(out[c])
    for c in ["operational_status"]:
        if c in out.columns:
            out[c] = out[c].map(equipment_status_badge).fillna(out[c])
    for c in ["role"]:
        if c in out.columns:
            out[c] = out[c].map(role_badge).fillna(out[c])
    for c in ["active", "is_active", "association_active", "requires_operator", "training_completed", "capacity_enforced", "blocks_booking"]:
        if c in out.columns:
            out[c] = out[c].map(yes_no)
    for c in [
        "start_datetime",
        "end_datetime",
        "created_at",
        "updated_at",
        "changed_at",
        "inactive_at",
        "occurrence_datetime",
        "expiration_date",
        "supply_lot_expiration_date",
        "received_date",
        "movement_date",
        "planned_date",
        "planned_end_date",
        "performed_date",
        "next_date",
        "conclusion_date",
        "start_date",
        "end_date",
        "requested_date",
        "expected_date",
        "completed_date",
    ]:
        if c in out.columns:
            out[c] = out[c].map(_format_datetime)
    return out.rename(columns=COLUMN_LABELS)


def load_reference_data(conn):
    database_url = _database_url()
    with perf_timer("Listas de referência"):
        equipment, users, projects = _cached_reference_data(
            _database_fingerprint(database_url),
            _database_url_value=database_url,
        )
    operators = users[users["role"].isin(["manager", "operator", "admin"])] if not users.empty else users
    return equipment, users, projects, operators


def page_dashboard(conn):
    hero()
    if current_access_role() == "member":
        st.subheader("Painel operacional")
        st.caption("Use o menu lateral para reservar equipamentos, consultar documentos, reportar problemas e registrar consumo de insumos.")

        cards = [
            ("Reservas", "Criar e acompanhar suas reservas de equipamentos."),
            ("Equipamentos e documentos", "Consultar equipamentos, POPs e documentos operacionais disponíveis."),
            ("Manutenção", "Reportar problemas observados nos equipamentos."),
            ("Insumos", "Consultar itens, documentos e registrar saída/consumo."),
        ]
        cols = st.columns(2)
        for idx, (title, description) in enumerate(cards):
            with cols[idx % 2]:
                st.markdown(
                    f"""
                    <div class="soft-card">
                    <b>{title}</b><br>
                    {description}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
        st.info("Relatórios, exportações e visão administrativa ficam disponíveis para Gerente ou Administrador.")
        return

    st.subheader("Visão geral da base LabCim")
    counts = cached_table_counts(conn)
    metrics = [
        ("Equipamentos", counts["equipment"]),
        ("Usuários", counts["users"]),
        ("Projetos", counts["projects"]),
        ("Reservas", counts["bookings"]),
        ("Preventivas", counts.get("maintenance_preventive", 0)),
        ("Corretivas", counts.get("maintenance_corrective", 0)),
        ("Insumos", counts.get("supplies", 0)),
    ]
    cols = st.columns(len(metrics))
    for col, (label, value) in zip(cols, metrics):
        with col:
            st.metric(label, value)

    st.markdown("---")
    equipment = query_df(conn, "SELECT lab_unit, active, COUNT(*) AS total FROM equipment GROUP BY lab_unit, active")
    users = query_df(conn, "SELECT lab_unit, role, COUNT(*) AS total FROM users GROUP BY lab_unit, role")
    bookings = query_df(
        conn,
        """
        SELECT b.id, e.equipment_code, e.equipment_name, u.full_name, p.project_name,
               ps.title AS service_title,
               b.start_datetime, b.end_datetime, b.status, b.sample_count
        FROM bookings b
        JOIN equipment e ON e.id=b.equipment_id
        JOIN users u ON u.id=b.user_id
        LEFT JOIN projects p ON p.id=b.project_id
        LEFT JOIN project_services ps ON ps.id=b.service_id
        ORDER BY b.start_datetime DESC
        LIMIT 12
        """,
    )

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("#### Equipamentos por unidade")
        if not equipment.empty:
            fig = px.bar(
                equipment,
                x="lab_unit",
                y="total",
                color="active",
                barmode="group",
                labels={"lab_unit": "Unidade", "total": "Quantidade", "active": "Ativo"},
                color_discrete_sequence=[LAB_CYAN, LAB_BLUE],
            )
            fig.update_layout(height=360, margin=dict(l=20, r=20, t=20, b=20))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Nenhum equipamento cadastrado.")
    with c2:
        st.markdown("#### Usuários por perfil")
        if not users.empty:
            fig = px.bar(
                users,
                x="role",
                y="total",
                color="lab_unit",
                barmode="group",
                labels={"role": "Perfil", "total": "Quantidade", "lab_unit": "Unidade"},
                color_discrete_sequence=[LAB_BLUE, LAB_CYAN, "#6BAED6"],
            )
            fig.update_layout(height=360, margin=dict(l=20, r=20, t=20, b=20))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Nenhum usuário cadastrado.")

    st.markdown("#### Últimas reservas")
    if bookings.empty:
        st.info("Ainda não há reservas registradas.")
    else:
        st.dataframe(_display_df(bookings), use_container_width=True, hide_index=True)

    supplies = query_df(conn, "SELECT * FROM supplies WHERE active=1")
    if not supplies.empty:
        supplies["alerta"] = supplies.apply(_supply_alert_status, axis=1)
        critical = supplies[supplies["alerta"].isin(["Estoque baixo", "Vencido", "Vence em até 60 dias"])]
        if not critical.empty:
            st.markdown("#### Alertas de insumos")
            st.dataframe(
                _display_df(critical[["alerta", "supply_name", "current_quantity", "unit", "minimum_quantity", "expiration_date", "location"]]),
                use_container_width=True,
                hide_index=True,
            )


def _select_index_by_code(equipment: pd.DataFrame, code: str | None) -> int:
    if not code or equipment.empty:
        return 0
    codes = [str(c).upper() for c in equipment["equipment_code"].tolist()]
    code = str(code).upper()
    return codes.index(code) if code in codes else 0


def _equipment_row_by_code(equipment: pd.DataFrame, code: str | None):
    if not code or equipment.empty or "equipment_code" not in equipment.columns:
        return None
    code_text = str(code).strip().upper()
    matches = equipment[equipment["equipment_code"].astype(str).str.strip().str.upper() == code_text]
    if matches.empty:
        return None
    return matches.iloc[0]


def _reservable_equipment(equipment: pd.DataFrame) -> pd.DataFrame:
    if equipment.empty:
        return equipment.copy()
    status = equipment.get("operational_status", pd.Series("", index=equipment.index)).fillna("").astype(str).str.strip().str.lower()
    return equipment[(equipment["active"] == 1) & (status != "inactive")].copy()


def _user_options(users: pd.DataFrame) -> list[str]:
    return users.apply(lambda r: f"{clean_value(r.get('full_name'))} ({clean_value(r.get('department'), 'sem vínculo')})", axis=1).tolist()


def _project_options(projects: pd.DataFrame) -> list[str]:
    if projects.empty:
        return ["Sem projeto específico"]
    return ["Sem projeto específico"] + projects.apply(
        lambda r: f"{clean_value(r.get('project_code'), 'Sem código')} — {clean_value(r.get('project_name'))}",
        axis=1,
    ).tolist()


def _optional_user_index(users: pd.DataFrame, user_id) -> int:
    if users.empty or is_blank(user_id):
        return 0
    ids = users["id"].astype(int).tolist()
    try:
        user_id = int(user_id)
    except Exception:
        return 0
    return ids.index(user_id) + 1 if user_id in ids else 0


def _project_services_for_project(project_id: int | None, *, active_only: bool = True) -> pd.DataFrame:
    if project_id is None:
        return pd.DataFrame()
    services = cached_project_services(active_only=active_only)
    if services.empty:
        return services
    return services[services["project_id"].astype(int) == int(project_id)].copy()


def _service_options(services: pd.DataFrame, *, include_project: bool = False) -> list[str]:
    if services.empty:
        return ["Sem serviço/análise específico"]
    if include_project:
        return ["Sem serviço/análise específico"] + services.apply(
            lambda r: (
                f"{clean_value(r.get('project_code'), 'Sem código')} — "
                f"{clean_value(r.get('project_name'))} · "
                f"{clean_value(r.get('service_code'), 'Sem código')} — {clean_value(r.get('title'))}"
            ),
            axis=1,
        ).tolist()
    return ["Sem serviço/análise específico"] + services.apply(
        lambda r: f"{clean_value(r.get('service_code'), 'Sem código')} — {clean_value(r.get('title'))}",
        axis=1,
    ).tolist()


def _service_id_from_label(services: pd.DataFrame, label: str, *, include_project: bool = False) -> int | None:
    options = _service_options(services, include_project=include_project)
    if label == "Sem serviço/análise específico" or services.empty:
        return None
    return int(services.iloc[options.index(label) - 1]["id"])


def _operator_options(operators: pd.DataFrame, *, placeholder: str = "Selecionar depois") -> list[str]:
    if operators.empty:
        return [placeholder]
    return [placeholder] + operators.apply(lambda r: f"{clean_value(r.get('full_name'))} ({role_badge(clean_value(r.get('role'), 'member'))})", axis=1).tolist()


def _booking_query_for_equipment(conn, equipment_id: int, start_date: date, end_date: date, include_cancelled: bool = True) -> pd.DataFrame:
    start_iso = datetime.combine(start_date, time.min).isoformat(timespec="minutes")
    end_iso = datetime.combine(end_date + timedelta(days=1), time.min).isoformat(timespec="minutes")
    sql = """
        SELECT b.id, e.equipment_code, e.equipment_name, u.full_name AS solicitante,
               p.project_name, ps.service_code, ps.title AS service_title,
               op.full_name AS operador, perf.full_name AS executante,
               b.start_datetime, b.end_datetime,
               b.sample_count, b.purpose, b.status
        FROM bookings b
        JOIN equipment e ON e.id=b.equipment_id
        JOIN users u ON u.id=b.user_id
        LEFT JOIN users op ON op.id=b.operator_id
        LEFT JOIN users perf ON perf.id=b.performed_by_id
        LEFT JOIN projects p ON p.id=b.project_id
        LEFT JOIN project_services ps ON ps.id=b.service_id
        WHERE b.equipment_id = ?
          AND b.start_datetime >= ?
          AND b.start_datetime < ?
    """
    params = [equipment_id, start_iso, end_iso]
    if not include_cancelled:
        sql += " AND b.status != 'cancelled'"
    sql += " ORDER BY b.start_datetime"
    return query_df(conn, sql, params)


def _calendar_events_for_equipment(conn, equipment_id: int, start_date: date, end_date: date, include_cancelled: bool = False) -> pd.DataFrame:
    """Eventos de reserva e manutenção que cruzam o intervalo informado."""
    start_iso = datetime.combine(start_date, time.min).isoformat(timespec="minutes")
    end_iso = datetime.combine(end_date + timedelta(days=1), time.min).isoformat(timespec="minutes")

    bookings_sql = """
        SELECT 'booking' AS event_type, b.id, e.equipment_code, e.equipment_name,
               u.full_name AS solicitante, op.full_name AS operador, perf.full_name AS executante,
               p.project_name, ps.service_code, ps.title AS service_title,
               b.start_datetime AS start_datetime, b.end_datetime AS end_datetime,
               b.sample_count, b.purpose, b.status
        FROM bookings b
        JOIN equipment e ON e.id=b.equipment_id
        JOIN users u ON u.id=b.user_id
        LEFT JOIN users op ON op.id=b.operator_id
        LEFT JOIN users perf ON perf.id=b.performed_by_id
        LEFT JOIN projects p ON p.id=b.project_id
        LEFT JOIN project_services ps ON ps.id=b.service_id
        WHERE b.equipment_id = ?
          AND b.start_datetime < ?
          AND b.end_datetime > ?
    """
    params = [equipment_id, end_iso, start_iso]
    if not include_cancelled:
        bookings_sql += " AND b.status != 'cancelled'"
    booking_events = query_df(conn, bookings_sql + " ORDER BY b.start_datetime", params)

    maintenance_events = query_df(
        conn,
        """
        SELECT 'maintenance' AS event_type, mp.id, e.equipment_code, e.equipment_name,
               NULL AS solicitante, NULL AS operador, NULL AS executante,
               NULL AS project_name,
               NULL AS service_code,
               NULL AS service_title,
               mp.planned_date AS start_datetime,
               COALESCE(mp.planned_end_date, mp.planned_date) AS end_datetime,
               NULL AS sample_count,
               mp.description AS purpose,
               COALESCE(mp.status, 'pendente') AS status,
               mp.activity_type,
               mp.blocks_booking
        FROM maintenance_preventive mp
        JOIN equipment e ON e.id=mp.equipment_id
        WHERE mp.equipment_id = ?
          AND mp.planned_date IS NOT NULL
          AND mp.planned_date <= ?
          AND COALESCE(mp.planned_end_date, mp.planned_date) >= ?
        ORDER BY mp.planned_date
        """,
        [equipment_id, end_date.isoformat(), start_date.isoformat()],
    )
    if booking_events.empty:
        return maintenance_events
    if maintenance_events.empty:
        return booking_events
    return pd.concat([booking_events, maintenance_events], ignore_index=True, sort=False)


def _coerce_event_datetime(value, end_of_day: bool = False) -> datetime:
    if not value:
        return datetime.max if end_of_day else datetime.min
    text = str(value)
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        parsed = datetime.fromisoformat(text + ("T23:59:59" if end_of_day else "T00:00:00"))
    if end_of_day and len(text) == 10:
        return datetime.combine(parsed.date(), time.max)
    return parsed


def _event_overlaps_day(row: pd.Series, day: date) -> bool:
    start_dt = _coerce_event_datetime(row.get("start_datetime"))
    end_dt = _coerce_event_datetime(row.get("end_datetime"), end_of_day=True)
    day_start = datetime.combine(day, time.min)
    day_end = datetime.combine(day, time.max)
    return start_dt <= day_end and end_dt >= day_start


def _event_time_label(row: pd.Series, day: date) -> str:
    if row.get("event_type") == "maintenance":
        start_dt = _coerce_event_datetime(row.get("start_datetime"))
        end_dt = _coerce_event_datetime(row.get("end_datetime"), end_of_day=True)
        if start_dt.date() == end_dt.date():
            return "manutenção"
        return f"{start_dt.strftime('%d/%m')}–{end_dt.strftime('%d/%m')}"
    start_dt = _coerce_event_datetime(row.get("start_datetime"))
    end_dt = _coerce_event_datetime(row.get("end_datetime"), end_of_day=True)
    if start_dt.date() == day and end_dt.date() == day:
        return f"{start_dt.strftime('%H:%M')}–{end_dt.strftime('%H:%M')}"
    if start_dt.date() == day:
        return f"{start_dt.strftime('%H:%M')}→"
    if end_dt.date() == day:
        return f"→{end_dt.strftime('%H:%M')}"
    return "continuação"


def _event_css_class(row: pd.Series, selected_eq: pd.Series) -> str:
    if row.get("event_type") == "maintenance":
        return "calendar-pill calendar-pill-maintenance"
    if str(row.get("status")) == "done":
        return "calendar-pill calendar-pill-done"
    if str(row.get("status")) == "cancelled":
        return "calendar-pill calendar-pill-cancelled"
    if selected_eq.get("operational_status") == "restricted":
        return "calendar-pill calendar-pill-restricted"
    return "calendar-pill"


def _render_event_pill(row: pd.Series, day: date, selected_eq: pd.Series) -> str:
    css = _event_css_class(row, selected_eq)
    time_label = escape(_event_time_label(row, day))
    if row.get("event_type") == "maintenance":
        title = f"{clean_value(row.get('activity_type'), 'Manutenção')} #{int(row.get('id')) if pd.notna(row.get('id')) else ''}"
        desc = clean_value(row.get("purpose"))
        status = clean_value(row.get("status"))
        return (
            f"<span class='{css}'><b>{time_label}</b><br>"
            f"{escape(str(title))}<br><small>{escape(str(desc))} · {escape(str(status))}</small></span>"
        )
    title = f"Reserva #{int(row.get('id')) if pd.notna(row.get('id')) else ''}"
    who = clean_value(row.get("solicitante"), "Solicitante não informado")
    samples = ""
    if pd.notna(row.get("sample_count")) and row.get("sample_count") not in ["", None]:
        try:
            samples = f" · {int(row.get('sample_count'))} am."
        except Exception:
            samples = f" · {row.get('sample_count')} am."
    status = status_badge(str(row.get("status")))
    return (
        f"<span class='{css}'><b>{time_label}</b><br>"
        f"{escape(title)} · {escape(str(who))}<br><small>{escape(status)}{escape(samples)}</small></span>"
    )


def _week_start(day: date) -> date:
    return day - timedelta(days=day.weekday())


def _render_week_calendar(events: pd.DataFrame, selected_eq: pd.Series, week_start: date) -> str:
    day_names = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    days = [week_start + timedelta(days=i) for i in range(7)]
    html = ["<div class='calendar-shell'><div class='calendar-grid-week'>"]
    today = date.today()
    for label, day in zip(day_names, days):
        cls = "calendar-day calendar-today" if day == today else "calendar-day"
        html.append(
            f"<div class='{cls}'><div class='calendar-head'>{label}<br>"
            f"<span class='calendar-date'>{day.strftime('%d/%m/%Y')}</span></div>"
        )
        if events.empty:
            html.append("<div class='calendar-more'>Livre</div>")
        else:
            day_events = events[events.apply(lambda r: _event_overlaps_day(r, day), axis=1)].copy()
            if day_events.empty:
                html.append("<div class='calendar-more'>Livre</div>")
            else:
                day_events["_sort"] = day_events["start_datetime"].astype(str)
                day_events = day_events.sort_values("_sort")
                for _, row in day_events.head(4).iterrows():
                    html.append(_render_event_pill(row, day, selected_eq))
                if len(day_events) > 4:
                    html.append(f"<div class='calendar-more'>+{len(day_events) - 4} evento(s)</div>")
        html.append("</div>")
    html.append("</div></div>")
    return "".join(html)


def _month_grid_days(year: int, month: int) -> list[date]:
    first = date(year, month, 1)
    start = _week_start(first)
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    last = next_month - timedelta(days=1)
    end = last + timedelta(days=(6 - last.weekday()))
    days = []
    d = start
    while d <= end:
        days.append(d)
        d += timedelta(days=1)
    return days


def _render_month_calendar(events: pd.DataFrame, selected_eq: pd.Series, anchor_day: date) -> str:
    days = _month_grid_days(anchor_day.year, anchor_day.month)
    day_names = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    html = ["<div class='calendar-shell'>"]
    html.append("<div class='calendar-grid-month'>")
    for label in day_names:
        html.append(f"<div class='calendar-head' style='text-align:center'>{label}</div>")
    today = date.today()
    for day in days:
        cls = "calendar-day"
        if day.month != anchor_day.month:
            cls += " calendar-day-muted"
        if day == today:
            cls += " calendar-today"
        html.append(f"<div class='{cls}' style='min-height:112px'><div class='calendar-head'>{day.day}</div>")
        if not events.empty:
            day_events = events[events.apply(lambda r: _event_overlaps_day(r, day), axis=1)].copy()
            if not day_events.empty:
                for _, row in day_events.head(2).iterrows():
                    html.append(_render_event_pill(row, day, selected_eq))
                if len(day_events) > 2:
                    html.append(f"<div class='calendar-more'>+{len(day_events) - 2}</div>")
        html.append("</div>")
    html.append("</div></div>")
    return "".join(html)


def page_reservas(conn):
    hero()
    st.subheader("Agenda funcional dos equipamentos")
    st.caption("Consulte a agenda por equipamento, crie reservas e acompanhe conflitos, status e documentação operacional.")

    equipment, users, projects, operators = load_reference_data(conn)
    if equipment.empty or users.empty:
        st.warning("Cadastre/importe equipamentos e usuários antes de criar reservas.")
        return

    active_equipment = _reservable_equipment(equipment)
    active_users = users[users["active"] == 1].copy()
    active_projects = projects[projects["active"] == 1].copy()
    booking_user = current_user()
    current_user_id = _current_user_id()
    can_choose_requester = can_manage_master_data()
    if active_equipment.empty or active_users.empty:
        st.warning("É necessário ter ao menos um equipamento disponível para reserva e um usuário ativo.")
        return

    qr_eq = st.query_params.get("eq", None)
    qr_view = clean_input(st.query_params.get("view")).lower()
    qr_issue = None
    qr_selection_code = qr_eq
    if qr_eq:
        qr_row = _equipment_row_by_code(equipment, qr_eq)
        if qr_row is None:
            qr_issue = "QR Code aponta para equipamento não encontrado. Selecione o equipamento manualmente."
            qr_selection_code = None
        else:
            qr_status = clean_input(qr_row.get("operational_status")).lower()
            if not truthy(qr_row.get("active")) or qr_status == "inactive":
                qr_issue = (
                    "QR Code aponta para equipamento inativo. "
                    f"{clean_value(qr_row.get('equipment_code'))} — {clean_value(qr_row.get('equipment_name'))} não está disponível para nova reserva."
                )
                qr_selection_code = None
    if qr_issue:
        st.warning(qr_issue)

    eq_labels = _equipment_options(active_equipment)
    manual_placeholder = "Selecione um equipamento"
    selected_index = _select_index_by_code(active_equipment, qr_selection_code)

    top1, top2 = st.columns([2, 1])
    with top1:
        if qr_issue:
            eq_options = [manual_placeholder] + eq_labels
            eq_label = st.selectbox("Equipamento", eq_options, index=0, key="booking_eq_manual_after_qr_issue")
            if eq_label == manual_placeholder:
                st.info("Selecione manualmente um equipamento disponível para continuar.")
                return
        else:
            eq_label = st.selectbox("Equipamento", eq_labels, index=selected_index, key="booking_eq_main")
    equipment_id = _equipment_id_from_label(active_equipment, eq_label)
    selected_eq = active_equipment[active_equipment["id"] == equipment_id].iloc[0]
    operator_required = truthy(selected_eq.get("requires_operator"))
    with top2:
        st.metric("Status do equipamento", equipment_status_badge(selected_eq.get("operational_status") or "available"))

    cap_text = "-"
    if pd.notna(selected_eq.get("max_sample_capacity")) and selected_eq.get("max_sample_capacity"):
        cap_text = f"{int(selected_eq.get('max_sample_capacity'))} {clean_value(selected_eq.get('capacity_unit'), 'amostras')}"
    unavailable = clean_input(selected_eq.get("unavailable_functions"))
    st.markdown(
        f"""
        <div class="soft-card">
        <b>{clean_value(selected_eq.get('equipment_code'))} — {clean_value(selected_eq.get('equipment_name'))}</b><br>
        Unidade: {clean_value(selected_eq.get('lab_unit'))} · Local: {clean_value(selected_eq.get('location'))} ·
        Responsável: {clean_value(selected_eq.get('responsible_name'))} · Requer operador: {yes_no(selected_eq.get('requires_operator', 0))}<br>
        Capacidade usual: {cap_text} · Gestor técnico: {clean_value(selected_eq.get('technical_manager'))}
        </div>
        """,
        unsafe_allow_html=True,
    )
    if selected_eq.get("operational_status") == "restricted":
        st.warning(f"Equipamento em uso restrito. Funcionalidades indisponíveis: {unavailable or 'não especificadas'}.")
    elif selected_eq.get("operational_status") == "maintenance":
        st.error("Equipamento marcado como em manutenção. Novas reservas serão bloqueadas.")

    with st.container(border=True):
        st.markdown("#### Documentação operacional")
        if qr_view == "pop":
            st.info("QR de documentação operacional aberto. Consulte abaixo o POP e os documentos cadastrados deste equipamento.")
        d1, d2 = st.columns([1, 2])
        with d1:
            pop_download_button(selected_eq, key_prefix="booking_pop")
        with d2:
            st.caption(
                f"{clean_value(selected_eq.get('pop_title'), 'Sem POP cadastrado')} · "
                f"Versão: {clean_value(selected_eq.get('pop_version'))} · "
                f"Responsável: {clean_value(selected_eq.get('pop_responsible'))}"
            )
            if not is_blank(selected_eq.get("document_notes")):
                st.caption(clean_value(selected_eq.get("document_notes")))
        if _equipment_document_rows(conn, equipment_id):
            st.markdown("##### Documentos cadastrados")
            render_equipment_document_downloads(
                conn,
                selected_eq,
                key_prefix=f"booking_equipment_documents_{equipment_id}",
                show_empty=False,
                can_deactivate=False,
            )

    confirmation = st.session_state.get("booking_confirmation")
    if confirmation:
        st.markdown(
            f"""
            <div class="success-card">
                <div class="success-card-title">✅ Reserva confirmada</div>
                <div>
                    <b>{escape(clean_value(confirmation.get("equipment")))}</b><br>
                    Solicitante: {escape(clean_value(confirmation.get("user")))} ·
                    Período: {escape(clean_value(confirmation.get("period")))} ·
                    Amostras: {escape(clean_value(confirmation.get("samples")))}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Ocultar confirmação", key="dismiss_booking_confirmation"):
            st.session_state.pop("booking_confirmation", None)
            st.rerun()

    st.markdown("#### Escolha uma ação")
    booking_section = st.radio(
        "Escolha uma ação",
        [
            "Calendário semanal",
            "Calendário mensal",
            "Agenda linear",
            "Nova reserva",
            "Gerenciar reservas",
        ],
        horizontal=True,
        key="booking_section",
        label_visibility="collapsed",
    )

    if booking_section == "Calendário semanal":
        st.markdown("### Calendário semanal")
        c1, c2 = st.columns([1, 2])
        with c1:
            selected_week_day = st.date_input("Semana de referência", value=date.today(), format="DD/MM/YYYY", key="calendar_week_day")
        with c2:
            include_cancelled_week = st.checkbox("Mostrar reservas canceladas", value=False, key="calendar_week_cancelled")
        current_week_start = _week_start(selected_week_day)
        current_week_end = current_week_start + timedelta(days=6)
        week_events = _calendar_events_for_equipment(
            conn,
            equipment_id,
            current_week_start,
            current_week_end,
            include_cancelled=include_cancelled_week,
        )
        st.caption(f"Semana de {current_week_start.strftime('%d/%m/%Y')} a {current_week_end.strftime('%d/%m/%Y')}.")
        st.markdown(_render_week_calendar(week_events, selected_eq, current_week_start), unsafe_allow_html=True)
        if not week_events.empty:
            with st.expander("Ver eventos da semana em tabela"):
                st.dataframe(_display_df(week_events.drop(columns=[c for c in ["activity_type", "blocks_booking"] if c in week_events.columns])), use_container_width=True, hide_index=True)

    elif booking_section == "Calendário mensal":
        st.markdown("### Calendário mensal")
        c1, c2 = st.columns([1, 2])
        with c1:
            selected_month_day = st.date_input("Mês de referência", value=date.today(), format="DD/MM/YYYY", key="calendar_month_day")
        with c2:
            include_cancelled_month = st.checkbox("Mostrar canceladas no mês", value=False, key="calendar_month_cancelled")
        month_start = date(selected_month_day.year, selected_month_day.month, 1)
        if selected_month_day.month == 12:
            month_end = date(selected_month_day.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(selected_month_day.year, selected_month_day.month + 1, 1) - timedelta(days=1)
        month_events = _calendar_events_for_equipment(
            conn,
            equipment_id,
            month_start,
            month_end,
            include_cancelled=include_cancelled_month,
        )
        st.caption(f"{month_start.strftime('%m/%Y')} · {len(month_events)} evento(s) no mês para este equipamento.")
        st.markdown(_render_month_calendar(month_events, selected_eq, selected_month_day), unsafe_allow_html=True)

    elif booking_section == "Agenda linear":
        st.markdown("### Agenda linear")
        st.caption("Visual técnico complementar, útil quando há muitos eventos no mesmo período.")
        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            start_day = st.date_input("Início da visualização", value=date.today(), format="DD/MM/YYYY", key="agenda_start")
        with c2:
            days = st.selectbox("Período", [7, 14, 30, 60], index=1, format_func=lambda x: f"{x} dias", key="agenda_days")
        with c3:
            include_cancelled = st.checkbox("Mostrar canceladas", value=False, key="agenda_include_cancelled")
        end_day = start_day + timedelta(days=int(days))
        agenda_df = _booking_query_for_equipment(conn, equipment_id, start_day, end_day, include_cancelled=include_cancelled)

        if agenda_df.empty:
            st.info("Não há reservas para este equipamento no período selecionado.")
        else:
            graph_df = agenda_df.copy()
            graph_df["Início"] = pd.to_datetime(graph_df["start_datetime"])
            graph_df["Fim"] = pd.to_datetime(graph_df["end_datetime"])
            graph_df["Início formatado"] = graph_df["start_datetime"].map(_format_datetime)
            graph_df["Fim formatado"] = graph_df["end_datetime"].map(_format_datetime)
            graph_df["Reserva"] = graph_df.apply(
                lambda r: f"#{r['id']} · {r['solicitante']} · {status_badge(r['status'])}", axis=1
            )
            graph_df["Status"] = graph_df["status"].map(status_badge)
            fig = px.timeline(
                graph_df,
                x_start="Início",
                x_end="Fim",
                y="Reserva",
                color="Status",
                custom_data=["Início formatado", "Fim formatado"],
                hover_data={"Início": False, "Fim": False, "Reserva": False},
                color_discrete_sequence=[LAB_BLUE, LAB_CYAN, "#6BAED6", "#9ECAE1"],
            )
            fig.update_traces(hovertemplate="<b>%{y}</b><br>Início: %{customdata[0]}<br>Fim: %{customdata[1]}<extra></extra>")
            fig.update_yaxes(autorange="reversed")
            fig.update_layout(height=max(300, min(720, 70 + 36 * len(graph_df))), margin=dict(l=20, r=20, t=20, b=20))
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(_display_df(agenda_df), use_container_width=True, hide_index=True)

    elif booking_section == "Nova reserva":
        st.markdown("### Criar reserva")
        project_id = None
        service_id = None
        pc1, pc2 = st.columns(2)
        with pc1:
            project_options = _project_options(active_projects)
            project_label = st.selectbox("Projeto", project_options, key="booking_project")
            if project_label != "Sem projeto específico" and not active_projects.empty:
                project_id = _project_id_from_label(active_projects, project_label)
        with pc2:
            if project_id is not None:
                project_services = _project_services_for_project(project_id, active_only=True)
                service_label = st.selectbox(
                    "Serviço/análise",
                    _service_options(project_services),
                    key=f"booking_service_{project_id}",
                )
                service_id = _service_id_from_label(project_services, service_label)
            else:
                st.caption("Selecione um projeto para vincular um serviço/análise.")

        with st.form("form_nova_reserva", clear_on_submit=False):
            c1, c2, c3 = st.columns(3)
            with c1:
                if can_choose_requester:
                    user_placeholder = "Selecione o solicitante"
                    user_labels = [user_placeholder] + _user_options(active_users)
                    user_label = st.selectbox("Solicitante", user_labels, key="booking_user")
                    user_id = None
                    if user_label != user_placeholder:
                        user_id = int(active_users.iloc[user_labels.index(user_label) - 1]["id"])
                else:
                    user_id = current_user_id
                    user_label = clean_value(booking_user.get("full_name"))
                    st.text_input("Solicitante", value=user_label, disabled=True, key="booking_user_display")
                    st.caption("Reservas de membro são registradas no próprio nome.")
                booking_date = st.date_input("Data", value=date.today(), format="DD/MM/YYYY", key="booking_date")
            with c2:
                start_t = st.time_input("Horário inicial", value=time(9, 0), step=timedelta(minutes=30), key="booking_start")
                end_t = st.time_input("Horário final", value=time(10, 0), step=timedelta(minutes=30), key="booking_end")
                sample_count = st.number_input("Número de amostras", min_value=0, step=1, value=0, key="booking_samples")
            with c3:
                operator_id = None
                if operator_required:
                    operator_placeholder = "Selecione o operador"
                    op_options = _operator_options(operators, placeholder=operator_placeholder)
                    op_label = st.selectbox("Operador", op_options, key="booking_operator")
                    if op_label != operator_placeholder and not operators.empty:
                        operator_id = int(operators.iloc[op_options.index(op_label) - 1]["id"])
                    st.caption("Este equipamento requer operador. Selecione o responsável pela operação.")
                else:
                    st.info("Este equipamento não exige operador obrigatório.")

            purpose = st.text_area("Finalidade / observações", placeholder="Ex.: análise MEV com EDS; ensaio mecânico; preparação de amostras...", key="booking_purpose")
            submitted = st.form_submit_button("Nova Reserva", type="primary")

        if submitted:
            start_dt = datetime.combine(booking_date, start_t)
            end_dt = datetime.combine(booking_date, end_t)
            max_capacity = None if is_blank(selected_eq.get("max_sample_capacity")) else int(selected_eq.get("max_sample_capacity"))
            capacity_unit = clean_value(selected_eq.get("capacity_unit"), "amostras")
            if user_id is None:
                st.error("Selecione o solicitante para criar a reserva.")
            elif not can_choose_requester and user_id != current_user_id:
                st.error("Usuário comum só pode criar reserva no próprio nome.")
            elif end_dt <= start_dt:
                st.error("O horário final precisa ser maior que o horário inicial.")
            elif operator_required and operator_id is None:
                st.error("Selecione o operador responsável para este equipamento.")
            elif max_capacity and sample_count and int(sample_count) > max_capacity and truthy(selected_eq.get("capacity_enforced")):
                st.error(f"A quantidade excede a capacidade máxima cadastrada para este equipamento ({max_capacity} {capacity_unit}).")
            else:
                if max_capacity and sample_count and int(sample_count) > max_capacity:
                    st.warning("A quantidade informada excede a capacidade usual. Como o bloqueio rígido não está ativo, a reserva será tentada mesmo assim.")
                ok, msg, booking_id = create_booking(
                    conn,
                    equipment_id=equipment_id,
                    user_id=user_id,
                    project_id=project_id,
                    service_id=service_id,
                    operator_id=operator_id,
                    performed_by_id=user_id,
                    start_iso=start_dt.isoformat(timespec="minutes"),
                    end_iso=end_dt.isoformat(timespec="minutes"),
                    sample_count=int(sample_count) if sample_count else None,
                    purpose=purpose,
                    changed_by_id=current_user_id,
                )
                if ok:
                    st.session_state["booking_confirmation"] = {
                        "id": booking_id,
                        "equipment": f"{clean_value(selected_eq.get('equipment_code'))} — {clean_value(selected_eq.get('equipment_name'))}",
                        "user": user_label,
                        "period": f"{start_dt.strftime('%d/%m/%Y %H:%M')} a {end_dt.strftime('%H:%M')}",
                        "samples": str(int(sample_count)) if sample_count else "-",
                    }
                    clear_app_caches()
                    st.rerun()
                else:
                    st.error(msg)

    elif booking_section == "Gerenciar reservas":
        st.markdown("### Gerenciar reservas")
        booking_status_feedback = st.session_state.pop("booking_status_feedback", None)
        if booking_status_feedback:
            st.success(booking_status_feedback)
        c1, c2 = st.columns([1, 1])
        with c1:
            status_label = st.selectbox("Filtrar por status", ["Todos"] + list(STATUS_LABELS.values()), key="manage_status_filter")
        with c2:
            manage_days = st.slider("Próximos dias", 1, 90, 30, key="manage_days")
        manage_start = date.today() - timedelta(days=1)
        manage_end = date.today() + timedelta(days=int(manage_days))
        manage_df = _booking_query_for_equipment(conn, equipment_id, manage_start, manage_end, include_cancelled=True)
        if status_label != "Todos" and not manage_df.empty:
            internal_status = STATUS_REVERSE[status_label]
            manage_df = manage_df[manage_df["status"] == internal_status]

        if manage_df.empty:
            st.info("Nenhuma reserva encontrada para os filtros atuais.")
        else:
            st.dataframe(_display_df(manage_df), use_container_width=True, hide_index=True)
            options = manage_df["id"].tolist()
            def _booking_format(x):
                row = manage_df[manage_df["id"] == x].iloc[0]
                return f"#{x} · {row['solicitante']} · {_format_datetime(row['start_datetime'])} · {status_badge(row['status'])}"
            booking_id = st.selectbox("Selecionar reserva", options, format_func=_booking_format, key="manage_booking_id")
            booking_row = conn.execute("SELECT user_id, status FROM bookings WHERE id = ?", [int(booking_id)]).fetchone()
            booking_owner_id = int(booking_row["user_id"]) if booking_row else None
            booking_status = str(booking_row["status"]) if booking_row else ""
            can_manage_any_booking = can_manage_master_data()
            is_own_booking = current_user_id is not None and booking_owner_id == current_user_id
            can_cancel_own_booking = is_own_booking and booking_status == "scheduled"

            status_reason = ""
            if can_manage_any_booking or can_cancel_own_booking:
                status_reason = st.text_area(
                    "Observação/justificativa da mudança de status",
                    placeholder="Obrigatória para cancelamento, não comparecimento e correções de reservas já finalizadas.",
                    height=90,
                    key=f"booking_status_reason_{booking_id}",
                )

            def _submit_booking_status_change(new_status: str) -> None:
                reason = clean_input(status_reason)
                if _booking_status_reason_required(booking_status, new_status, is_manager=can_manage_any_booking) and not reason:
                    st.error("Informe uma justificativa para esta mudança de status.")
                    return
                ok, msg = change_booking_status(
                    conn,
                    int(booking_id),
                    new_status,
                    changed_by_id=current_user_id,
                    reason=reason or None,
                    source="ui",
                )
                if ok:
                    st.session_state["booking_status_feedback"] = msg
                    clear_app_caches()
                    st.rerun()
                else:
                    st.error(msg)

            if can_manage_any_booking:
                a1, a2, a3 = st.columns(3)
                with a1:
                    if st.button("Marcar como concluída", key="booking_mark_done"):
                        _submit_booking_status_change("done")
                with a2:
                    if st.button("Cancelar reserva", key="booking_cancel"):
                        _submit_booking_status_change("cancelled")
                with a3:
                    if st.button("Marcar como não compareceu", key="booking_no_show"):
                        _submit_booking_status_change("no_show")
            elif is_own_booking:
                st.caption("Você pode cancelar apenas suas próprias reservas agendadas. Conclusão e não comparecimento são ações de Gerente/Administrador.")
                if can_cancel_own_booking:
                    if st.button("Cancelar minha reserva", key="booking_cancel_own"):
                        _submit_booking_status_change("cancelled")
                else:
                    st.info("Esta reserva não está agendada e não pode ser cancelada por membro comum.")
            else:
                st.info("Você pode consultar esta reserva, mas alterações de status são restritas ao solicitante ou a Gerente/Administrador.")

            with st.expander("Histórico da reserva selecionada"):
                render_booking_status_history(conn, int(booking_id))


def page_table(conn, table_name: str, title: str):
    hero()
    st.subheader(title)
    df = query_df(conn, f"SELECT * FROM {table_name}")
    if df.empty:
        st.info("Nenhum registro encontrado.")
    else:
        display = _display_df(df)
        st.dataframe(display, use_container_width=True, hide_index=True)
        st.download_button(
            "Baixar CSV",
            data=display.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"{table_name}.csv",
            mime="text/csv",
        )


def page_usuarios(conn):
    hero()
    st.subheader("Usuários")
    st.caption("Consulta de usuários, perfis de acesso, vínculo e treinamento. Criação, edição e alteração de perfil são restritas a Administrador.")

    if not can_view_users_directory():
        st.info("A consulta completa de usuários é restrita a Gerente ou Administrador.")
        return

    _, users, _, _ = load_reference_data(conn)
    display_cols = ["full_name", "email", "phone_e164", "role", "lab_unit", "department", "advisor_name", "training_completed", "active", "notes"]
    if users.empty:
        st.info("Nenhum usuário cadastrado.")
    else:
        st.dataframe(_display_df(users[[c for c in display_cols if c in users.columns]]), use_container_width=True, hide_index=True)

    if not can_manage_users():
        st.info("Criação, edição, ativação/inativação e alteração de perfil são restritas a Administrador.")
        return

    st.markdown("### Incluir ou atualizar usuário")
    mode = st.radio("Modo", ["Novo usuário", "Editar usuário existente"], horizontal=True, key="user_edit_mode")
    selected = None
    user_id = None
    if mode == "Editar usuário existente":
        if users.empty:
            st.info("Cadastre um usuário antes de editar.")
            return
        labels = users.apply(lambda r: f"{clean_value(r.get('full_name'))} ({role_badge(clean_value(r.get('role'), 'member'))})", axis=1).tolist()
        label = st.selectbox("Selecionar usuário", labels, key="user_edit_select")
        selected = users.iloc[labels.index(label)]
        user_id = int(selected["id"])

    with st.form("form_user_master"):
        c1, c2, c3 = st.columns(3)
        with c1:
            full_name = st.text_input("Nome completo", value=clean_input(selected.get("full_name")) if selected is not None else "")
            email = st.text_input("E-mail", value=clean_input(selected.get("email")) if selected is not None else "")
            phone = st.text_input("Celular/WhatsApp", value=clean_input(selected.get("phone_e164")) if selected is not None else "")
        with c2:
            current_role = clean_value(selected.get("role"), "member") if selected is not None else "member"
            if current_role == "operator":
                current_role = "manager"
            role_label = st.selectbox(
                "Perfil",
                list(ROLE_LABELS.values()),
                index=list(ROLE_LABELS.keys()).index(current_role) if current_role in ROLE_LABELS else 0,
            )
            lab_unit = st.text_input("Unidade/laboratório", value=clean_input(selected.get("lab_unit")) if selected is not None else "LabCim")
            department = st.text_input("Departamento/programa", value=clean_input(selected.get("department")) if selected is not None else "")
        with c3:
            advisor_name = st.text_input("Orientador(a)", value=clean_input(selected.get("advisor_name")) if selected is not None else "")
            training_completed = st.checkbox("Treinamento concluído", value=truthy(selected.get("training_completed")) if selected is not None else False)
            active = st.checkbox("Usuário ativo", value=truthy(selected.get("active")) if selected is not None else True)
        notes = st.text_area("Observações", value=clean_input(selected.get("notes")) if selected is not None else "")
        submitted = st.form_submit_button("Salvar usuário", type="primary")

    if submitted:
        payload = dict(
            full_name=full_name,
            email=email.strip() or None,
            phone_e164=phone.strip() or None,
            role=ROLE_REVERSE[role_label],
            lab_unit=lab_unit.strip() or None,
            department=department.strip() or None,
            advisor_name=advisor_name.strip() or None,
            training_completed=int(training_completed),
            active=int(active),
            notes=notes.strip() or None,
        )
        if mode == "Novo usuário":
            ok, msg = create_user(conn, **payload)
        else:
            ok, msg = update_user(conn, int(user_id), **payload)
        if ok:
            st.success(msg)
            clear_app_caches()
            st.rerun()
        else:
            st.error(msg)


def page_projetos(conn):
    hero()
    st.subheader("Projetos")
    st.caption("Cadastro simples de projetos e serviços/análises para rastrear reservas, uso de equipamentos e consumo de insumos.")

    _, users, projects, _ = load_reference_data(conn)
    user_names = {}
    if not users.empty:
        user_names = {int(row["id"]): clean_value(row.get("full_name")) for _, row in users.iterrows()}
    project_display = projects.copy()
    if not project_display.empty:
        project_display["requester_name"] = project_display["requester_id"].map(lambda value: user_names.get(int(value), "-") if not is_blank(value) else "-")
        project_display["coordinator_name"] = project_display["coordinator_id"].map(lambda value: user_names.get(int(value), "-") if not is_blank(value) else "-")
    display_cols = [
        "project_code", "project_name", "objective", "funding_source",
        "requester_name", "coordinator_name", "status", "start_date",
        "end_date", "active", "notes"
    ]
    if projects.empty:
        st.info("Nenhum projeto cadastrado.")
    else:
        st.dataframe(_display_df(project_display[[c for c in display_cols if c in project_display.columns]]), use_container_width=True, hide_index=True)

    if not can_manage_master_data():
        admin_required_message("incluir ou atualizar projetos e serviços/análises")
    else:
        st.markdown("### Incluir ou atualizar projeto")
        mode = st.radio("Modo", ["Novo projeto", "Editar projeto existente"], horizontal=True, key="project_edit_mode")
        selected = None
        project_id = None
        if mode == "Editar projeto existente":
            if projects.empty:
                st.info("Cadastre um projeto antes de editar.")
                selected = None
            else:
                labels = projects.apply(lambda r: f"{clean_value(r.get('project_code'), 'Sem código')} — {clean_value(r.get('project_name'))}", axis=1).tolist()
                label = st.selectbox("Selecionar projeto", labels, key="project_edit_select")
                selected = projects.iloc[labels.index(label)]
                project_id = int(selected["id"])

        if mode == "Novo projeto" or selected is not None:
            with st.form("form_project_master"):
                c1, c2, c3 = st.columns(3)
                with c1:
                    project_code = st.text_input("Código interno do projeto", value=clean_input(selected.get("project_code")) if selected is not None else "")
                    project_name = st.text_input("Nome do projeto", value=clean_input(selected.get("project_name")) if selected is not None else "")
                    funding_source = st.text_input("Fonte/convênio/financiamento", value=clean_input(selected.get("funding_source")) if selected is not None else "")
                with c2:
                    objective = st.text_area("Objetivo", value=clean_input(selected.get("objective")) if selected is not None else "", height=120)
                    status = st.selectbox(
                        "Status",
                        PROJECT_STATUSES,
                        index=_option_index(PROJECT_STATUSES, selected.get("status") if selected is not None else "em andamento"),
                    )
                with c3:
                    requester_options = ["Não informado"] + _user_options(users)
                    requester_label = st.selectbox(
                        "Solicitante",
                        requester_options,
                        index=_optional_user_index(users, selected.get("requester_id")) if selected is not None else 0,
                        key="project_requester",
                    )
                    coordinator_label = st.selectbox(
                        "Coordenador/responsável",
                        requester_options,
                        index=_optional_user_index(users, selected.get("coordinator_id")) if selected is not None else 0,
                        key="project_coordinator",
                    )
                    start_value = _date_input_value(selected.get("start_date"), None) if selected is not None else None
                    end_value = _date_input_value(selected.get("end_date"), None) if selected is not None else None
                    start_date_value = st.date_input("Data de início", value=start_value, key="project_start_date")
                    end_date_value = st.date_input("Data de conclusão/término", value=end_value, key="project_end_date")
                    active = st.checkbox("Projeto ativo", value=truthy(selected.get("active")) if selected is not None else True)
                notes = st.text_area("Observações", value=clean_input(selected.get("notes")) if selected is not None else "")
                submitted = st.form_submit_button("Salvar projeto", type="primary")

            if submitted:
                if start_date_value and end_date_value and start_date_value > end_date_value:
                    st.error("A data de início não pode ser posterior à data de conclusão/término.")
                    return
                payload = dict(
                    project_code=project_code.strip() or None,
                    project_name=project_name,
                    objective=objective.strip() or None,
                    funding_source=funding_source.strip() or None,
                    requester_id=_user_id_from_label(users, requester_label),
                    coordinator_id=_user_id_from_label(users, coordinator_label),
                    status=status,
                    start_date=start_date_value.isoformat() if start_date_value else None,
                    end_date=end_date_value.isoformat() if end_date_value else None,
                    active=int(active),
                    notes=notes.strip() or None,
                )
                if mode == "Novo projeto":
                    ok, msg = create_project(conn, **payload)
                else:
                    ok, msg = update_project(conn, int(project_id), **payload)
                if ok:
                    st.success(msg)
                    clear_app_caches()
                    st.rerun()
                else:
                    st.error(msg)

    st.markdown("### Serviços/análises")
    active_projects = projects[projects["active"] == 1].copy() if not projects.empty else projects
    if active_projects.empty:
        st.info("Cadastre um projeto ativo para registrar serviços/análises.")
        return
    service_project_label = st.selectbox(
        "Projeto vinculado",
        _project_options(active_projects),
        key="service_project_select",
    )
    if service_project_label == "Sem projeto específico":
        st.info("Selecione um projeto para listar ou cadastrar serviços/análises.")
        return

    service_project_id = _project_id_from_label(active_projects, service_project_label)
    services = _project_services_for_project(service_project_id, active_only=False)
    service_cols = [
        "service_code", "title", "service_type", "status", "requested_date",
        "expected_date", "completed_date", "requester_name", "responsible_name",
        "active", "notes"
    ]
    if services.empty:
        st.caption("Nenhum serviço/análise cadastrado para este projeto.")
    else:
        st.dataframe(_display_df(services[[c for c in service_cols if c in services.columns]]), use_container_width=True, hide_index=True)

    if not can_manage_master_data():
        return

    service_mode = st.radio("Modo do serviço/análise", ["Novo serviço/análise", "Editar serviço/análise existente"], horizontal=True, key="service_edit_mode")
    selected_service = None
    service_id = None
    if service_mode == "Editar serviço/análise existente":
        if services.empty:
            st.info("Cadastre um serviço/análise antes de editar.")
            return
        service_labels = services.apply(lambda r: f"{clean_value(r.get('service_code'), 'Sem código')} — {clean_value(r.get('title'))}", axis=1).tolist()
        service_label = st.selectbox("Selecionar serviço/análise", service_labels, key="service_edit_select")
        selected_service = services.iloc[service_labels.index(service_label)]
        service_id = int(selected_service["id"])

    with st.form("form_project_service"):
        c1, c2, c3 = st.columns(3)
        with c1:
            service_code = st.text_input("Código do serviço/análise", value=clean_input(selected_service.get("service_code")) if selected_service is not None else "")
            title = st.text_input("Título", value=clean_input(selected_service.get("title")) if selected_service is not None else "")
            service_type = st.selectbox(
                "Tipo de serviço/análise",
                SERVICE_TYPES,
                index=_option_index(SERVICE_TYPES, selected_service.get("service_type") if selected_service is not None else "Análise"),
            )
        with c2:
            service_status = st.selectbox(
                "Status",
                SERVICE_STATUSES,
                index=_option_index(SERVICE_STATUSES, selected_service.get("status") if selected_service is not None else "em andamento"),
                key="project_service_status",
            )
            requested_date = st.date_input("Data solicitada", value=_date_input_value(selected_service.get("requested_date"), date.today()) if selected_service is not None else date.today(), key="service_requested_date")
            expected_date = st.date_input("Data prevista", value=_date_input_value(selected_service.get("expected_date")) if selected_service is not None else None, key="service_expected_date")
            completed_date = st.date_input("Data concluída", value=_date_input_value(selected_service.get("completed_date")) if selected_service is not None else None, key="service_completed_date")
        with c3:
            user_options = ["Não informado"] + _user_options(users)
            service_requester_label = st.selectbox(
                "Solicitante",
                user_options,
                index=_optional_user_index(users, selected_service.get("requester_id")) if selected_service is not None else 0,
                key="service_requester",
            )
            responsible_label = st.selectbox(
                "Responsável",
                user_options,
                index=_optional_user_index(users, selected_service.get("responsible_id")) if selected_service is not None else 0,
                key="service_responsible",
            )
            service_active = st.checkbox("Serviço/análise ativo", value=truthy(selected_service.get("active")) if selected_service is not None else True)
        service_notes = st.text_area("Observações do serviço/análise", value=clean_input(selected_service.get("notes")) if selected_service is not None else "")
        service_submitted = st.form_submit_button("Salvar serviço/análise", type="primary")

    if service_submitted:
        if completed_date and requested_date and completed_date < requested_date:
            st.error("A data concluída não pode ser anterior à data solicitada.")
        else:
            service_payload = dict(
                project_id=service_project_id,
                service_code=service_code.strip() or None,
                title=title.strip(),
                service_type=service_type,
                requester_id=_user_id_from_label(users, service_requester_label),
                responsible_id=_user_id_from_label(users, responsible_label),
                status=service_status,
                requested_date=requested_date.isoformat() if requested_date else None,
                expected_date=expected_date.isoformat() if expected_date else None,
                completed_date=completed_date.isoformat() if completed_date else None,
                notes=service_notes.strip() or None,
                active=int(service_active),
            )
            if service_mode == "Novo serviço/análise":
                ok, msg, _ = create_project_service(conn, **service_payload)
            else:
                ok, msg = update_project_service(conn, int(service_id), **service_payload)
            if ok:
                st.success(msg)
                clear_app_caches()
                st.rerun()
            else:
                st.error(msg)

    if selected_service is not None and truthy(selected_service.get("active")):
        with st.expander("Inativar serviço/análise", expanded=False):
            st.caption("Use inativação para lançamentos encerrados por erro, sem exclusão definitiva.")
            if st.button("Inativar serviço/análise", key=f"inactivate_service_{int(selected_service['id'])}"):
                ok, msg = inactivate_project_service(conn, int(selected_service["id"]))
                if ok:
                    st.success(msg)
                    clear_app_caches()
                    st.rerun()
                else:
                    st.error(msg)


def page_equipamentos(conn):
    hero()
    st.subheader("Equipamentos")
    st.caption("Cadastro operacional simples: status, capacidade, funcionalidades indisponíveis, documentação e localização. O cadastro mestre fica restrito a gerente/administrador.")

    equipment, _, _, _ = load_reference_data(conn)
    display_cols = [
        "equipment_code",
        "equipment_name",
        "lab_unit",
        "location",
        "operational_status",
        "max_sample_capacity",
        "capacity_unit",
        "capacity_enforced",
        "unavailable_functions",
        "technical_manager",
        "pop_title",
        "pop_version",
        "pop_responsible",
        "requires_operator",
        "active",
    ]
    if equipment.empty:
        st.info("Nenhum equipamento cadastrado.")
    else:
        existing_cols = [c for c in display_cols if c in equipment.columns]
        st.dataframe(_display_df(equipment[existing_cols]), use_container_width=True, hide_index=True)

    tab_oper, tab_parts, tab_docs, tab_master = st.tabs([
        "Atualizar dados operacionais",
        "Peças de reposição",
        "Documentos",
        "Cadastro mestre",
    ])

    with tab_oper:
        if equipment.empty:
            st.info("Cadastre um equipamento antes de atualizar dados operacionais.")
        elif not can_edit_operational_data():
            st.info("Para atualizar dados operacionais, use perfil Gerente ou Administrador.")
        else:
            eq_label = st.selectbox("Selecionar equipamento", _equipment_options(equipment), key="equip_edit_select")
            equipment_id = _equipment_id_from_label(equipment, eq_label)
            selected = equipment[equipment["id"] == equipment_id].iloc[0]

            with st.form("form_equip_operational"):
                c1, c2, c3 = st.columns(3)
                with c1:
                    current_status = selected.get("operational_status") or "available"
                    status_label = st.selectbox(
                        "Status operacional",
                        list(EQUIPMENT_STATUS_LABELS.values()),
                        index=list(EQUIPMENT_STATUS_LABELS.keys()).index(current_status) if current_status in EQUIPMENT_STATUS_LABELS else 0,
                    )
                    st.caption("Uso restrito sinaliza limitação operacional, mas não bloqueia reserva automaticamente. Para inativar, use o cadastro mestre.")
                    location = st.text_input("Localização", value=clean_input(selected.get("location")), placeholder="Ex.: Lab Tecnológico, Sala de Raios X...")
                    technical_manager = st.text_input("Gestor técnico", value=clean_input(selected.get("technical_manager")) or clean_input(selected.get("responsible_name")))
                with c2:
                    raw_capacity = selected.get("max_sample_capacity")
                    initial_capacity = int(raw_capacity) if not is_blank(raw_capacity) and raw_capacity else 0
                    max_sample_capacity = st.number_input("Capacidade máxima por reserva", min_value=0, value=initial_capacity, step=1)
                    capacity_unit = st.text_input("Unidade da capacidade", value=clean_value(selected.get("capacity_unit"), "amostras"))
                    capacity_enforced = st.checkbox("Bloquear acima da capacidade", value=truthy(selected.get("capacity_enforced")))
                    st.caption("Capacidade usual apenas alerta; bloqueio rígido impede reserva acima da capacidade.")
                with c3:
                    unavailable_functions = st.text_area(
                        "Funcionalidades indisponíveis",
                        value=clean_input(selected.get("unavailable_functions")),
                        placeholder="Ex.: EDS indisponível; vácuo parcial; forno sem rampa automática...",
                    )
                    notes = st.text_area("Observações operacionais", value=clean_input(selected.get("notes")))

                st.markdown("#### Documentação operacional")
                d1, d2, d3 = st.columns(3)
                with d1:
                    pop_title = st.text_input("Título do POP", value=clean_input(selected.get("pop_title")), placeholder="Ex.: POP - Autoclave")
                    pop_version = st.text_input("Versão do POP", value=clean_input(selected.get("pop_version")), placeholder="Ex.: v1, Rev. 02")
                with d2:
                    pop_path = st.text_input("Arquivo/link do POP", value=clean_input(selected.get("pop_path")), placeholder="Ex.: assets/pops/POP_Autoclave.pdf")
                    pop_updated_at = st.text_input("Data de atualização do POP", value=clean_input(selected.get("pop_updated_at")), placeholder="Ex.: 19/06/2026")
                with d3:
                    pop_responsible = st.text_input("Responsável pelo POP", value=clean_input(selected.get("pop_responsible")) or clean_input(selected.get("technical_manager")) or clean_input(selected.get("responsible_name")))
                    document_notes = st.text_area("Observações documentais", value=clean_input(selected.get("document_notes")))

                submitted = st.form_submit_button("Salvar dados operacionais", type="primary")

            pop_download_button(selected, key_prefix="equip_pop")

            if submitted:
                old_status = clean_input(selected.get("operational_status")) or "available"
                new_status = EQUIPMENT_STATUS_REVERSE[status_label]
                if new_status == "inactive":
                    st.error("Para inativar equipamento, use o cadastro mestre e informe uma justificativa.")
                else:
                    update_equipment_operational_info(
                        conn,
                        equipment_id,
                        location=location.strip() or None,
                        operational_status=new_status,
                        unavailable_functions=unavailable_functions.strip() or None,
                        max_sample_capacity=int(max_sample_capacity) if max_sample_capacity else None,
                        capacity_unit=capacity_unit.strip() or "amostras",
                        capacity_enforced=int(capacity_enforced),
                        technical_manager=technical_manager.strip() or None,
                        pop_title=pop_title.strip() or None,
                        pop_path=pop_path.strip() or None,
                        pop_version=pop_version.strip() or None,
                        pop_updated_at=pop_updated_at.strip() or None,
                        pop_responsible=pop_responsible.strip() or None,
                        document_notes=document_notes.strip() or None,
                        notes=notes.strip() or None,
                    )
                    if new_status == "maintenance" and old_status != "maintenance":
                        sent, total = notify_equipment_maintenance(
                            conn,
                            equipment_id=equipment_id,
                            title="equipamento em manutenção",
                            message=(
                                "O equipamento foi marcado como EM MANUTENÇÃO no sistema.\n"
                                f"Observações: {notes.strip() or unavailable_functions.strip() or 'Sem observações adicionais.'}"
                            ),
                            related_table="equipment",
                            related_id=equipment_id,
                        )
                        if total:
                            st.info(f"Notificação de manutenção registrada para {total} destinatário(s). Enviadas: {sent}.")
                    st.success("Dados operacionais do equipamento atualizados.")
                    clear_app_caches()
                    st.rerun()

    with tab_parts:
        st.markdown("### Peças de reposição associadas")
        if equipment.empty:
            st.info("Cadastre um equipamento antes de consultar peças de reposição.")
        else:
            eq_label = st.selectbox("Selecionar equipamento", _equipment_options(equipment), key="equipment_spare_parts_select")
            equipment_id = _equipment_id_from_label(equipment, eq_label)
            selected = equipment[equipment["id"] == equipment_id].iloc[0]
            st.caption(
                f"{clean_value(selected.get('equipment_code'))} — {clean_value(selected.get('equipment_name'))} · "
                f"Local: {clean_value(selected.get('location'))}"
            )
            spare_parts = list_spare_parts_for_equipment(conn, equipment_id)
            render_equipment_spare_parts(spare_parts)

    with tab_docs:
        st.markdown("### Documentos operacionais")
        if equipment.empty:
            st.info("Cadastre um equipamento antes de consultar documentos.")
        else:
            eq_label = st.selectbox("Selecionar equipamento", _equipment_options(equipment), key="equipment_documents_select")
            equipment_id = _equipment_id_from_label(equipment, eq_label)
            selected = equipment[equipment["id"] == equipment_id].iloc[0]
            st.caption(
                f"{clean_value(selected.get('equipment_code'))} — {clean_value(selected.get('equipment_name'))} · "
                f"Local: {clean_value(selected.get('location'))}"
            )
            render_equipment_documents_section(
                conn,
                selected,
                key_prefix=f"equipment_documents_{equipment_id}",
                can_manage=can_manage_master_data(),
            )

    with tab_master:
        if not can_manage_master_data():
            admin_required_message("incluir ou atualizar equipamentos")
        else:
            mode = st.radio("Modo", ["Novo equipamento", "Editar equipamento existente"], horizontal=True, key="equipment_master_mode")
            selected = None
            equipment_id = None
            if mode == "Editar equipamento existente":
                if equipment.empty:
                    st.info("Cadastre um equipamento antes de editar.")
                    return
                eq_label = st.selectbox("Selecionar equipamento", _equipment_options(equipment), key="equipment_master_select")
                equipment_id = _equipment_id_from_label(equipment, eq_label)
                selected = equipment[equipment["id"] == equipment_id].iloc[0]

            with st.form("form_equipment_master"):
                c1, c2, c3 = st.columns(3)
                with c1:
                    equipment_code = st.text_input("Código/patrimônio", value=clean_input(selected.get("equipment_code")) if selected is not None else "")
                    equipment_name = st.text_input("Nome do equipamento", value=clean_input(selected.get("equipment_name")) if selected is not None else "")
                    lab_unit = st.text_input("Unidade/laboratório", value=clean_input(selected.get("lab_unit")) if selected is not None else "LabCim")
                    location = st.text_input("Localização", value=clean_input(selected.get("location")) if selected is not None else "")
                with c2:
                    responsible_name = st.text_input("Responsável", value=clean_input(selected.get("responsible_name")) if selected is not None else "")
                    responsible_phone = st.text_input("Telefone do responsável", value=clean_input(selected.get("responsible_phone")) if selected is not None else "")
                    technical_manager = st.text_input("Gestor técnico", value=clean_input(selected.get("technical_manager")) if selected is not None else "")
                    requires_operator = st.checkbox("Requer operador", value=truthy(selected.get("requires_operator")) if selected is not None else False)
                    st.caption("Quando marcado, o operador passa a ser obrigatório na criação da reserva.")
                with c3:
                    current_status = selected.get("operational_status") if selected is not None else "available"
                    status_label = st.selectbox(
                        "Status operacional",
                        list(EQUIPMENT_STATUS_LABELS.values()),
                        index=list(EQUIPMENT_STATUS_LABELS.keys()).index(current_status) if current_status in EQUIPMENT_STATUS_LABELS else 0,
                        key="equipment_master_status",
                    )
                    st.caption("Uso restrito sinaliza limitação operacional, mas não bloqueia reserva automaticamente. Inativação deve corresponder ao campo ativo.")
                    raw_capacity = selected.get("max_sample_capacity") if selected is not None else None
                    initial_capacity = int(raw_capacity) if not is_blank(raw_capacity) and raw_capacity else 0
                    max_sample_capacity = st.number_input("Capacidade máxima", min_value=0, value=initial_capacity, step=1, key="equipment_master_capacity")
                    capacity_unit = st.text_input("Unidade da capacidade", value=clean_value(selected.get("capacity_unit"), "amostras") if selected is not None else "amostras")
                    capacity_enforced = st.checkbox("Bloquear acima da capacidade", value=truthy(selected.get("capacity_enforced")) if selected is not None else False, key="equipment_master_enforce")
                    st.caption("Capacidade usual apenas alerta; bloqueio rígido impede reserva acima da capacidade.")
                    active = st.checkbox("Equipamento ativo", value=truthy(selected.get("active")) if selected is not None else True)

                unavailable_functions = st.text_area("Funcionalidades indisponíveis", value=clean_input(selected.get("unavailable_functions")) if selected is not None else "")
                notes = st.text_area("Observações", value=clean_input(selected.get("notes")) if selected is not None else "")
                st.markdown("#### POP / documentação operacional")
                d1, d2, d3 = st.columns(3)
                with d1:
                    pop_title = st.text_input("Título do POP", value=clean_input(selected.get("pop_title")) if selected is not None else "")
                    pop_version = st.text_input("Versão do POP", value=clean_input(selected.get("pop_version")) if selected is not None else "")
                with d2:
                    pop_path = st.text_input("Arquivo/link do POP", value=clean_input(selected.get("pop_path")) if selected is not None else "")
                    pop_updated_at = st.text_input("Data de atualização do POP", value=clean_input(selected.get("pop_updated_at")) if selected is not None else "")
                with d3:
                    pop_responsible = st.text_input("Responsável pelo POP", value=clean_input(selected.get("pop_responsible")) if selected is not None else "")
                    document_notes = st.text_area("Observações documentais", value=clean_input(selected.get("document_notes")) if selected is not None else "")
                inactive_reason = ""
                if mode == "Editar equipamento existente" and selected is not None and truthy(selected.get("active")):
                    inactive_reason = st.text_area(
                        "Motivo da inativação",
                        value="",
                        placeholder="Obrigatório se desmarcar Equipamento ativo.",
                        key="equipment_inactive_reason",
                    )

                submitted = st.form_submit_button("Salvar equipamento", type="primary")

            if submitted:
                new_operational_status = EQUIPMENT_STATUS_REVERSE[status_label]
                is_inactivation = mode == "Editar equipamento existente" and selected is not None and truthy(selected.get("active")) and not active
                if not active:
                    new_operational_status = "inactive"
                payload = dict(
                    equipment_code=equipment_code,
                    equipment_name=equipment_name,
                    lab_unit=lab_unit.strip() or None,
                    location=location.strip() or None,
                    requires_operator=int(requires_operator),
                    responsible_name=responsible_name.strip() or None,
                    responsible_phone=responsible_phone.strip() or None,
                    active=int(active),
                    operational_status=new_operational_status,
                    unavailable_functions=unavailable_functions.strip() or None,
                    max_sample_capacity=int(max_sample_capacity) if max_sample_capacity else None,
                    capacity_unit=capacity_unit.strip() or "amostras",
                    capacity_enforced=int(capacity_enforced),
                    technical_manager=technical_manager.strip() or None,
                    pop_title=pop_title.strip() or None,
                    pop_path=pop_path.strip() or None,
                    pop_version=pop_version.strip() or None,
                    pop_updated_at=pop_updated_at.strip() or None,
                    pop_responsible=pop_responsible.strip() or None,
                    document_notes=document_notes.strip() or None,
                    notes=notes.strip() or None,
                )
                if mode == "Novo equipamento":
                    ok, msg = create_equipment(conn, **payload)
                else:
                    old_status = clean_input(selected.get("operational_status")) or "available"
                    if is_inactivation and not inactive_reason.strip():
                        ok, msg = False, "Informe o motivo da inativação do equipamento."
                    elif active and new_operational_status == "inactive":
                        ok, msg = False, "Para usar status Inativo, desmarque Equipamento ativo e informe uma justificativa."
                    else:
                        ok, msg = update_equipment_master(
                            conn,
                            int(equipment_id),
                            **payload,
                            inactive_reason=inactive_reason.strip() or None,
                            inactive_by_id=_current_user_id(),
                        )
                    if ok and payload["operational_status"] == "maintenance" and old_status != "maintenance":
                        sent, total = notify_equipment_maintenance(
                            conn,
                            equipment_id=int(equipment_id),
                            title="equipamento em manutenção",
                            message=(
                                "O equipamento foi marcado como EM MANUTENÇÃO no cadastro mestre.\n"
                                f"Observações: {payload.get('notes') or payload.get('unavailable_functions') or 'Sem observações adicionais.'}"
                            ),
                            related_table="equipment",
                            related_id=int(equipment_id),
                        )
                        if total:
                            st.info(f"Notificação de manutenção registrada para {total} destinatário(s). Enviadas: {sent}.")
                if ok:
                    st.success(msg)
                    clear_app_caches()
                    st.rerun()
                else:
                    st.error(msg)

    with st.expander("Biblioteca de POPs disponíveis no projeto"):
        docs = sorted(POP_DIR.glob("*.pdf")) if POP_DIR.exists() else []
        if not docs:
            st.info("Nenhum PDF de POP encontrado em assets/pops.")
        else:
            doc_names = [doc.name for doc in docs]
            selected_doc_name = st.selectbox("PDF disponível", doc_names, key="library_pop_select")
            selected_doc = docs[doc_names.index(selected_doc_name)]
            st.caption(selected_doc.name)
            st.download_button(
                f"📄 Baixar {selected_doc.name}",
                data=selected_doc.read_bytes(),
                file_name=selected_doc.name,
                mime="application/pdf",
                key=f"library_pop_{selected_doc.name}",
            )



def _equipment_options(equipment: pd.DataFrame) -> list[str]:
    return equipment.apply(lambda r: f"{clean_value(r.get('equipment_code'))} — {clean_value(r.get('equipment_name'))}", axis=1).tolist()


def _equipment_id_from_label(equipment: pd.DataFrame, label: str) -> int:
    labels = _equipment_options(equipment)
    return int(equipment.iloc[labels.index(label)]["id"])


def _current_user_id() -> int | None:
    user = current_user()
    try:
        return int(user.get("id")) if user.get("id") is not None else None
    except Exception:
        return None


def _attachment_ref(attachment_id: int) -> str:
    return f"attachment:{int(attachment_id)}"


def _attachment_id_from_ref(path_value) -> int | None:
    path_text = clean_input(path_value)
    if not path_text.lower().startswith("attachment:"):
        return None
    try:
        return int(path_text.split(":", 1)[1])
    except Exception:
        return None


def _ensure_storage_ready_for_upload(*uploaded_files) -> bool:
    if not any(uploaded_file is not None for uploaded_file in uploaded_files):
        return True
    try:
        active_storage_backend()
        return True
    except StorageConfigurationError as exc:
        st.error(str(exc))
        return False


def _save_upload(
    conn,
    uploaded_file,
    *,
    entity_type: str,
    entity_id: int,
    attachment_role: str,
    notes: str | None = None,
) -> str | None:
    if uploaded_file is None:
        return None
    backend = active_storage_backend()
    content = uploaded_file.getvalue()
    stored = backend.save_file(
        entity_type=entity_type,
        entity_id=int(entity_id),
        original_filename=uploaded_file.name,
        content=content,
        mime_type=getattr(uploaded_file, "type", None),
    )
    attachment_id = create_attachment(
        conn,
        entity_type=entity_type,
        entity_id=int(entity_id),
        attachment_role=attachment_role,
        original_filename=stored.original_filename,
        storage_key=stored.storage_key,
        storage_backend=stored.storage_backend,
        mime_type=stored.mime_type,
        file_size=stored.file_size,
        sha256=stored.sha256,
        uploaded_by_id=_current_user_id(),
        notes=notes,
    )
    return _attachment_ref(attachment_id)


def _supply_options(supplies: pd.DataFrame) -> list[str]:
    def _label(r: pd.Series) -> str:
        qty = 0.0 if is_blank(r.get("current_quantity")) else float(r.get("current_quantity"))
        item_type = clean_value(r.get("supply_type"), "Insumo")
        code = clean_input(r.get("supply_code"))
        code_text = f"{code} · " if code else ""
        return f"{int(r['id'])} — {code_text}{clean_value(r.get('supply_name'))} · {item_type} · saldo: {qty:g} {clean_value(r.get('unit'), '')}"
    return supplies.apply(_label, axis=1).tolist()


def _supply_id_from_label(supplies: pd.DataFrame, label: str) -> int:
    labels = _supply_options(supplies)
    return int(supplies.iloc[labels.index(label)]["id"])


def _supply_type_value(row: pd.Series | None) -> str:
    value = clean_input(row.get("supply_type")) if row is not None else ""
    return value if value in SUPPLY_TYPES else "Insumo"


def _is_spare_part(row: pd.Series | None) -> bool:
    return _supply_type_value(row) == "Peça de reposição"


def _spare_part_stock_status(row: pd.Series) -> str:
    qty = float(row.get("current_quantity") or 0)
    min_qty = float(row.get("minimum_quantity") or 0)
    return "Abaixo do mínimo" if qty < min_qty else "OK"


def _equipment_ids_from_labels(equipment: pd.DataFrame, labels: list[str]) -> list[int]:
    if equipment.empty:
        return []
    return [_equipment_id_from_label(equipment, label) for label in labels]


def render_equipment_spare_parts(spare_parts: pd.DataFrame) -> None:
    if spare_parts.empty:
        st.info("Nenhuma peça de reposição associada a este equipamento.")
        return
    display = spare_parts.copy()
    display["stock_status"] = display.apply(_spare_part_stock_status, axis=1)
    cols = [
        "supply_name",
        "supply_code",
        "manufacturer_code",
        "manufacturer",
        "current_quantity",
        "unit",
        "minimum_quantity",
        "stock_status",
        "location",
        "compatible_model_family",
        "association_notes",
    ]
    st.dataframe(
        _display_df(display[[c for c in cols if c in display.columns]]),
        use_container_width=True,
        hide_index=True,
    )


def _option_index(options: list[str], value: str | None, default: int = 0) -> int:
    value = clean_input(value)
    return options.index(value) if value in options else default


def _maintenance_status_requires_justification(previous_status: str | None, new_status: str, *, creating: bool) -> bool:
    previous = clean_input(previous_status)
    new = clean_input(new_status)
    if not creating and previous == new:
        return False
    if new in MAINTENANCE_JUSTIFICATION_STATUSES:
        return True
    return new == "pendente" and not creating


def _datetime_input_defaults(value, fallback: datetime | None = None) -> tuple[date, time]:
    fallback = fallback or datetime.now().replace(second=0, microsecond=0)
    if not is_blank(value):
        try:
            parsed = datetime.fromisoformat(str(value))
            return parsed.date(), parsed.time().replace(second=0, microsecond=0)
        except Exception:
            try:
                parsed = pd.to_datetime(value).to_pydatetime()
                return parsed.date(), parsed.time().replace(second=0, microsecond=0)
            except Exception:
                pass
    return fallback.date(), fallback.time().replace(second=0, microsecond=0)


def _maintenance_reference_label(row: pd.Series) -> str:
    reference = clean_input(row.get("scheduled_reference")) or clean_input(row.get("next_date")) or clean_input(row.get("planned_date"))
    return _format_datetime(reference) if reference else "sem próxima data"


def render_maintenance_status_history(conn, *, entity_type: str, entity_id: int) -> None:
    history = list_maintenance_status_history(conn, entity_type=entity_type, entity_id=entity_id)
    if history.empty:
        st.caption("Nenhuma alteração de status registrada.")
        return
    cols = ["previous_status", "new_status", "justification", "changed_by_name", "changed_at"]
    st.dataframe(
        _display_df(history[[c for c in cols if c in history.columns]]),
        use_container_width=True,
        hide_index=True,
    )


def _select_index_by_supply_id(supplies: pd.DataFrame, supply_id: str | int | None) -> int:
    if supply_id is None or supplies.empty:
        return 0
    try:
        supply_id = int(supply_id)
    except Exception:
        return 0
    ids = supplies["id"].astype(int).tolist()
    return ids.index(supply_id) if supply_id in ids else 0


def _supply_lots_for_supply(supply_lots: pd.DataFrame, supply_id: int, *, active_only: bool = False) -> pd.DataFrame:
    if supply_lots.empty:
        return supply_lots
    lots = supply_lots[supply_lots["supply_id"].astype(int) == int(supply_id)].copy()
    if active_only and "is_active" in lots.columns:
        lots = lots[lots["is_active"].fillna(1).astype(int) == 1]
    return lots


def _lot_expiration_status(expiration_value) -> str:
    if is_blank(expiration_value):
        return "Sem validade"
    try:
        expiration = datetime.fromisoformat(str(expiration_value)).date()
    except Exception:
        return "Validade inválida"
    today = date.today()
    if expiration < today:
        return "Vencido"
    if expiration <= today + timedelta(days=60):
        return "Vence em até 60 dias"
    return "OK"


def _days_until_date(value) -> int | None:
    if is_blank(value):
        return None
    try:
        target = datetime.fromisoformat(str(value)).date()
    except Exception:
        return None
    return (target - date.today()).days


def _lot_display_df(lots: pd.DataFrame) -> pd.DataFrame:
    if lots.empty:
        return lots
    out = lots.copy()
    out["lot_status"] = out["expiration_date"].map(_lot_expiration_status) if "expiration_date" in out.columns else "Sem validade"
    cols = [
        "lot_code",
        "expiration_date",
        "lot_status",
        "current_quantity",
        "unit",
        "initial_quantity",
        "supplier_name",
        "location",
        "received_date",
        "is_active",
        "notes",
    ]
    return _display_df(out[[c for c in cols if c in out.columns]])


def _supply_lot_options(lots: pd.DataFrame) -> list[str]:
    def _label(row: pd.Series) -> str:
        qty = float(row.get("current_quantity") or 0)
        unit = clean_value(row.get("unit"), "")
        return (
            f"{int(row['id'])} — lote {clean_value(row.get('lot_code'))} · "
            f"saldo: {qty:g} {unit} · validade: {_format_datetime(row.get('expiration_date'))}"
        )

    return lots.apply(_label, axis=1).tolist()


def _supply_lot_id_from_label(lots: pd.DataFrame, label: str) -> int | None:
    if label == "Sem lote específico" or lots.empty:
        return None
    options = ["Sem lote específico"] + _supply_lot_options(lots)
    return int(lots.iloc[options.index(label) - 1]["id"])


def _render_attachment_download(attachment_row, label: str, key: str) -> bool:
    if not attachment_row:
        return False
    try:
        backend = storage_backend_for_name(attachment_row["storage_backend"])
        filename = clean_value(attachment_row["original_filename"], "arquivo")
        mime = clean_value(attachment_row["mime_type"], "application/octet-stream")
        if isinstance(backend, R2StorageBackend):
            url = backend.generate_download_url(attachment_row["storage_key"], filename)
            st.link_button(label, url)
            return True
        if isinstance(backend, LocalStorageBackend):
            st.download_button(
                label,
                data=backend.get_file_bytes(attachment_row["storage_key"]),
                file_name=filename,
                mime=mime,
                key=key,
            )
            return True
    except Exception as exc:
        st.warning(f"Não foi possível abrir o anexo persistido: {exc}")
    return False


def _format_file_size(size_value) -> str:
    try:
        size = int(size_value or 0)
    except Exception:
        return "-"
    if size <= 0:
        return "-"
    value = float(size)
    for unit in ["B", "KB", "MB", "GB"]:
        if value < 1024 or unit == "GB":
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{value:.1f} GB"


def _attachment_metadata_caption(attachment_row) -> str:
    details = [
        _format_file_size(attachment_row["file_size"] if "file_size" in attachment_row.keys() else None),
        _format_datetime(attachment_row["uploaded_at"] if "uploaded_at" in attachment_row.keys() else None),
    ]
    backend = clean_input(attachment_row["storage_backend"] if "storage_backend" in attachment_row.keys() else "")
    if backend:
        details.append(f"armazenado em {backend.upper() if backend == 'r2' else backend}")
    return " · ".join(detail for detail in details if detail and detail != "-")


def render_attachment_list(
    conn,
    *,
    entity_type: str,
    entity_id: int,
    attachment_role: str,
    legacy_path=None,
    key_prefix: str,
    title: str | None = "Anexos cadastrados",
    empty_message: str = "Nenhum anexo cadastrado.",
    attachment_rows: list | None = None,
) -> None:
    if title:
        st.markdown(f"##### {title}")
    rows = attachment_rows
    if rows is None:
        rows = list_attachments(
            conn,
            entity_type=entity_type,
            entity_id=int(entity_id),
            attachment_role=attachment_role,
        )
    rendered = False
    for row in rows:
        rendered = True
        filename = clean_value(row["original_filename"], "arquivo")
        st.caption(f"{filename} · {_attachment_metadata_caption(row)}")
        _render_attachment_download(row, "Baixar", f"{key_prefix}_attachment_{int(row['id'])}")

    listed_ids = {int(row["id"]) for row in rows}
    legacy_attachment_id = _attachment_id_from_ref(legacy_path)
    if legacy_attachment_id is not None and legacy_attachment_id not in listed_ids:
        legacy_attachment = get_attachment(conn, legacy_attachment_id)
        if legacy_attachment is not None:
            rendered = True
            filename = clean_value(legacy_attachment["original_filename"], "arquivo")
            st.caption(f"{filename} · {_attachment_metadata_caption(legacy_attachment)}")
            _render_attachment_download(legacy_attachment, "Baixar", f"{key_prefix}_legacy_attachment_{legacy_attachment_id}")

    if not rendered and not is_blank(legacy_path):
        rendered = True
        _download_or_link_document(conn, legacy_path, "Baixar anexo legado", f"{key_prefix}_legacy")

    if not rendered:
        st.caption(empty_message)


def _equipment_document_role_label(role: str | None) -> str:
    return EQUIPMENT_DOCUMENT_ROLE_LABELS.get(clean_input(role), clean_value(role, "Documento"))


def _equipment_document_rows(conn, equipment_id: int) -> list:
    rows = list_attachments(
        conn,
        entity_type="equipment",
        entity_id=int(equipment_id),
        active_only=True,
    )
    role_order = {role: idx for idx, role in enumerate(EQUIPMENT_DOCUMENT_ROLE_LABELS)}
    return sorted(
        rows,
        key=lambda row: (
            0 if clean_input(row["attachment_role"]) == "pop" else 1,
            role_order.get(clean_input(row["attachment_role"]), 999),
        ),
    )


def render_equipment_document_downloads(
    conn,
    equipment_row,
    *,
    key_prefix: str,
    show_empty: bool = False,
    can_deactivate: bool = False,
) -> int:
    rows = _equipment_document_rows(conn, int(equipment_row["id"]))
    if not rows:
        if show_empty:
            st.caption("Nenhum documento cadastrado via anexos.")
        return 0

    for row in rows:
        role_label = _equipment_document_role_label(row["attachment_role"])
        filename = clean_value(row["original_filename"], "arquivo")
        notes = clean_input(row["notes"] if "notes" in row.keys() else "")
        caption = f"{role_label}: {filename} · {_attachment_metadata_caption(row)}"
        if notes:
            caption += f" · {notes}"
        st.caption(caption)
        if can_deactivate:
            col_download, col_action = st.columns([4, 1])
            with col_download:
                _render_attachment_download(row, "Baixar", f"{key_prefix}_attachment_{int(row['id'])}")
            with col_action:
                if st.button("Inativar", key=f"{key_prefix}_deactivate_{int(row['id'])}"):
                    deactivate_attachment(conn, int(row["id"]))
                    st.success("Documento inativado.")
                    clear_app_caches()
                    st.rerun()
        else:
            _render_attachment_download(row, "Baixar", f"{key_prefix}_attachment_{int(row['id'])}")
    return len(rows)


def render_equipment_documents_section(
    conn,
    equipment_row,
    *,
    key_prefix: str,
    can_manage: bool,
) -> None:
    equipment_id = int(equipment_row["id"])
    with st.expander("Documentos do equipamento", expanded=False):
        legacy_path = clean_input(equipment_row.get("pop_path"))
        if legacy_path:
            st.markdown("##### POP legado")
            _download_or_link_document(conn, legacy_path, "Baixar/abrir POP legado", f"{key_prefix}_legacy_pop")
            st.caption(
                f"{clean_value(equipment_row.get('pop_title'), 'POP do equipamento')} · "
                f"Versão: {clean_value(equipment_row.get('pop_version'))} · "
                f"Responsável: {clean_value(equipment_row.get('pop_responsible'))}"
            )

        st.markdown("##### Documentos cadastrados")
        render_equipment_document_downloads(
            conn,
            equipment_row,
            key_prefix=f"{key_prefix}_docs",
            show_empty=True,
            can_deactivate=can_manage,
        )

        if not can_manage:
            st.caption("Envio e inativação de documentos ficam disponíveis para Gerente ou Administrador.")
            return

        st.markdown("##### Enviar documento")
        role_label = st.selectbox(
            "Tipo do documento",
            list(EQUIPMENT_DOCUMENT_ROLE_LABELS.values()),
            key=f"{key_prefix}_role",
        )
        upload = st.file_uploader("Arquivo", key=f"{key_prefix}_upload")
        notes = st.text_area("Observação", value="", key=f"{key_prefix}_notes")
        if st.button("Salvar documento", type="primary", key=f"{key_prefix}_save"):
            if upload is None:
                st.error("Selecione um arquivo para enviar.")
            elif _ensure_storage_ready_for_upload(upload):
                role = EQUIPMENT_DOCUMENT_ROLE_REVERSE[role_label]
                _save_upload(
                    conn,
                    upload,
                    entity_type="equipment",
                    entity_id=equipment_id,
                    attachment_role=role,
                    notes=notes.strip() or None,
                )
                st.success("Documento do equipamento cadastrado.")
                clear_app_caches()
                st.rerun()


def render_supply_lots_section(conn, supply_row: pd.Series, supply_lots: pd.DataFrame) -> None:
    supply_id = int(supply_row["id"])
    supply_unit = clean_input(supply_row.get("unit")) or "kg"
    supply_location = clean_input(supply_row.get("location"))
    active_lots = _supply_lots_for_supply(supply_lots, supply_id, active_only=True)

    st.markdown("#### Lotes do insumo")
    if active_lots.empty:
        st.caption("Nenhum lote ativo cadastrado.")
    else:
        st.dataframe(_lot_display_df(active_lots), use_container_width=True, hide_index=True)
        with st.expander("Certificados de análise", expanded=False):
            for _, lot in active_lots.iterrows():
                st.markdown(f"##### Lote {clean_value(lot.get('lot_code'))}")
                render_attachment_list(
                    conn,
                    entity_type="supply_lot",
                    entity_id=int(lot["id"]),
                    attachment_role="analysis_certificate",
                    legacy_path=lot.get("certificate_path"),
                    key_prefix=f"supply_lot_{int(lot['id'])}_certificate",
                    title=None,
                    empty_message="Nenhum certificado de análise cadastrado.",
                )

    if not can_manage_master_data():
        st.caption("Cadastro e edição de lotes ficam disponíveis para Gerente ou Administrador.")
        return

    with st.expander("Criar lote", expanded=False):
        with st.form(f"form_create_lot_{supply_id}"):
            lc1, lc2, lc3 = st.columns(3)
            with lc1:
                lot_code = st.text_input("Código do lote *", key=f"lot_code_new_{supply_id}")
                expiration_date = st.date_input("Validade", value=None, key=f"lot_exp_new_{supply_id}")
                received_date = st.date_input("Recebido em", value=date.today(), key=f"lot_received_new_{supply_id}")
            with lc2:
                supplier_name = st.text_input("Fornecedor", key=f"lot_supplier_new_{supply_id}")
                initial_quantity = st.number_input(
                    "Quantidade inicial",
                    min_value=0.0,
                    value=0.0,
                    step=1.0,
                    key=f"lot_initial_new_{supply_id}",
                    help="Se informada, uma entrada de estoque será registrada para manter rastreabilidade.",
                )
                unit = st.text_input("Unidade", value=supply_unit, key=f"lot_unit_new_{supply_id}")
            with lc3:
                location = st.text_input("Localização", value=supply_location, key=f"lot_location_new_{supply_id}")
                certificate_upload = st.file_uploader(
                    "Certificado de análise",
                    type=["pdf", "png", "jpg", "jpeg", "xlsx"],
                    key=f"lot_certificate_new_{supply_id}",
                )
                notes = st.text_area("Observações", key=f"lot_notes_new_{supply_id}")
            create_lot_submitted = st.form_submit_button("Criar lote", type="primary")

        if create_lot_submitted:
            if not lot_code.strip():
                st.error("Informe o código do lote.")
            elif not _ensure_storage_ready_for_upload(certificate_upload):
                pass
            else:
                try:
                    lot_id = create_supply_lot(
                        conn,
                        supply_id=supply_id,
                        lot_code=lot_code.strip(),
                        expiration_date=expiration_date.isoformat() if expiration_date else None,
                        received_date=received_date.isoformat() if received_date else None,
                        supplier_name=supplier_name.strip() or None,
                        initial_quantity=float(initial_quantity),
                        current_quantity=0.0,
                        unit=unit.strip() or supply_unit,
                        location=location.strip() or None,
                        notes=notes.strip() or None,
                    )
                    if initial_quantity:
                        ok, msg, _ = create_supply_movement(
                            conn,
                            supply_id=supply_id,
                            supply_lot_id=lot_id,
                            movement_type="entrada",
                            movement_date=(received_date or date.today()).isoformat(),
                            quantity=float(initial_quantity),
                            user_id=_current_user_id(),
                            project_id=None,
                            service_id=None,
                            purpose="Saldo inicial do lote.",
                            document_path=None,
                        )
                        if not ok:
                            st.error(msg)
                            clear_app_caches()
                            st.rerun()
                    if certificate_upload is not None:
                        certificate_ref = _save_upload(
                            conn,
                            certificate_upload,
                            entity_type="supply_lot",
                            entity_id=lot_id,
                            attachment_role="analysis_certificate",
                        )
                        update_legacy_attachment_path(
                            conn,
                            table="supply_lots",
                            row_id=lot_id,
                            column="certificate_path",
                            value=certificate_ref,
                        )
                    st.success("Lote cadastrado com sucesso.")
                    clear_app_caches()
                    st.rerun()
                except ValueError as exc:
                    st.error(str(exc))

    if active_lots.empty:
        return

    with st.expander("Editar lote", expanded=False):
        edit_label = st.selectbox(
            "Lote para editar",
            _supply_lot_options(active_lots),
            key=f"lot_edit_select_{supply_id}",
        )
        edit_lot = active_lots[active_lots["id"].astype(int) == int(_supply_lot_id_from_label(active_lots, edit_label))].iloc[0]
        with st.form(f"form_edit_lot_{int(edit_lot['id'])}"):
            le1, le2, le3 = st.columns(3)
            with le1:
                edit_lot_code = st.text_input("Código do lote *", value=clean_input(edit_lot.get("lot_code")), key=f"lot_code_edit_{int(edit_lot['id'])}")
                edit_expiration_date = st.date_input(
                    "Validade",
                    value=_date_input_value(edit_lot.get("expiration_date")),
                    key=f"lot_exp_edit_{int(edit_lot['id'])}",
                )
                edit_received_date = st.date_input(
                    "Recebido em",
                    value=_date_input_value(edit_lot.get("received_date")),
                    key=f"lot_received_edit_{int(edit_lot['id'])}",
                )
            with le2:
                edit_supplier_name = st.text_input("Fornecedor", value=clean_input(edit_lot.get("supplier_name")), key=f"lot_supplier_edit_{int(edit_lot['id'])}")
                edit_initial_quantity = float(edit_lot.get("initial_quantity") or 0)
                st.caption(f"Quantidade inicial registrada: {edit_initial_quantity:g} {clean_value(edit_lot.get('unit'), supply_unit)}")
                edit_unit = st.text_input("Unidade", value=clean_input(edit_lot.get("unit")) or supply_unit, key=f"lot_unit_edit_{int(edit_lot['id'])}")
            with le3:
                edit_location = st.text_input("Localização", value=clean_input(edit_lot.get("location")), key=f"lot_location_edit_{int(edit_lot['id'])}")
                edit_certificate_upload = st.file_uploader(
                    "Novo certificado de análise",
                    type=["pdf", "png", "jpg", "jpeg", "xlsx"],
                    key=f"lot_certificate_edit_{int(edit_lot['id'])}",
                )
                edit_notes = st.text_area("Observações", value=clean_input(edit_lot.get("notes")), key=f"lot_notes_edit_{int(edit_lot['id'])}")
            edit_lot_submitted = st.form_submit_button("Salvar lote", type="primary")

        if edit_lot_submitted:
            if not edit_lot_code.strip():
                st.error("Informe o código do lote.")
            elif not _ensure_storage_ready_for_upload(edit_certificate_upload):
                pass
            else:
                try:
                    certificate_final = clean_input(edit_lot.get("certificate_path")) or None
                    if edit_certificate_upload is not None:
                        certificate_final = _save_upload(
                            conn,
                            edit_certificate_upload,
                            entity_type="supply_lot",
                            entity_id=int(edit_lot["id"]),
                            attachment_role="analysis_certificate",
                        )
                    update_supply_lot(
                        conn,
                        int(edit_lot["id"]),
                        lot_code=edit_lot_code.strip(),
                        expiration_date=edit_expiration_date.isoformat() if edit_expiration_date else None,
                        received_date=edit_received_date.isoformat() if edit_received_date else None,
                        supplier_name=edit_supplier_name.strip() or None,
                        initial_quantity=float(edit_initial_quantity),
                        current_quantity=float(edit_lot.get("current_quantity") or 0),
                        unit=edit_unit.strip() or supply_unit,
                        location=edit_location.strip() or None,
                        certificate_path=certificate_final,
                        notes=edit_notes.strip() or None,
                        is_active=int(edit_lot.get("is_active") if not is_blank(edit_lot.get("is_active")) else 1),
                    )
                    st.success("Lote atualizado com sucesso.")
                    clear_app_caches()
                    st.rerun()
                except ValueError as exc:
                    st.error(str(exc))

    with st.expander("Inativar lote", expanded=False):
        inactive_label = st.selectbox(
            "Lote para inativar",
            _supply_lot_options(active_lots),
            key=f"lot_inactivate_select_{supply_id}",
        )
        inactive_lot_id = _supply_lot_id_from_label(active_lots, inactive_label)
        confirm_inactivate = st.checkbox(
            "Confirmo que este lote deve ficar indisponível para novas movimentações.",
            key=f"lot_inactivate_confirm_{supply_id}",
        )
        if st.button("Inativar lote", key=f"lot_inactivate_button_{supply_id}"):
            if not confirm_inactivate:
                st.error("Confirme a inativação antes de continuar.")
            else:
                ok, msg = inactivate_supply_lot(conn, int(inactive_lot_id))
                (st.success if ok else st.error)(msg)
                if ok:
                    clear_app_caches()
                    st.rerun()


def _download_or_link_document(
    conn,
    path_value,
    label: str,
    key: str,
    *,
    entity_type: str | None = None,
    entity_id: int | None = None,
    attachment_role: str | None = None,
) -> None:
    if entity_type and entity_id is not None:
        latest = get_latest_attachment_for_entity(
            conn,
            entity_type=entity_type,
            entity_id=int(entity_id),
            attachment_role=attachment_role,
        )
        if _render_attachment_download(latest, label, key):
            return

    attachment_id = _attachment_id_from_ref(path_value)
    if attachment_id is not None and _render_attachment_download(get_attachment(conn, attachment_id), label, key):
        return

    path_text = clean_input(path_value)
    if not path_text:
        return
    if path_text.lower().startswith(("http://", "https://")):
        st.link_button(label, path_text)
        return
    doc_path = _resolve_local_doc(path_text)
    if doc_path:
        mime = "application/pdf" if doc_path.suffix.lower() == ".pdf" else "application/octet-stream"
        st.download_button(label, data=doc_path.read_bytes(), file_name=doc_path.name, mime=mime, key=key)


def render_supply_quick_card(conn, supply_row: pd.Series) -> None:
    alert = _supply_alert_status(supply_row)
    qty = 0.0 if is_blank(supply_row.get("current_quantity")) else float(supply_row.get("current_quantity"))
    min_qty = 0.0 if is_blank(supply_row.get("minimum_quantity")) else float(supply_row.get("minimum_quantity"))
    type_line = f"Tipo: {clean_value(supply_row.get('supply_type'), 'Insumo')} · "
    if _is_spare_part(supply_row):
        type_line += (
            f"Código interno: {clean_value(supply_row.get('supply_code'))} · "
            f"Código fabricante: {clean_value(supply_row.get('manufacturer_code'))}<br>"
            f"Modelo/família compatível: {clean_value(supply_row.get('compatible_model_family'))}<br>"
        )
    st.markdown(
        f"""
        <div class="soft-card">
        <b>{clean_value(supply_row.get('supply_name'))}</b><br>
        {type_line}
        Categoria: {clean_value(supply_row.get('category'))} · Estado: {clean_value(supply_row.get('physical_state'))} ·
        Fabricante: {clean_value(supply_row.get('manufacturer'))}<br>
        Saldo: <b>{qty:g} {clean_value(supply_row.get('unit'), '')}</b> · Estoque mínimo: {min_qty:g} {clean_value(supply_row.get('unit'), '')}<br>
        Lote: {clean_value(supply_row.get('lot'))} · Validade: {_format_datetime(supply_row.get('expiration_date'))} ·
        Localização: {clean_value(supply_row.get('location'))}<br>
        Responsável: {clean_value(supply_row.get('responsible_name'))} · Status: <b>{alert}</b>
        </div>
        """,
        unsafe_allow_html=True,
    )
    c1, c2 = st.columns(2)
    with c1:
        _download_or_link_document(
            conn,
            supply_row.get("safety_doc_path"),
            "📄 Baixar/abrir FDS/FISPQ",
            f"supply_safety_qr_{int(supply_row['id'])}",
            entity_type="supply",
            entity_id=int(supply_row["id"]),
            attachment_role="safety_doc",
        )
    with c2:
        _download_or_link_document(
            conn,
            supply_row.get("technical_doc_path"),
            "📎 Baixar/abrir ficha técnica",
            f"supply_technical_qr_{int(supply_row['id'])}",
            entity_type="supply",
            entity_id=int(supply_row["id"]),
            attachment_role="technical_doc",
        )


def _project_id_from_label(projects: pd.DataFrame, label: str) -> int | None:
    options = _project_options(projects)
    if label == "Sem projeto específico" or projects.empty:
        return None
    return int(projects.iloc[options.index(label) - 1]["id"])


def _user_id_from_label(users: pd.DataFrame, label: str) -> int | None:
    options = ["Não informado"] + _user_options(users)
    if label == "Não informado" or users.empty:
        return None
    return int(users.iloc[options.index(label) - 1]["id"])


def _supply_alert_status(row: pd.Series) -> str:
    qty = float(row.get("current_quantity") or 0)
    min_qty = float(row.get("minimum_quantity") or 0)
    if min_qty and qty < min_qty:
        return "Estoque baixo"
    exp = row.get("expiration_date")
    if not is_blank(exp):
        try:
            exp_date = datetime.fromisoformat(str(exp)).date()
            today = date.today()
            if exp_date < today:
                return "Vencido"
            if exp_date <= today + timedelta(days=60):
                return "Vence em até 60 dias"
        except Exception:
            pass
    if not _is_spare_part(row) and is_blank(row.get("safety_doc_path")):
        return "Sem FDS/FISPQ"
    return "OK"


def page_insumos(conn):
    hero()
    st.subheader("Insumos e almoxarifado")
    st.caption("Controle simples de estoque: cadastro mínimo, saldo atual, lote, validade, localização, documentos e histórico de movimentações.")

    database_url = _database_url()
    with perf_timer("Dados de insumos"):
        supplies, supply_lots, users, projects, equipment = _cached_supply_page_data(
            _database_fingerprint(database_url),
            _database_url_value=database_url,
        )

    qr_supply_id = st.query_params.get("sid", None)
    if qr_supply_id and not supplies.empty:
        try:
            qr_supply_id_int = int(qr_supply_id)
            qr_supply = supplies[supplies["id"].astype(int) == qr_supply_id_int]
        except Exception:
            qr_supply = pd.DataFrame()
        if not qr_supply.empty:
            st.markdown("### Ficha rápida do insumo")
            render_supply_quick_card(conn, qr_supply.iloc[0])
            st.info("QR de insumo detectado. Use a aba **Movimentar estoque** para registrar entrada, saída, descarte ou ajuste.")
        else:
            st.warning("QR de insumo detectado, mas o insumo não foi encontrado no banco atual.")

    tab_visao, tab_cadastro, tab_mov, tab_hist = st.tabs([
        "Visão geral",
        "Cadastrar/editar insumo",
        "Movimentar estoque",
        "Histórico",
    ])

    with tab_visao:
        st.markdown("### Visão geral do almoxarifado")
        active_supplies = supplies[supplies["active"] == 1].copy() if not supplies.empty else supplies
        if supplies.empty:
            st.info("Nenhum insumo cadastrado ainda.")
        else:
            active_supplies["alerta"] = active_supplies.apply(_supply_alert_status, axis=1)
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Itens ativos", int((supplies["active"] == 1).sum()))
            k2.metric("Estoque baixo", int((active_supplies["alerta"] == "Estoque baixo").sum()))
            k3.metric("Vencidos", int((active_supplies["alerta"] == "Vencido").sum()))
            k4.metric("Sem FDS/FISPQ", int((active_supplies["alerta"] == "Sem FDS/FISPQ").sum()))

            alert_df = active_supplies[active_supplies["alerta"] != "OK"].copy()
            if not alert_df.empty:
                st.markdown("#### Alertas")
                st.dataframe(
                    _display_df(alert_df[[
                        "alerta", "supply_type", "supply_name", "supply_code", "category",
                        "current_quantity", "unit", "minimum_quantity", "lot",
                        "expiration_date", "location", "responsible_name"
                    ]]),
                    use_container_width=True,
                    hide_index=True,
                )
            st.markdown("#### Estoque atual")
            cols = [
                "supply_type", "supply_name", "supply_code", "manufacturer_code",
                "commercial_name", "manufacturer", "category", "physical_state",
                "compatible_model_family", "current_quantity", "unit", "minimum_quantity",
                "lot", "expiration_date", "location", "responsible_name", "active"
            ]
            st.dataframe(_display_df(supplies[[c for c in cols if c in supplies.columns]]), use_container_width=True, hide_index=True)
            if can_export_inventory():
                st.download_button(
                    "Baixar estoque em CSV",
                    data=_display_df(supplies).to_csv(index=False).encode("utf-8-sig"),
                    file_name="labcim_insumos_estoque.csv",
                    mime="text/csv",
                )
            else:
                st.caption("Exportação completa do estoque é restrita a Gerente ou Administrador.")

    with tab_cadastro:
        st.markdown("### Cadastro de item de estoque")
        if not can_manage_master_data():
            st.info("Cadastro, edição estrutural, lotes, certificados e associação de peças são restritos a Gerente ou Administrador.")
        else:
            mode = st.radio("Modo", ["Novo item", "Editar item existente"], horizontal=True, key="supply_edit_mode")
            selected_supply = None
            if mode == "Editar item existente":
                if supplies.empty:
                    st.info("Cadastre um item de estoque antes de editar.")
                    return
                label = st.selectbox("Selecionar item", _supply_options(supplies), key="supply_edit_select")
                selected_supply = supplies[supplies["id"] == _supply_id_from_label(supplies, label)].iloc[0]

            current_supply_type = _supply_type_value(selected_supply)
            supply_type_key = f"supply_type_{mode}_{int(selected_supply['id']) if selected_supply is not None else 'new'}"
            supply_type = st.radio(
                "Tipo de item",
                SUPPLY_TYPES,
                index=SUPPLY_TYPES.index(current_supply_type),
                horizontal=True,
                key=supply_type_key,
            )
            is_spare_part = supply_type == "Peça de reposição"
            selected_equipment_ids: list[int] = []
            current_item_is_spare_part = _is_spare_part(selected_supply)

            with st.form("form_supply"):
                c1, c2, c3 = st.columns(3)
                with c1:
                    supply_name = st.text_input(
                        "Nome da peça *" if is_spare_part else "Nome do insumo *",
                        value=clean_input(selected_supply.get("supply_name")) if selected_supply is not None else "",
                        placeholder="Ex.: rotor, vedação, filtro" if is_spare_part else "Ex.: Cimento Portland Classe G",
                    )
                    supply_code = clean_input(selected_supply.get("supply_code")) if selected_supply is not None else ""
                    manufacturer_code = clean_input(selected_supply.get("manufacturer_code")) if selected_supply is not None else ""
                    compatible_model_family = clean_input(selected_supply.get("compatible_model_family")) if selected_supply is not None else ""
                    if is_spare_part:
                        supply_code = st.text_input("Código interno", value=supply_code, placeholder="Ex.: PR-0001")
                    commercial_name = st.text_input("Nome comercial", value=clean_input(selected_supply.get("commercial_name")) if selected_supply is not None else "")
                    manufacturer = st.text_input("Fabricante", value=clean_input(selected_supply.get("manufacturer")) if selected_supply is not None else "")
                    if is_spare_part:
                        manufacturer_code = st.text_input("Código do fabricante", value=manufacturer_code, placeholder="Ex.: part number, SKU ou referência do fabricante")
                        category = st.text_input(
                            "Categoria",
                            value=clean_input(selected_supply.get("category")) if selected_supply is not None else "",
                            placeholder="Ex.: filtro, vedação, sensor, placa eletrônica",
                        )
                    else:
                        category = st.selectbox(
                            "Categoria",
                            ["Cimento", "Aditivo", "Sal", "Polímero", "Pozolana", "Carga mineral", "Lavador/espaçador", "Reagente", "Consumível", "Outro"],
                            index=0,
                            key="supply_category",
                        )
                        if selected_supply is not None and clean_input(selected_supply.get("category")):
                            category = st.text_input("Categoria cadastrada", value=clean_input(selected_supply.get("category")))
                with c2:
                    if is_spare_part:
                        physical_state = clean_input(selected_supply.get("physical_state")) if current_item_is_spare_part and selected_supply is not None else ""
                        physical_state = physical_state or "Não se aplica"
                        application_function = clean_input(selected_supply.get("application_function")) if current_item_is_spare_part and selected_supply is not None else ""
                        addition_mode = clean_input(selected_supply.get("addition_mode")) if current_item_is_spare_part and selected_supply is not None else ""
                        addition_mode = addition_mode or "Não se aplica"
                        compatible_model_family = st.text_input("Modelo/família compatível", value=compatible_model_family, placeholder="Ex.: Reômetro modelo X, Autoclave série Y")
                        unit = clean_input(selected_supply.get("unit")) if current_item_is_spare_part and selected_supply is not None else ""
                        unit = unit or "unidade"
                    else:
                        physical_state = st.selectbox("Estado físico", ["Sólido", "Líquido", "Gás", "Pasta/suspensão", "Outro"], key="supply_state")
                        if selected_supply is not None and clean_input(selected_supply.get("physical_state")):
                            physical_state = st.text_input("Estado físico cadastrado", value=clean_input(selected_supply.get("physical_state")))
                        application_function = st.text_input("Função/aplicação", value=clean_input(selected_supply.get("application_function")) if selected_supply is not None else "", placeholder="Ex.: retardador, expansivo, salmoura, cimento base")
                        addition_mode = st.selectbox("Modo de adição", ["Não se aplica", "Misturado a seco", "Água de mistura", "Solução", "Outro"], key="supply_addition")
                        if selected_supply is not None and clean_input(selected_supply.get("addition_mode")):
                            addition_mode = st.text_input("Modo de adição cadastrado", value=clean_input(selected_supply.get("addition_mode")))
                        unit = st.selectbox("Unidade de controle", ["kg", "g", "L", "mL", "unidade", "frasco", "saco"], key="supply_unit")
                        if selected_supply is not None and clean_input(selected_supply.get("unit")):
                            unit = st.text_input("Unidade cadastrada", value=clean_input(selected_supply.get("unit")))
                with c3:
                    initial_qty_default = float(selected_supply.get("current_quantity") or 0) if selected_supply is not None else 0.0
                    current_quantity = st.number_input("Saldo inicial/atual", min_value=0.0, value=initial_qty_default, step=1.0, disabled=(mode == "Editar item existente"), help="Depois do cadastro, o saldo deve ser alterado por movimentações.")
                    min_qty_default = float(selected_supply.get("minimum_quantity") or 0) if selected_supply is not None else 0.0
                    minimum_quantity = st.number_input("Estoque mínimo", min_value=0.0, value=min_qty_default, step=1.0)
                    lot = st.text_input("Lote", value=clean_input(selected_supply.get("lot")) if selected_supply is not None else "")
                    expiration_date = _date_input_value(selected_supply.get("expiration_date")) if is_spare_part and selected_supply is not None else None
                    if not is_spare_part:
                        expiration_date = st.date_input(
                            "Validade",
                            value=None if selected_supply is None or is_blank(selected_supply.get("expiration_date")) else datetime.fromisoformat(str(selected_supply.get("expiration_date"))).date(),
                            key="supply_expiration",
                        )
                    location = st.text_input("Localização", value=clean_input(selected_supply.get("location")) if selected_supply is not None else "", placeholder="Ex.: Almoxarifado 1, armário A")
                    responsible_name = st.text_input("Responsável", value=clean_input(selected_supply.get("responsible_name")) if selected_supply is not None else "")

                density = selected_supply.get("density") if selected_supply is not None else None
                recommended_concentration = clean_input(selected_supply.get("recommended_concentration")) if selected_supply is not None else ""
                recommended_temperature = clean_input(selected_supply.get("recommended_temperature")) if selected_supply is not None else ""
                characterization_summary = clean_input(selected_supply.get("characterization_summary")) if selected_supply is not None else ""
                safety_doc_path = clean_input(selected_supply.get("safety_doc_path")) if selected_supply is not None else ""
                technical_doc_path = clean_input(selected_supply.get("technical_doc_path")) if selected_supply is not None else ""
                safety_upload = None
                technical_upload = None
                if not is_spare_part:
                    st.markdown("#### Dados técnicos opcionais")
                    t1, t2, t3 = st.columns(3)
                    with t1:
                        density_default = float(selected_supply.get("density") or 0) if selected_supply is not None and not is_blank(selected_supply.get("density")) else 0.0
                        density = st.number_input("Massa específica", min_value=0.0, value=density_default, step=0.01)
                        recommended_concentration = st.text_input("Faixa de concentração", value=recommended_concentration, placeholder="Ex.: 0,5–3,0% BWOC")
                    with t2:
                        recommended_temperature = st.text_input("Faixa de temperatura", value=recommended_temperature, placeholder="Ex.: 25–90 °C")
                        characterization_summary = st.text_area("Caracterização resumida", value=characterization_summary, placeholder="Ex.: FRX/DRX realizados; arquivo anexado...")
                    with t3:
                        safety_doc_path = st.text_input("FDS/FISPQ existente ou link", value=safety_doc_path)
                        technical_doc_path = st.text_input("Ficha técnica/caracterização existente ou link", value=technical_doc_path)
                        safety_upload = st.file_uploader("Anexar FDS/FISPQ", type=["pdf", "png", "jpg", "jpeg"], key="safety_doc_upload")
                        technical_upload = st.file_uploader("Anexar ficha/caracterização", type=["pdf", "png", "jpg", "jpeg", "xlsx"], key="technical_doc_upload")

                if is_spare_part:
                    st.markdown("#### Equipamentos associados")
                    if equipment.empty:
                        st.info("Cadastre equipamentos antes de associar peças de reposição.")
                    else:
                        linked_equipment = (
                            list_equipment_for_spare_part(conn, int(selected_supply["id"]))
                            if selected_supply is not None
                            else pd.DataFrame()
                        )
                        linked_ids = set(linked_equipment["id"].astype(int).tolist()) if not linked_equipment.empty else set()
                        equipment_options = _equipment_options(equipment)
                        default_equipment_labels = [
                            label for label in equipment_options
                            if _equipment_id_from_label(equipment, label) in linked_ids
                        ]
                        selected_equipment_labels = st.multiselect(
                            "Equipamentos associados",
                            equipment_options,
                            default=default_equipment_labels,
                            key=f"spare_equipment_links_{int(selected_supply['id']) if selected_supply is not None else 'new'}",
                        )
                        selected_equipment_ids = _equipment_ids_from_labels(equipment, selected_equipment_labels)

                notes = st.text_area("Observações", value=clean_input(selected_supply.get("notes")) if selected_supply is not None else "")
                active = st.checkbox("Item ativo", value=True if selected_supply is None else truthy(selected_supply.get("active")))
                submitted = st.form_submit_button("Salvar insumo", type="primary")

            if selected_supply is not None and not is_spare_part:
                st.markdown("#### Documentos cadastrados")
                a1, a2 = st.columns(2)
                with a1:
                    render_attachment_list(
                        conn,
                        entity_type="supply",
                        entity_id=int(selected_supply["id"]),
                        attachment_role="safety_doc",
                        legacy_path=selected_supply.get("safety_doc_path"),
                        key_prefix=f"supply_{int(selected_supply['id'])}_safety_doc",
                        title="FDS/FISPQ",
                        empty_message="Nenhuma FDS/FISPQ cadastrada.",
                    )
                with a2:
                    render_attachment_list(
                        conn,
                        entity_type="supply",
                        entity_id=int(selected_supply["id"]),
                        attachment_role="technical_doc",
                        legacy_path=selected_supply.get("technical_doc_path"),
                        key_prefix=f"supply_{int(selected_supply['id'])}_technical_doc",
                        title="Ficha técnica/caracterização",
                        empty_message="Nenhuma ficha técnica/caracterização cadastrada.",
                    )

            if selected_supply is not None:
                render_supply_lots_section(conn, selected_supply, supply_lots)

            if submitted:
                if not can_manage_master_data():
                    st.error("Cadastro/edição estrutural de insumos exige perfil Gerente ou Administrador.")
                elif not supply_name.strip():
                    st.error("Informe o nome do item.")
                elif not _ensure_storage_ready_for_upload(safety_upload, technical_upload):
                    pass
                else:
                    if mode == "Novo item":
                        supply_id = create_supply(
                            conn,
                            supply_type=supply_type,
                            supply_name=supply_name.strip(),
                            supply_code=supply_code.strip() or None,
                            commercial_name=commercial_name.strip() or None,
                            manufacturer=manufacturer.strip() or None,
                            manufacturer_code=manufacturer_code.strip() or None,
                            category=category.strip() or None,
                            physical_state=physical_state.strip() or None,
                            application_function=application_function.strip() or None,
                            addition_mode=addition_mode.strip() or None,
                            compatible_model_family=compatible_model_family.strip() or None,
                            unit=unit.strip() or "kg",
                            current_quantity=0.0,
                            minimum_quantity=float(minimum_quantity),
                            lot=lot.strip() or None,
                            expiration_date=expiration_date.isoformat() if expiration_date else None,
                            location=location.strip() or None,
                            responsible_name=responsible_name.strip() or None,
                            safety_doc_path=safety_doc_path.strip() or None,
                            technical_doc_path=technical_doc_path.strip() or None,
                            density=float(density) if density else None,
                            recommended_concentration=recommended_concentration.strip() or None,
                            recommended_temperature=recommended_temperature.strip() or None,
                            characterization_summary=characterization_summary.strip() or None,
                            notes=notes.strip() or None,
                        )
                        if is_spare_part:
                            set_spare_part_equipment_links(
                                conn,
                                supply_id=supply_id,
                                equipment_ids=selected_equipment_ids,
                            )
                        if current_quantity:
                            create_supply_movement(
                                conn,
                                supply_id=supply_id,
                                movement_type="entrada",
                                movement_date=date.today().isoformat(),
                                quantity=float(current_quantity),
                                user_id=None,
                                project_id=None,
                                purpose="Saldo inicial cadastrado.",
                                document_path=None,
                            )
                        if safety_upload is not None:
                            safety_ref = _save_upload(
                                conn,
                                safety_upload,
                                entity_type="supply",
                                entity_id=supply_id,
                                attachment_role="safety_doc",
                            )
                            update_legacy_attachment_path(
                                conn,
                                table="supplies",
                                row_id=supply_id,
                                column="safety_doc_path",
                                value=safety_ref,
                            )
                        if technical_upload is not None:
                            technical_ref = _save_upload(
                                conn,
                                technical_upload,
                                entity_type="supply",
                                entity_id=supply_id,
                                attachment_role="technical_doc",
                            )
                            update_legacy_attachment_path(
                                conn,
                                table="supplies",
                                row_id=supply_id,
                                column="technical_doc_path",
                                value=technical_ref,
                            )
                        st.success("Item cadastrado com sucesso.")
                        clear_app_caches()
                        st.rerun()
                    else:
                        supply_id = int(selected_supply["id"])
                        safety_final = safety_doc_path.strip() or None
                        technical_final = technical_doc_path.strip() or None
                        if safety_upload is not None:
                            safety_final = _save_upload(
                                conn,
                                safety_upload,
                                entity_type="supply",
                                entity_id=supply_id,
                                attachment_role="safety_doc",
                            )
                        if technical_upload is not None:
                            technical_final = _save_upload(
                                conn,
                                technical_upload,
                                entity_type="supply",
                                entity_id=supply_id,
                                attachment_role="technical_doc",
                            )
                        update_supply(
                            conn,
                            supply_id,
                            supply_type=supply_type,
                            supply_name=supply_name.strip(),
                            supply_code=supply_code.strip() or None,
                            commercial_name=commercial_name.strip() or None,
                            manufacturer=manufacturer.strip() or None,
                            manufacturer_code=manufacturer_code.strip() or None,
                            category=category.strip() or None,
                            physical_state=physical_state.strip() or None,
                            application_function=application_function.strip() or None,
                            addition_mode=addition_mode.strip() or None,
                            compatible_model_family=compatible_model_family.strip() or None,
                            unit=unit.strip() or "kg",
                            minimum_quantity=float(minimum_quantity),
                            lot=lot.strip() or None,
                            expiration_date=expiration_date.isoformat() if expiration_date else None,
                            location=location.strip() or None,
                            responsible_name=responsible_name.strip() or None,
                            safety_doc_path=safety_final,
                            technical_doc_path=technical_final,
                            density=float(density) if density else None,
                            recommended_concentration=recommended_concentration.strip() or None,
                            recommended_temperature=recommended_temperature.strip() or None,
                            characterization_summary=characterization_summary.strip() or None,
                            active=int(active),
                            notes=notes.strip() or None,
                        )
                        set_spare_part_equipment_links(
                            conn,
                            supply_id=supply_id,
                            equipment_ids=selected_equipment_ids if is_spare_part else [],
                        )
                        st.success("Item atualizado com sucesso.")
                        clear_app_caches()
                        st.rerun()

    with tab_mov:
        st.markdown("### Movimentar estoque")
        active_supplies = supplies[supplies["active"] == 1].copy() if not supplies.empty else supplies
        if active_supplies.empty:
            st.info("Cadastre ao menos um item ativo para movimentar estoque.")
        else:
            sc1, sc2 = st.columns([1.35, 1])
            with sc1:
                supply_label = st.selectbox(
                    "Item de estoque",
                    _supply_options(active_supplies),
                    index=_select_index_by_supply_id(active_supplies, qr_supply_id),
                    key="movement_supply",
                )
                supply_id = _supply_id_from_label(active_supplies, supply_label)
                selected_movement_supply = active_supplies[active_supplies["id"].astype(int) == int(supply_id)].iloc[0]
            with sc2:
                active_lots = _supply_lots_for_supply(supply_lots, int(supply_id), active_only=True)
                selected_lot_id = None
                selected_lot = None
                if active_lots.empty:
                    st.caption("Este item não tem lotes ativos. A movimentação pode seguir sem lote.")
                else:
                    lot_label = st.selectbox(
                        "Lote (opcional)",
                        ["Sem lote específico"] + _supply_lot_options(active_lots),
                        key=f"movement_lot_{int(supply_id)}",
                    )
                    selected_lot_id = _supply_lot_id_from_label(active_lots, lot_label)
                    if selected_lot_id is not None:
                        selected_lot = active_lots[active_lots["id"].astype(int) == int(selected_lot_id)].iloc[0]

            if selected_lot is not None:
                lot_status = _lot_expiration_status(selected_lot.get("expiration_date"))
                lot_message = (
                    f"Lote {clean_value(selected_lot.get('lot_code'))} · "
                    f"saldo {float(selected_lot.get('current_quantity') or 0):g} {clean_value(selected_lot.get('unit'), clean_value(selected_movement_supply.get('unit'), ''))} · "
                    f"validade {_format_datetime(selected_lot.get('expiration_date'))} · "
                    f"fornecedor {clean_value(selected_lot.get('supplier_name'))} · "
                    f"local {clean_value(selected_lot.get('location'))}"
                )
                if lot_status == "Vencido":
                    st.error(f"{lot_message} · {lot_status}")
                elif lot_status == "Vence em até 60 dias":
                    st.warning(f"{lot_message} · {lot_status}")
                else:
                    st.caption(f"{lot_message} · {lot_status}")

            movement_project_id = None
            movement_service_id = None
            pc1, pc2 = st.columns(2)
            with pc1:
                project_label = st.selectbox("Projeto", _project_options(projects), key="movement_project")
                if project_label != "Sem projeto específico" and not projects.empty:
                    movement_project_id = _project_id_from_label(projects, project_label)
            with pc2:
                if movement_project_id is not None:
                    movement_services = _project_services_for_project(movement_project_id, active_only=True)
                    movement_service_label = st.selectbox(
                        "Serviço/análise",
                        _service_options(movement_services),
                        key=f"movement_service_{movement_project_id}",
                    )
                    movement_service_id = _service_id_from_label(movement_services, movement_service_label)
                else:
                    st.caption("Selecione um projeto para vincular um serviço/análise.")

            movement_type_options = ["entrada", "saída", "descarte", "ajuste positivo", "ajuste negativo"]
            if not can_manage_inventory_adjustments():
                movement_type_options = ["saída"]
                st.info("Seu perfil permite registrar apenas saída/consumo. Entradas, descartes e ajustes são restritos a Gerente ou Administrador.")

            movement_user_id = None
            with st.form("form_supply_movement"):
                c1, c2, c3 = st.columns(3)
                with c1:
                    movement_type = st.selectbox("Tipo de movimentação", movement_type_options)
                    movement_date = st.date_input("Data", value=date.today(), key="movement_date")
                with c2:
                    quantity = st.number_input("Quantidade", min_value=0.0, value=0.0, step=1.0, key="movement_qty")
                    if can_manage_inventory_adjustments():
                        user_label = st.selectbox("Responsável pela movimentação", ["Não informado"] + _user_options(users), key="movement_user")
                        movement_user_id = _user_id_from_label(users, user_label)
                    else:
                        movement_user_id = _current_user_id()
                        st.caption(f"Responsável pela movimentação: {clean_value(current_user().get('full_name'), 'usuário autenticado')}")
                with c3:
                    purpose = st.text_area("Finalidade/observação", placeholder="Ex.: preparo de pasta; recebimento de material; descarte por vencimento...")
                    movement_doc = st.file_uploader("Anexo da movimentação", type=["pdf", "png", "jpg", "jpeg", "xlsx"], key="movement_doc")
                move_submitted = st.form_submit_button("Registrar movimentação", type="primary")

            st.markdown("#### Ficha do item selecionado")
            fc1, fc2, fc3, fc4 = st.columns(4)
            fc1.metric("Saldo total", f"{float(selected_movement_supply.get('current_quantity') or 0):g} {clean_value(selected_movement_supply.get('unit'), '')}")
            fc2.metric("Estoque mínimo", f"{float(selected_movement_supply.get('minimum_quantity') or 0):g} {clean_value(selected_movement_supply.get('unit'), '')}")
            fc3.metric("Lotes ativos", len(active_lots))
            fc4.metric("Status", _supply_alert_status(selected_movement_supply))
            if active_lots.empty:
                st.caption("Nenhum lote ativo cadastrado para este item.")
            else:
                st.dataframe(_lot_display_df(active_lots), use_container_width=True, hide_index=True)

            supply_history = query_df(
                conn,
                """
                SELECT sm.id, sm.movement_type, sm.movement_date, sm.quantity,
                       COALESCE(sm.unit, s.unit) AS unit, sl.lot_code AS supply_lot_code,
                       u.full_name AS responsible_name, p.project_name,
                       ps.service_code, ps.title AS service_title, sm.purpose, sm.created_at
                FROM supply_movements sm
                JOIN supplies s ON s.id = sm.supply_id
                LEFT JOIN supply_lots sl ON sl.id = sm.supply_lot_id
                LEFT JOIN users u ON u.id = sm.user_id
                LEFT JOIN projects p ON p.id = sm.project_id
                LEFT JOIN project_services ps ON ps.id = sm.service_id
                WHERE sm.supply_id = ?
                ORDER BY sm.movement_date DESC, sm.id DESC
                LIMIT 20
                """,
                [int(supply_id)],
            )
            if supply_history.empty:
                st.caption("Nenhuma movimentação registrada para este item.")
            else:
                st.dataframe(_display_df(supply_history), use_container_width=True, hide_index=True)

            st.markdown("#### Documentos do item selecionado")
            d1, d2 = st.columns(2)
            with d1:
                render_attachment_list(
                    conn,
                    entity_type="supply",
                    entity_id=int(supply_id),
                    attachment_role="safety_doc",
                    legacy_path=selected_movement_supply.get("safety_doc_path"),
                    key_prefix=f"movement_supply_{int(supply_id)}_safety_doc",
                    title="FDS/FISPQ",
                    empty_message="Nenhuma FDS/FISPQ cadastrada.",
                )
            with d2:
                render_attachment_list(
                    conn,
                    entity_type="supply",
                    entity_id=int(supply_id),
                    attachment_role="technical_doc",
                    legacy_path=selected_movement_supply.get("technical_doc_path"),
                    key_prefix=f"movement_supply_{int(supply_id)}_technical_doc",
                    title="Ficha técnica/caracterização",
                    empty_message="Nenhuma ficha técnica/caracterização cadastrada.",
                )

            if move_submitted:
                negative_movement = movement_type in {"saída", "descarte", "ajuste negativo"}
                lot_balance = float(selected_lot.get("current_quantity") or 0) if selected_lot is not None else None
                if not can_manage_inventory_adjustments() and movement_user_id is None:
                    st.error("Não foi possível identificar o usuário autenticado. Faça login novamente.")
                elif selected_lot is not None and negative_movement and float(quantity) > float(lot_balance or 0) + 1e-9:
                    st.error(f"Saldo insuficiente no lote. Saldo atual do lote: {float(lot_balance or 0):g} {clean_value(selected_lot.get('unit'), clean_value(selected_movement_supply.get('unit'), ''))}.")
                elif _ensure_storage_ready_for_upload(movement_doc):
                    ok, msg, movement_id = create_supply_movement(
                        conn,
                        supply_id=supply_id,
                        supply_lot_id=selected_lot_id,
                        movement_type=movement_type,
                        movement_date=movement_date.isoformat(),
                        quantity=float(quantity),
                        user_id=movement_user_id,
                        project_id=movement_project_id,
                        service_id=movement_service_id,
                        purpose=purpose.strip() or None,
                        document_path=None,
                    )
                    if ok and movement_doc is not None and movement_id is not None:
                        doc_ref = _save_upload(
                            conn,
                            movement_doc,
                            entity_type="supply_movement",
                            entity_id=movement_id,
                            attachment_role="movement_document",
                        )
                        update_legacy_attachment_path(
                            conn,
                            table="supply_movements",
                            row_id=movement_id,
                            column="document_path",
                            value=doc_ref,
                        )
                    (st.success if ok else st.error)(msg)
                    if ok:
                        clear_app_caches()
                        st.rerun()

    with tab_hist:
        st.markdown("### Histórico de movimentações")
        hist = query_df(
            conn,
            """
            SELECT sm.id, s.supply_name, sl.lot_code AS supply_lot_code,
                   sm.movement_type, sm.movement_date, sm.quantity, sm.unit,
                   u.full_name AS responsible_name, p.project_name,
                   ps.service_code, ps.title AS service_title,
                   sm.purpose, sm.document_path, sm.created_at
            FROM supply_movements sm
            JOIN supplies s ON s.id = sm.supply_id
            LEFT JOIN supply_lots sl ON sl.id = sm.supply_lot_id
            LEFT JOIN users u ON u.id = sm.user_id
            LEFT JOIN projects p ON p.id = sm.project_id
            LEFT JOIN project_services ps ON ps.id = sm.service_id
            ORDER BY sm.movement_date DESC, sm.id DESC
            """,
        )
        if hist.empty:
            st.info("Ainda não há movimentações registradas.")
        else:
            st.dataframe(_display_df(hist), use_container_width=True, hide_index=True)
            st.markdown("#### Anexos cadastrados")
            movement_attachments = _attachment_rows_by_entity(
                conn,
                entity_type="supply_movement",
                attachment_role="movement_document",
                entity_ids=[int(value) for value in hist["id"].tolist() if not is_blank(value)],
            )
            shown_movements = 0
            for _, movement in hist.iterrows():
                movement_id = int(movement["id"])
                attachment_rows = movement_attachments.get(movement_id, [])
                legacy_path = movement.get("document_path")
                if not attachment_rows and is_blank(legacy_path):
                    continue
                shown_movements += 1
                with st.expander(f"Movimentação #{movement_id} · {clean_value(movement.get('supply_name'))}", expanded=False):
                    render_attachment_list(
                        conn,
                        entity_type="supply_movement",
                        entity_id=movement_id,
                        attachment_role="movement_document",
                        legacy_path=legacy_path,
                        key_prefix=f"supply_movement_{movement_id}",
                        title="Documento/anexo",
                        empty_message="Nenhum documento/anexo cadastrado.",
                        attachment_rows=attachment_rows,
                    )
            if shown_movements == 0:
                st.caption("Nenhum anexo cadastrado.")
            if can_export_inventory():
                st.download_button(
                    "Baixar histórico em CSV",
                    data=_display_df(hist).to_csv(index=False).encode("utf-8-sig"),
                    file_name="labcim_insumos_movimentacoes.csv",
                    mime="text/csv",
                )
            else:
                st.caption("Exportação do histórico de movimentações é restrita a Gerente ou Administrador.")


def page_manutencao(conn):
    hero()
    can_manage_maintenance = can_edit_operational_data()
    if can_manage_maintenance:
        st.subheader("Manutenção e suporte")
        st.caption("Controle preventivo/calibração e tickets corretivos com edição, histórico de status e inativação auditável.")
    else:
        st.subheader("Reportar problema em equipamento")
        st.caption("Use esta tela para registrar falhas, ruídos, quebras, mensagens de erro ou necessidade de suporte. A equipe do laboratório acompanhará o ticket.")
    equipment, users, _, _ = load_reference_data(conn)
    if equipment.empty:
        st.warning("Cadastre/importe equipamentos antes de registrar manutenções.")
        return
    qr_view = clean_input(st.query_params.get("view")).lower()
    qr_equipment_code = st.query_params.get("eq") if qr_view == "manutencao" else None
    qr_maintenance_equipment_id = None
    qr_maintenance_message = ""
    qr_maintenance_issue = ""
    if qr_equipment_code:
        qr_row = _equipment_row_by_code(equipment, qr_equipment_code)
        if qr_row is None:
            qr_maintenance_issue = "QR de manutenção aponta para equipamento não encontrado. Selecione o equipamento manualmente."
        else:
            qr_status = clean_input(qr_row.get("operational_status")).lower()
            if not truthy(qr_row.get("active")) or qr_status == "inactive":
                qr_maintenance_issue = (
                    "QR de manutenção aponta para equipamento inativo. "
                    f"{clean_value(qr_row.get('equipment_code'))} — {clean_value(qr_row.get('equipment_name'))} não será selecionado automaticamente."
                )
            else:
                qr_maintenance_equipment_id = int(qr_row["id"])
                qr_maintenance_message = "QR de manutenção aberto. Registre abaixo o problema observado neste equipamento."

    if not can_manage_maintenance:
        if qr_maintenance_message:
            st.info(qr_maintenance_message)
        elif qr_maintenance_issue:
            st.warning(qr_maintenance_issue)

        st.caption("Manutenção preventiva, calibração e edição de registros são gerenciadas por Gerente ou Administrador.")
        st.markdown("### Reportar problema")
        with st.container(border=True):
            c1, c2 = st.columns(2)
            with c1:
                equipment_options = _equipment_options(equipment)
                manual_placeholder = "Selecione um equipamento"
                corr_select_options = equipment_options
                corr_equipment_index = 0
                if qr_maintenance_equipment_id is not None:
                    equipment_ids = equipment["id"].astype(int).tolist()
                    if qr_maintenance_equipment_id in equipment_ids:
                        corr_equipment_index = equipment_ids.index(qr_maintenance_equipment_id)
                elif qr_maintenance_issue:
                    corr_select_options = [manual_placeholder] + equipment_options
                eq_label = st.selectbox("Equipamento", corr_select_options, index=corr_equipment_index, key="member_corr_eq")
                equipment_id = None
                selected = pd.Series(dtype=object)
                if eq_label == manual_placeholder:
                    st.info("Selecione manualmente um equipamento para abrir o ticket.")
                else:
                    equipment_id = _equipment_id_from_label(equipment, eq_label)
                    selected = equipment[equipment["id"] == equipment_id].iloc[0]
                    st.info(f"**Local:** {clean_value(selected.get('location'))}  \n**Patrimônio/código:** {clean_value(selected.get('equipment_code'))}")
                reporter_id = _current_user_id()
                st.caption(f"Ticket registrado por: {clean_value(current_user().get('full_name'))}")
                title = st.text_input("Resumo do problema", placeholder="Ex.: Microscópio não liga", key="member_corr_title")
                occurrence_date = st.date_input("Data da ocorrência", value=date.today(), key="member_corr_occurrence_date")
                occurrence_time = st.time_input("Hora da ocorrência", value=datetime.now().time().replace(second=0, microsecond=0), step=timedelta(minutes=15), key="member_corr_occurrence_time")
            with c2:
                description = st.text_area("Descrição do problema *", placeholder="Explique a falha, mensagem de erro, contexto de uso, sintomas observados...", key="member_corr_desc")
                priority = st.selectbox("Prioridade sugerida", ["alta", "média", "baixa"], index=2, key="member_corr_priority")
                attachment = st.file_uploader("Anexo opcional (foto, vídeo, print)", type=["png", "jpg", "jpeg", "pdf", "mp4", "mov"], key="member_corr_attach")

            st.markdown("#### Peças de reposição associadas ao equipamento")
            if equipment_id is None:
                st.caption("Selecione um equipamento para consultar peças associadas.")
            else:
                render_equipment_spare_parts(list_spare_parts_for_equipment(conn, equipment_id))

            if st.button("Abrir ticket corretivo", type="primary", disabled=equipment_id is None, key="member_open_corrective"):
                if equipment_id is None:
                    st.error("Selecione um equipamento para abrir o ticket.")
                elif not title.strip() or not description.strip():
                    st.error("Informe o resumo e a descrição do problema.")
                elif not _ensure_storage_ready_for_upload(attachment):
                    pass
                else:
                    occurrence_dt = datetime.combine(occurrence_date, occurrence_time).isoformat(timespec="minutes")
                    ticket_id = create_corrective_ticket(
                        conn,
                        equipment_id=equipment_id,
                        reporter_id=reporter_id,
                        title=title.strip(),
                        description=description.strip(),
                        occurrence_datetime=occurrence_dt,
                        impact="baixo",
                        priority=priority,
                        attachment_path=None,
                        assigned_to=clean_input(selected.get("responsible_name")),
                        initial_diagnosis=None,
                        probable_cause=None,
                        operator_trained="não informado",
                        external_supplier_needed=0,
                        corrective_action=None,
                        replaced_parts=None,
                        costs=None,
                        downtime_hours=None,
                        conclusion_date=None,
                        status="aberto",
                        notify_technical=1,
                        notify_manager=1,
                        notify_supplier=0,
                        notify_reporter=1,
                    )
                    if attachment is not None:
                        attachment_ref = _save_upload(
                            conn,
                            attachment,
                            entity_type="maintenance_corrective",
                            entity_id=ticket_id,
                            attachment_role="corrective_attachment",
                        )
                        update_legacy_attachment_path(
                            conn,
                            table="maintenance_corrective",
                            row_id=ticket_id,
                            column="attachment_path",
                            value=attachment_ref,
                        )
                    st.success("Ticket corretivo registrado.")
                    clear_app_caches()
                    st.rerun()

        st.markdown("### Tickets corretivos")
        corr_df = query_df(
            conn,
            """
            SELECT mc.id, mc.equipment_id, e.equipment_code, e.equipment_name, e.location,
                   u.full_name AS reporter, mc.title, mc.priority,
                   mc.status, mc.occurrence_datetime, mc.created_at, mc.attachment_path
            FROM maintenance_corrective mc
            JOIN equipment e ON e.id = mc.equipment_id
            LEFT JOIN users u ON u.id = mc.reporter_id
            WHERE COALESCE(mc.is_active, 1) = 1
            ORDER BY CASE mc.status WHEN 'aberto' THEN 0 WHEN 'em análise' THEN 1 WHEN 'aguardando peça' THEN 2 ELSE 3 END,
                     mc.created_at DESC
            """,
        )
        if corr_df.empty:
            st.info("Nenhum ticket corretivo registrado.")
        else:
            member_ticket_cols = [
                "id", "equipment_code", "equipment_name", "location",
                "reporter", "title", "priority", "status",
                "occurrence_datetime", "created_at",
            ]
            st.dataframe(_display_df(corr_df[[c for c in member_ticket_cols if c in corr_df.columns]]), use_container_width=True, hide_index=True)
            st.markdown("#### Detalhes do ticket")
            ticket_options = [None] + corr_df["id"].astype(int).tolist()
            ticket_labels = {
                int(ticket["id"]): (
                    f"Ticket #{int(ticket['id'])} · "
                    f"{clean_value(ticket.get('equipment_code'))} · "
                    f"{clean_value(ticket.get('title'))}"
                )
                for _, ticket in corr_df.iterrows()
            }
            selected_ticket_id = st.selectbox(
                "Ver detalhes do ticket",
                ticket_options,
                format_func=lambda ticket_id: (
                    "Selecione um ticket"
                    if ticket_id is None
                    else ticket_labels.get(int(ticket_id), f"Ticket #{ticket_id}")
                ),
                key="member_ticket_detail_id",
            )
            if selected_ticket_id is None:
                st.caption("Selecione um ticket para consultar anexos, peças associadas e histórico.")
            else:
                selected_ticket = corr_df[corr_df["id"].astype(int) == int(selected_ticket_id)].iloc[0]
                with st.container(border=True):
                    legacy_path = selected_ticket.get("attachment_path")
                    st.caption(
                        f"{clean_value(selected_ticket.get('equipment_code'))} · "
                        f"{clean_value(selected_ticket.get('equipment_name'))} · "
                        f"Status: {clean_value(selected_ticket.get('status'))}"
                    )
                    render_attachment_list(
                        conn,
                        entity_type="maintenance_corrective",
                        entity_id=int(selected_ticket["id"]),
                        attachment_role="corrective_attachment",
                        legacy_path=legacy_path,
                        key_prefix=f"member_corrective_{int(selected_ticket['id'])}",
                        title="Anexo",
                        empty_message="Nenhum anexo cadastrado.",
                    )
                    st.markdown("##### Peças de reposição associadas ao equipamento")
                    render_equipment_spare_parts(list_spare_parts_for_equipment(conn, int(selected_ticket["equipment_id"])))
                    st.markdown("##### Histórico de status")
                    render_maintenance_status_history(conn, entity_type="corrective", entity_id=int(selected_ticket["id"]))
        return

    maintenance_section = st.radio(
        "Escolha uma seção",
        [
            "Preventiva e calibração",
            "Corretiva e suporte",
            "Indicadores e histórico",
        ],
        horizontal=True,
        key="maintenance_section",
    )

    if maintenance_section == "Preventiva e calibração":
        st.markdown("### Manutenção preventiva e calibração")
        st.write("Registro de atividades planejadas, periódicas e obrigatórias: preventiva, calibração interna/externa e inspeções.")
        if not can_manage_maintenance:
            st.info("A gestão de manutenção preventiva é restrita a Gerente ou Administrador.")

        with st.container(border=True):
            c1, c2, c3 = st.columns([1.2, 1, 1])
            with c1:
                eq_label = st.selectbox("Equipamento", _equipment_options(equipment), key="prev_eq")
                equipment_id = _equipment_id_from_label(equipment, eq_label)
                selected = equipment[equipment["id"] == equipment_id].iloc[0]
                st.info(f"**Local:** {clean_value(selected.get('location'))}  \n**Responsável:** {clean_value(selected.get('responsible_name'))}")
                activity_type = st.selectbox(
                    "Tipo da atividade",
                    ["Preventiva", "Calibração interna", "Calibração externa", "Inspeção periódica"],
                    key="prev_activity_type",
                )
                status = st.selectbox("Status", PREVENTIVE_STATUSES, key="prev_status")
            with c2:
                description = st.text_area(
                    "Descrição da atividade",
                    placeholder="Ex.: troca de filtros, lubrificação, verificação de alinhamento, calibração de sensores...",
                    key="prev_desc",
                )
                periodicity = st.selectbox("Periodicidade", ["mensal", "trimestral", "semestral", "anual", "por horas de uso", "sob demanda"], key="prev_periodicity")
                planned_date = st.date_input("Data inicial prevista", value=date.today(), key="prev_planned")
                planned_end_date = st.date_input("Data final prevista", value=date.today(), key="prev_planned_end")
                performed_date = st.date_input("Data realizada", value=None, key="prev_done")
                execution_time = st.text_input("Tempo de execução", placeholder="Ex.: 2 h, 1 dia, 30 min", key="prev_execution_time")
            with c3:
                internal_responsible = st.text_input("Responsável interno", value=clean_input(selected.get("responsible_name")), key="prev_internal_responsible")
                external_supplier = st.text_input("Fornecedor externo", placeholder="Se houver", key="prev_external_supplier")
                supplier_contact = st.text_input("Contato do fornecedor", placeholder="Telefone/e-mail", key="prev_supplier_contact")
                service_order = st.text_input("OS / protocolo externo", key="prev_service_order")
                next_date = st.date_input("Próxima data", value=None, key="prev_next")

            c4, c5 = st.columns(2)
            with c4:
                checklist = st.file_uploader("Checklist anexado (PDF/imagem/formulário)", type=["pdf", "png", "jpg", "jpeg"], key="prev_check")
            with c5:
                certificate = st.file_uploader("Certificado de calibração", type=["pdf", "png", "jpg", "jpeg"], key="prev_cert")

            observations = st.text_area("Observações", key="prev_obs")
            blocks_booking = st.checkbox(
                "Bloquear novas reservas neste período",
                value=True,
                key="prev_blocks_booking",
                help="Use para manutenção, calibração ou inspeção que impeça uso do equipamento. Desmarque quando for apenas registro documental.",
            )
            st.markdown("#### Notificações futuras")
            n1, n2, n3, n4 = st.columns(4)
            notify_internal = n1.checkbox("Responsável interno", value=True, key="prev_notify_internal")
            notify_manager = n2.checkbox("Gestor do laboratório", value=True, key="prev_notify_manager")
            notify_supplier = n3.checkbox("Fornecedor", value=False, key="prev_notify_supplier")
            notify_users = n4.checkbox("Usuários do equipamento", value=False, key="prev_notify_users")
            prev_create_status_reason = ""
            if _maintenance_status_requires_justification("pendente", status, creating=True):
                prev_create_status_reason = st.text_area(
                    "Justificativa do status *",
                    placeholder="Explique brevemente o motivo do status selecionado.",
                    key="prev_create_status_reason",
                )

            if st.button("Registrar preventiva/calibração", type="primary", disabled=not can_manage_maintenance):
                if not description.strip():
                    st.error("Informe a descrição da atividade.")
                elif planned_end_date and planned_date and planned_end_date < planned_date:
                    st.error("A data final prevista não pode ser anterior à data inicial.")
                elif _maintenance_status_requires_justification("pendente", status, creating=True) and not prev_create_status_reason.strip():
                    st.error("Informe a justificativa para este status.")
                elif not _ensure_storage_ready_for_upload(checklist, certificate):
                    pass
                else:
                    base_status = "pendente" if status != "pendente" else status
                    preventive_id = create_preventive_activity(
                        conn,
                        equipment_id=equipment_id,
                        activity_type=activity_type,
                        description=description.strip(),
                        periodicity=periodicity,
                        planned_date=planned_date.isoformat() if planned_date else None,
                        planned_end_date=planned_end_date.isoformat() if planned_end_date else None,
                        performed_date=performed_date.isoformat() if performed_date else None,
                        execution_time=execution_time.strip() or None,
                        checklist_path=None,
                        internal_responsible=internal_responsible.strip() or None,
                        external_supplier=external_supplier.strip() or None,
                        supplier_contact=supplier_contact.strip() or None,
                        service_order=service_order.strip() or None,
                        status=base_status,
                        certificate_path=None,
                        observations=observations.strip() or None,
                        next_date=next_date.isoformat() if next_date else None,
                        blocks_booking=int(blocks_booking),
                        notify_internal=int(notify_internal),
                        notify_manager=int(notify_manager),
                        notify_supplier=int(notify_supplier),
                        notify_users=int(notify_users),
                    )
                    if checklist is not None:
                        checklist_ref = _save_upload(
                            conn,
                            checklist,
                            entity_type="maintenance_preventive",
                            entity_id=preventive_id,
                            attachment_role="preventive_checklist",
                        )
                        update_legacy_attachment_path(
                            conn,
                            table="maintenance_preventive",
                            row_id=preventive_id,
                            column="checklist_path",
                            value=checklist_ref,
                        )
                    if certificate is not None:
                        cert_ref = _save_upload(
                            conn,
                            certificate,
                            entity_type="maintenance_preventive",
                            entity_id=preventive_id,
                            attachment_role="preventive_certificate",
                        )
                        update_legacy_attachment_path(
                            conn,
                            table="maintenance_preventive",
                            row_id=preventive_id,
                            column="certificate_path",
                            value=cert_ref,
                        )
                    if status != base_status:
                        change_maintenance_status(
                            conn,
                            entity_type="preventive",
                            entity_id=preventive_id,
                            new_status=status,
                            justification=prev_create_status_reason.strip() or None,
                            changed_by_id=_current_user_id(),
                        )
                    if blocks_booking and status not in {"realizado", "cancelado"}:
                        sent, total = notify_equipment_maintenance(
                            conn,
                            equipment_id=equipment_id,
                            title="manutenção/calibração agendada",
                            message=(
                                f"Foi registrada uma atividade bloqueante no equipamento.\n"
                                f"Tipo: {activity_type}\n"
                                f"Período: {planned_date.strftime('%d/%m/%Y')} a {planned_end_date.strftime('%d/%m/%Y')}\n"
                                f"Descrição: {description.strip()}"
                            ),
                            related_table="maintenance_preventive",
                            related_id=preventive_id,
                            include_future_users=bool(notify_users),
                        )
                        if total:
                            st.info(f"Notificação de manutenção registrada para {total} destinatário(s). Enviadas: {sent}.")
                    st.success("Atividade preventiva/calibração registrada.")
                    clear_app_caches()
                    st.rerun()

        st.markdown("### Próximas preventivas/calibrações")
        prev_df = list_upcoming_preventive_maintenance(conn)
        if prev_df.empty:
            st.info("Nenhuma preventiva/calibração ativa pendente.")
        else:
            prev_display = prev_df.copy()
            prev_display["reference_label"] = prev_display.apply(_maintenance_reference_label, axis=1)
            prev_cols = [
                "id", "equipment_code", "equipment_name", "activity_type", "status",
                "reference_label", "planned_date", "planned_end_date", "next_date",
                "periodicity", "blocks_booking", "internal_responsible",
                "external_supplier", "service_order",
            ]
            st.dataframe(_display_df(prev_display[[c for c in prev_cols if c in prev_display.columns]]), use_container_width=True, hide_index=True)
            st.markdown("#### Anexos cadastrados")
            shown_preventive = 0
            for _, preventive in prev_df.iterrows():
                checklist_rows = list_attachments(
                    conn,
                    entity_type="maintenance_preventive",
                    entity_id=int(preventive["id"]),
                    attachment_role="preventive_checklist",
                )
                certificate_rows = list_attachments(
                    conn,
                    entity_type="maintenance_preventive",
                    entity_id=int(preventive["id"]),
                    attachment_role="preventive_certificate",
                )
                has_checklist = checklist_rows or not is_blank(preventive.get("checklist_path"))
                has_certificate = certificate_rows or not is_blank(preventive.get("certificate_path"))
                if not has_checklist and not has_certificate:
                    continue
                shown_preventive += 1
                title = f"Preventiva #{int(preventive['id'])} · {clean_value(preventive.get('equipment_code'))} · {clean_value(preventive.get('activity_type'))}"
                with st.expander(title, expanded=False):
                    p1, p2 = st.columns(2)
                    with p1:
                        render_attachment_list(
                            conn,
                            entity_type="maintenance_preventive",
                            entity_id=int(preventive["id"]),
                            attachment_role="preventive_checklist",
                            legacy_path=preventive.get("checklist_path"),
                            key_prefix=f"preventive_{int(preventive['id'])}_checklist",
                            title="Checklist",
                            empty_message="Nenhum checklist cadastrado.",
                        )
                    with p2:
                        render_attachment_list(
                            conn,
                            entity_type="maintenance_preventive",
                            entity_id=int(preventive["id"]),
                            attachment_role="preventive_certificate",
                            legacy_path=preventive.get("certificate_path"),
                            key_prefix=f"preventive_{int(preventive['id'])}_certificate",
                            title="Certificado",
                            empty_message="Nenhum certificado cadastrado.",
                        )
            if shown_preventive == 0:
                st.caption("Nenhum anexo cadastrado.")

        if can_manage_maintenance:
            st.markdown("### Editar ou inativar preventiva/calibração")
            prev_edit_df = query_df(
                conn,
                """
                SELECT mp.*, e.equipment_code, e.equipment_name, e.location
                FROM maintenance_preventive mp
                JOIN equipment e ON e.id = mp.equipment_id
                WHERE COALESCE(mp.is_active, 1) = 1
                ORDER BY COALESCE(NULLIF(mp.next_date, ''), NULLIF(mp.planned_date, ''), mp.created_at) DESC,
                         mp.id DESC
                """,
            )
        else:
            prev_edit_df = pd.DataFrame()
        if prev_edit_df.empty:
            if can_manage_maintenance:
                st.info("Nenhuma preventiva/calibração ativa para editar.")
        else:
            prev_options = prev_edit_df.apply(
                lambda r: (
                    f"#{int(r['id'])} · {clean_value(r.get('equipment_code'))} · "
                    f"{clean_value(r.get('activity_type'))} · {_maintenance_reference_label(r)} · "
                    f"{clean_value(r.get('status'), 'pendente')}"
                ),
                axis=1,
            ).tolist()
            prev_label = st.selectbox("Selecionar preventiva/calibração", prev_options, key="prev_edit_select")
            selected_prev = prev_edit_df.iloc[prev_options.index(prev_label)]
            prev_token = f"prev_edit_{int(selected_prev['id'])}"
            with st.container(border=True):
                c1, c2, c3 = st.columns([1.2, 1, 1])
                equipment_ids = equipment["id"].astype(int).tolist()
                selected_equipment_id = int(selected_prev["equipment_id"])
                equipment_index = equipment_ids.index(selected_equipment_id) if selected_equipment_id in equipment_ids else 0
                with c1:
                    eq_label = st.selectbox("Equipamento", _equipment_options(equipment), index=equipment_index, key=f"{prev_token}_eq")
                    edit_equipment_id = _equipment_id_from_label(equipment, eq_label)
                    edit_selected_eq = equipment[equipment["id"] == edit_equipment_id].iloc[0]
                    st.info(f"**Local:** {clean_value(edit_selected_eq.get('location'))}  \n**Responsável:** {clean_value(edit_selected_eq.get('responsible_name'))}")
                    prev_activity_options = ["Preventiva", "Calibração interna", "Calibração externa", "Inspeção periódica"]
                    edit_activity_type = st.selectbox(
                        "Tipo da atividade",
                        prev_activity_options,
                        index=_option_index(prev_activity_options, selected_prev.get("activity_type")),
                        key=f"{prev_token}_activity_type",
                    )
                    previous_status = clean_input(selected_prev.get("status")) or "pendente"
                    edit_status = st.selectbox(
                        "Status",
                        PREVENTIVE_STATUSES,
                        index=_option_index(PREVENTIVE_STATUSES, previous_status),
                        key=f"{prev_token}_status",
                    )
                with c2:
                    edit_description = st.text_area(
                        "Descrição da atividade",
                        value=clean_input(selected_prev.get("description")),
                        key=f"{prev_token}_desc",
                    )
                    prev_periodicity_options = ["mensal", "trimestral", "semestral", "anual", "por horas de uso", "sob demanda"]
                    edit_periodicity = st.selectbox(
                        "Periodicidade",
                        prev_periodicity_options,
                        index=_option_index(prev_periodicity_options, selected_prev.get("periodicity")),
                        key=f"{prev_token}_periodicity",
                    )
                    edit_planned_date = st.date_input("Data inicial prevista", value=_date_input_value(selected_prev.get("planned_date"), date.today()), key=f"{prev_token}_planned")
                    edit_planned_end_date = st.date_input("Data final prevista", value=_date_input_value(selected_prev.get("planned_end_date"), edit_planned_date), key=f"{prev_token}_planned_end")
                    edit_performed_date = st.date_input("Data realizada", value=_date_input_value(selected_prev.get("performed_date")), key=f"{prev_token}_done")
                    edit_execution_time = st.text_input("Tempo de execução", value=clean_input(selected_prev.get("execution_time")), key=f"{prev_token}_execution_time")
                with c3:
                    edit_internal_responsible = st.text_input("Responsável interno", value=clean_input(selected_prev.get("internal_responsible")), key=f"{prev_token}_internal_responsible")
                    edit_external_supplier = st.text_input("Fornecedor externo", value=clean_input(selected_prev.get("external_supplier")), key=f"{prev_token}_external_supplier")
                    edit_supplier_contact = st.text_input("Contato do fornecedor", value=clean_input(selected_prev.get("supplier_contact")), key=f"{prev_token}_supplier_contact")
                    edit_service_order = st.text_input("OS / protocolo externo", value=clean_input(selected_prev.get("service_order")), key=f"{prev_token}_service_order")
                    edit_next_date = st.date_input("Próxima data", value=_date_input_value(selected_prev.get("next_date")), key=f"{prev_token}_next")

                edit_status_reason = ""
                edit_requires_reason = _maintenance_status_requires_justification(previous_status, edit_status, creating=False)
                if edit_requires_reason:
                    edit_status_reason = st.text_area(
                        "Justificativa da alteração de status *",
                        placeholder="Explique brevemente o motivo do status selecionado.",
                        key=f"{prev_token}_status_reason",
                    )

                u1, u2 = st.columns(2)
                with u1:
                    edit_checklist = st.file_uploader("Novo checklist anexado", type=["pdf", "png", "jpg", "jpeg"], key=f"{prev_token}_check")
                with u2:
                    edit_certificate = st.file_uploader("Novo certificado de calibração", type=["pdf", "png", "jpg", "jpeg"], key=f"{prev_token}_cert")

                edit_observations = st.text_area("Observações", value=clean_input(selected_prev.get("observations")), key=f"{prev_token}_obs")
                edit_blocks_booking = st.checkbox(
                    "Bloquear novas reservas neste período",
                    value=truthy(selected_prev.get("blocks_booking")),
                    key=f"{prev_token}_blocks_booking",
                )
                n1, n2, n3, n4 = st.columns(4)
                edit_notify_internal = n1.checkbox("Responsável interno", value=truthy(selected_prev.get("notify_internal")), key=f"{prev_token}_notify_internal")
                edit_notify_manager = n2.checkbox("Gestor do laboratório", value=truthy(selected_prev.get("notify_manager")), key=f"{prev_token}_notify_manager")
                edit_notify_supplier = n3.checkbox("Fornecedor", value=truthy(selected_prev.get("notify_supplier")), key=f"{prev_token}_notify_supplier")
                edit_notify_users = n4.checkbox("Usuários do equipamento", value=truthy(selected_prev.get("notify_users")), key=f"{prev_token}_notify_users")

                if st.button("Atualizar preventiva/calibração", type="primary", key=f"{prev_token}_save"):
                    if not edit_description.strip():
                        st.error("Informe a descrição da atividade.")
                    elif edit_planned_end_date and edit_planned_date and edit_planned_end_date < edit_planned_date:
                        st.error("A data final prevista não pode ser anterior à data inicial.")
                    elif edit_requires_reason and not edit_status_reason.strip():
                        st.error("Informe a justificativa para este status.")
                    elif not _ensure_storage_ready_for_upload(edit_checklist, edit_certificate):
                        pass
                    else:
                        checklist_final = clean_input(selected_prev.get("checklist_path")) or None
                        certificate_final = clean_input(selected_prev.get("certificate_path")) or None
                        if edit_checklist is not None:
                            checklist_final = _save_upload(
                                conn,
                                edit_checklist,
                                entity_type="maintenance_preventive",
                                entity_id=int(selected_prev["id"]),
                                attachment_role="preventive_checklist",
                            )
                        if edit_certificate is not None:
                            certificate_final = _save_upload(
                                conn,
                                edit_certificate,
                                entity_type="maintenance_preventive",
                                entity_id=int(selected_prev["id"]),
                                attachment_role="preventive_certificate",
                            )
                        ok, msg = update_preventive_activity(
                            conn,
                            int(selected_prev["id"]),
                            equipment_id=edit_equipment_id,
                            activity_type=edit_activity_type,
                            description=edit_description.strip(),
                            periodicity=edit_periodicity,
                            planned_date=edit_planned_date.isoformat() if edit_planned_date else None,
                            planned_end_date=edit_planned_end_date.isoformat() if edit_planned_end_date else None,
                            performed_date=edit_performed_date.isoformat() if edit_performed_date else None,
                            execution_time=edit_execution_time.strip() or None,
                            checklist_path=checklist_final,
                            internal_responsible=edit_internal_responsible.strip() or None,
                            external_supplier=edit_external_supplier.strip() or None,
                            supplier_contact=edit_supplier_contact.strip() or None,
                            service_order=edit_service_order.strip() or None,
                            status=edit_status,
                            certificate_path=certificate_final,
                            observations=edit_observations.strip() or None,
                            next_date=edit_next_date.isoformat() if edit_next_date else None,
                            blocks_booking=int(edit_blocks_booking),
                            notify_internal=int(edit_notify_internal),
                            notify_manager=int(edit_notify_manager),
                            notify_supplier=int(edit_notify_supplier),
                            notify_users=int(edit_notify_users),
                            status_justification=edit_status_reason.strip() or None,
                            changed_by_id=_current_user_id(),
                        )
                        if ok:
                            st.success(msg)
                            clear_app_caches()
                            st.rerun()
                        else:
                            st.error(msg)

            with st.expander("Anexos e histórico da preventiva/calibração", expanded=False):
                p1, p2 = st.columns(2)
                with p1:
                    render_attachment_list(
                        conn,
                        entity_type="maintenance_preventive",
                        entity_id=int(selected_prev["id"]),
                        attachment_role="preventive_checklist",
                        legacy_path=selected_prev.get("checklist_path"),
                        key_prefix=f"{prev_token}_current_checklist",
                        title="Checklist",
                        empty_message="Nenhum checklist cadastrado.",
                    )
                with p2:
                    render_attachment_list(
                        conn,
                        entity_type="maintenance_preventive",
                        entity_id=int(selected_prev["id"]),
                        attachment_role="preventive_certificate",
                        legacy_path=selected_prev.get("certificate_path"),
                        key_prefix=f"{prev_token}_current_certificate",
                        title="Certificado",
                        empty_message="Nenhum certificado cadastrado.",
                    )
                st.markdown("##### Histórico de status")
                render_maintenance_status_history(conn, entity_type="preventive", entity_id=int(selected_prev["id"]))

            with st.expander("Inativar lançamento incorreto", expanded=False):
                inactive_reason = st.text_area("Motivo da inativação *", key=f"{prev_token}_inactive_reason")
                if st.button("Inativar preventiva/calibração", key=f"{prev_token}_inactivate"):
                    ok, msg = inactivate_maintenance_record(
                        conn,
                        entity_type="preventive",
                        entity_id=int(selected_prev["id"]),
                        inactive_reason=inactive_reason,
                        inactive_by_id=_current_user_id(),
                    )
                    if ok:
                        st.success(msg)
                        clear_app_caches()
                        st.rerun()
                    else:
                        st.error(msg)

        if can_manage_maintenance:
            inactive_prev_df = query_df(
                conn,
                """
                SELECT mp.id, e.equipment_code, e.equipment_name, mp.activity_type,
                       mp.status, mp.inactive_reason, mp.inactive_at, u.full_name AS inactive_by_name
                FROM maintenance_preventive mp
                JOIN equipment e ON e.id = mp.equipment_id
                LEFT JOIN users u ON u.id = mp.inactive_by_id
                WHERE COALESCE(mp.is_active, 1) = 0
                ORDER BY mp.inactive_at DESC, mp.id DESC
                """,
            )
        else:
            inactive_prev_df = pd.DataFrame()
        if not inactive_prev_df.empty:
            with st.expander("Preventivas/calibrações inativas para auditoria", expanded=False):
                st.dataframe(_display_df(inactive_prev_df), use_container_width=True, hide_index=True)

    elif maintenance_section == "Corretiva e suporte":
        st.markdown("### Manutenção corretiva e suporte")
        st.write("Tickets abertos por usuários quando há falha, quebra, ruído, anomalia operacional ou necessidade de suporte.")
        if qr_maintenance_message:
            st.info(qr_maintenance_message)
        elif qr_maintenance_issue:
            st.warning(qr_maintenance_issue)
        if not can_manage_maintenance:
            st.caption("Membros podem abrir tickets corretivos. Edição, alteração de status e inativação ficam restritas a Gerente ou Administrador.")

        with st.container(border=True):
            c1, c2 = st.columns(2)
            with c1:
                equipment_options = _equipment_options(equipment)
                manual_placeholder = "Selecione um equipamento"
                corr_select_options = equipment_options
                corr_equipment_index = 0
                if qr_maintenance_equipment_id is not None:
                    equipment_ids = equipment["id"].astype(int).tolist()
                    if qr_maintenance_equipment_id in equipment_ids:
                        corr_equipment_index = equipment_ids.index(qr_maintenance_equipment_id)
                elif qr_maintenance_issue:
                    corr_select_options = [manual_placeholder] + equipment_options
                eq_label = st.selectbox("Equipamento", corr_select_options, index=corr_equipment_index, key="corr_eq")
                equipment_id = None
                selected = pd.Series(dtype=object)
                if eq_label == manual_placeholder:
                    st.info("Selecione manualmente um equipamento para abrir o ticket.")
                else:
                    equipment_id = _equipment_id_from_label(equipment, eq_label)
                    selected = equipment[equipment["id"] == equipment_id].iloc[0]
                    st.info(f"**Local:** {clean_value(selected.get('location'))}  \n**Patrimônio/código:** {clean_value(selected.get('equipment_code'))}")
                reporter_id = None
                if can_manage_maintenance and not users.empty:
                    user_labels = ["Não informado"] + users.apply(lambda r: f"{clean_value(r.get('full_name'))} ({clean_value(r.get('role'), 'member')})", axis=1).tolist()
                    reporter_label = st.selectbox("Usuário que abriu o ticket", user_labels, key="corr_reporter")
                    if reporter_label != "Não informado":
                        reporter_id = int(users.iloc[user_labels.index(reporter_label) - 1]["id"])
                elif not can_manage_maintenance:
                    reporter_id = _current_user_id()
                    st.caption(f"Ticket registrado por: {clean_value(current_user().get('full_name'))}")
                title = st.text_input(
                    "Título do ticket" if can_manage_maintenance else "Resumo do problema",
                    placeholder="Ex.: Microscópio não liga",
                    key="corr_title",
                )
                occurrence_date = st.date_input("Data da ocorrência", value=date.today(), key="corr_occurrence_date")
                occurrence_time = st.time_input("Hora da ocorrência", value=datetime.now().time().replace(second=0, microsecond=0), step=timedelta(minutes=15), key="corr_occurrence_time")
            with c2:
                description = st.text_area("Descrição detalhada" if can_manage_maintenance else "Descrição do problema *", placeholder="Explique a falha, mensagem de erro, contexto de uso, sintomas observados...", key="corr_desc")
                if can_manage_maintenance:
                    impact = st.selectbox("Impacto", ["crítico", "moderado", "baixo"], index=2, key="corr_impact")
                else:
                    impact = "baixo"
                priority = st.selectbox("Prioridade sugerida", ["alta", "média", "baixa"], index=2, key="corr_priority")
                attachment = st.file_uploader("Anexos (foto, vídeo, print)", type=["png", "jpg", "jpeg", "pdf", "mp4", "mov"], key="corr_attach")

            st.markdown("#### Peças de reposição associadas ao equipamento")
            if equipment_id is None:
                st.caption("Selecione um equipamento para consultar peças associadas.")
            else:
                render_equipment_spare_parts(list_spare_parts_for_equipment(conn, equipment_id))

            if can_manage_maintenance:
                st.markdown("#### Diagnóstico e ações")
                d1, d2, d3 = st.columns(3)
                with d1:
                    assigned_to = st.text_input("Responsável pelo atendimento", value=clean_input(selected.get("responsible_name")), key="corr_assigned_to")
                    operator_trained = st.selectbox("Operador era treinado?", ["não informado", "sim", "não"], key="corr_operator_trained")
                    external_supplier_needed = st.checkbox("Necessita fornecedor externo", value=False, key="corr_external_supplier_needed")
                with d2:
                    initial_diagnosis = st.text_area("Diagnóstico inicial", key="diag")
                    probable_cause = st.text_area("Causa provável", key="cause")
                with d3:
                    corrective_action = st.text_area("Ação corretiva realizada", key="action")
                    replaced_parts = st.text_input("Peças substituídas", key="corr_replaced_parts")
                    costs = st.number_input("Custos envolvidos (R$)", min_value=0.0, value=0.0, step=50.0, key="corr_costs")
                    downtime_hours = st.number_input("Downtime (h)", min_value=0.0, value=0.0, step=0.5, key="corr_downtime_hours")

                status = st.selectbox("Status do ticket", CORRECTIVE_STATUSES, key="corr_status")
                corr_create_status_reason = ""
                if _maintenance_status_requires_justification("aberto", status, creating=True):
                    corr_create_status_reason = st.text_area(
                        "Justificativa do status *",
                        placeholder="Explique brevemente o motivo do status selecionado.",
                        key="corr_create_status_reason",
                    )
                conclusion_date = st.date_input("Data de conclusão", value=None, key="corr_conclusion")

                st.markdown("#### Notificações futuras")
                n1, n2, n3, n4 = st.columns(4)
                notify_technical = n1.checkbox("Responsável técnico", value=True, key="corr_notify_technical")
                notify_manager = n2.checkbox("Gestor do laboratório", value=True, key="corr_notify_manager")
                notify_supplier = n3.checkbox("Fornecedor", value=False, key="corr_notify_supplier")
                notify_reporter = n4.checkbox("Usuário que abriu", value=True, key="corr_notify_reporter")
            else:
                assigned_to = clean_input(selected.get("responsible_name"))
                operator_trained = "não informado"
                external_supplier_needed = False
                initial_diagnosis = ""
                probable_cause = ""
                corrective_action = ""
                replaced_parts = ""
                costs = 0.0
                downtime_hours = 0.0
                status = "aberto"
                corr_create_status_reason = ""
                conclusion_date = None
                notify_technical = True
                notify_manager = True
                notify_supplier = False
                notify_reporter = True

            if st.button("Abrir ticket corretivo", type="primary", disabled=equipment_id is None):
                if equipment_id is None:
                    st.error("Selecione um equipamento para abrir o ticket.")
                elif not title.strip() or not description.strip():
                    st.error("Informe o título e a descrição do ticket.")
                elif _maintenance_status_requires_justification("aberto", status, creating=True) and not corr_create_status_reason.strip():
                    st.error("Informe a justificativa para este status.")
                elif not _ensure_storage_ready_for_upload(attachment):
                    pass
                else:
                    occurrence_dt = datetime.combine(occurrence_date, occurrence_time).isoformat(timespec="minutes")
                    base_status = "aberto" if status != "aberto" else status
                    ticket_id = create_corrective_ticket(
                        conn,
                        equipment_id=equipment_id,
                        reporter_id=reporter_id,
                        title=title.strip(),
                        description=description.strip(),
                        occurrence_datetime=occurrence_dt,
                        impact=impact,
                        priority=priority,
                        attachment_path=None,
                        assigned_to=assigned_to.strip() or None,
                        initial_diagnosis=initial_diagnosis.strip() or None,
                        probable_cause=probable_cause.strip() or None,
                        operator_trained=operator_trained,
                        external_supplier_needed=int(external_supplier_needed),
                        corrective_action=corrective_action.strip() or None,
                        replaced_parts=replaced_parts.strip() or None,
                        costs=float(costs) if costs else None,
                        downtime_hours=float(downtime_hours) if downtime_hours else None,
                        conclusion_date=conclusion_date.isoformat() if conclusion_date else None,
                        status=base_status,
                        notify_technical=int(notify_technical),
                        notify_manager=int(notify_manager),
                        notify_supplier=int(notify_supplier),
                        notify_reporter=int(notify_reporter),
                    )
                    if attachment is not None:
                        attachment_ref = _save_upload(
                            conn,
                            attachment,
                            entity_type="maintenance_corrective",
                            entity_id=ticket_id,
                            attachment_role="corrective_attachment",
                        )
                        update_legacy_attachment_path(
                            conn,
                            table="maintenance_corrective",
                            row_id=ticket_id,
                            column="attachment_path",
                            value=attachment_ref,
                        )
                    if status != base_status:
                        change_maintenance_status(
                            conn,
                            entity_type="corrective",
                            entity_id=ticket_id,
                            new_status=status,
                            justification=corr_create_status_reason.strip() or None,
                            changed_by_id=_current_user_id(),
                        )
                    st.success("Ticket corretivo registrado.")
                    clear_app_caches()
                    st.rerun()

        st.markdown("### Tickets corretivos")
        if can_manage_maintenance:
            show_inactive_corr = st.checkbox("Mostrar tickets inativos para auditoria", value=False, key="corr_show_inactive")
        else:
            show_inactive_corr = False
        corr_df = query_df(
            conn,
            """
            SELECT mc.id, mc.equipment_id, e.equipment_code, e.equipment_name, e.location,
                   u.full_name AS reporter, mc.title, mc.impact, mc.priority,
                   mc.status, mc.occurrence_datetime, mc.assigned_to,
                   mc.downtime_hours, mc.costs, mc.created_at, mc.attachment_path,
                   mc.is_active, mc.inactive_reason, mc.inactive_at
            FROM maintenance_corrective mc
            JOIN equipment e ON e.id = mc.equipment_id
            LEFT JOIN users u ON u.id = mc.reporter_id
            WHERE (? = 1 OR COALESCE(mc.is_active, 1) = 1)
            ORDER BY CASE mc.status WHEN 'aberto' THEN 0 WHEN 'em análise' THEN 1 WHEN 'aguardando peça' THEN 2 ELSE 3 END,
                     mc.created_at DESC
            """,
            [int(show_inactive_corr)],
        )
        if corr_df.empty:
            st.info("Nenhum ticket corretivo registrado.")
        else:
            st.dataframe(_display_df(corr_df), use_container_width=True, hide_index=True)
            st.markdown("#### Detalhes, anexos e peças")
            for _, ticket in corr_df.iterrows():
                legacy_path = ticket.get("attachment_path")
                title = f"Ticket #{int(ticket['id'])} · {clean_value(ticket.get('equipment_code'))} · {clean_value(ticket.get('title'))}"
                with st.expander(title, expanded=False):
                    render_attachment_list(
                        conn,
                        entity_type="maintenance_corrective",
                        entity_id=int(ticket["id"]),
                        attachment_role="corrective_attachment",
                        legacy_path=legacy_path,
                        key_prefix=f"corrective_{int(ticket['id'])}",
                        title="Anexo",
                        empty_message="Nenhum anexo cadastrado.",
                    )
                    st.markdown("##### Peças de reposição associadas ao equipamento")
                    render_equipment_spare_parts(list_spare_parts_for_equipment(conn, int(ticket["equipment_id"])))
                    st.markdown("##### Histórico de status")
                    render_maintenance_status_history(conn, entity_type="corrective", entity_id=int(ticket["id"]))
            active_corr = corr_df[corr_df["is_active"].map(lambda value: True if is_blank(value) else truthy(value))]
            active_ids = active_corr[~active_corr["status"].isin(["concluído", "cancelado"])] ["id"].tolist() if can_manage_maintenance else []
            if not can_manage_maintenance:
                st.caption("Alteração de status de tickets corretivos é restrita a Gerente ou Administrador.")
            if active_ids:
                c1, c2 = st.columns([1, 1])
                with c1:
                    ticket_id = st.selectbox("Atualizar status do ticket", active_ids, format_func=lambda x: f"Ticket #{x}", key="corr_update_ticket_id")
                with c2:
                    new_status = st.selectbox("Novo status", ["em análise", "aguardando peça", "enviado para fornecedor", "concluído", "cancelado"], key="corr_update_new_status")
                current_ticket = active_corr[active_corr["id"] == ticket_id].iloc[0]
                current_status = clean_input(current_ticket.get("status"))
                quick_requires_reason = _maintenance_status_requires_justification(current_status, new_status, creating=False)
                quick_reason = ""
                if quick_requires_reason:
                    quick_reason = st.text_area("Justificativa da alteração de status *", key="corr_quick_status_reason")
                if st.button("Atualizar status com histórico"):
                    if quick_requires_reason and not quick_reason.strip():
                        st.error("Informe a justificativa para este status.")
                    else:
                        ok, msg = change_maintenance_status(
                            conn,
                            entity_type="corrective",
                            entity_id=int(ticket_id),
                            new_status=new_status,
                            justification=quick_reason.strip() or None,
                            changed_by_id=_current_user_id(),
                        )
                        if ok:
                            st.success(msg)
                            clear_app_caches()
                            st.rerun()
                        else:
                            st.error(msg)

        if can_manage_maintenance:
            st.markdown("### Editar ou inativar ticket corretivo")
            corr_edit_df = query_df(
                conn,
                """
                SELECT mc.*, e.equipment_code, e.equipment_name, e.location,
                       u.full_name AS reporter
                FROM maintenance_corrective mc
                JOIN equipment e ON e.id = mc.equipment_id
                LEFT JOIN users u ON u.id = mc.reporter_id
                WHERE COALESCE(mc.is_active, 1) = 1
                ORDER BY CASE mc.status
                           WHEN 'aberto' THEN 0
                           WHEN 'em análise' THEN 1
                           WHEN 'aguardando peça' THEN 2
                           WHEN 'enviado para fornecedor' THEN 3
                           WHEN 'concluído' THEN 4
                           WHEN 'cancelado' THEN 5
                           ELSE 6
                         END,
                         mc.created_at DESC
                """,
            )
        else:
            corr_edit_df = pd.DataFrame()
        if corr_edit_df.empty:
            if can_manage_maintenance:
                st.info("Nenhum ticket corretivo ativo para editar.")
        else:
            corr_options = corr_edit_df.apply(
                lambda r: (
                    f"#{int(r['id'])} · {clean_value(r.get('equipment_code'))} · "
                    f"{clean_value(r.get('title'))} · {clean_value(r.get('status'), 'aberto')}"
                ),
                axis=1,
            ).tolist()
            corr_label = st.selectbox("Selecionar ticket corretivo", corr_options, key="corr_edit_select")
            selected_corr = corr_edit_df.iloc[corr_options.index(corr_label)]
            corr_token = f"corr_edit_{int(selected_corr['id'])}"

            with st.container(border=True):
                c1, c2 = st.columns(2)
                equipment_ids = equipment["id"].astype(int).tolist()
                selected_equipment_id = int(selected_corr["equipment_id"])
                equipment_index = equipment_ids.index(selected_equipment_id) if selected_equipment_id in equipment_ids else 0
                with c1:
                    eq_label = st.selectbox("Equipamento", _equipment_options(equipment), index=equipment_index, key=f"{corr_token}_eq")
                    edit_equipment_id = _equipment_id_from_label(equipment, eq_label)
                    selected_eq = equipment[equipment["id"] == edit_equipment_id].iloc[0]
                    st.info(f"**Local:** {clean_value(selected_eq.get('location'))}  \n**Patrimônio/código:** {clean_value(selected_eq.get('equipment_code'))}")
                    reporter_id = None
                    if not users.empty:
                        user_labels = ["Não informado"] + _user_options(users)
                        reporter_index = 0
                        if not is_blank(selected_corr.get("reporter_id")):
                            reporter_ids = users["id"].astype(int).tolist()
                            selected_reporter_id = int(selected_corr.get("reporter_id"))
                            if selected_reporter_id in reporter_ids:
                                reporter_index = reporter_ids.index(selected_reporter_id) + 1
                        reporter_label = st.selectbox("Usuário que abriu o ticket", user_labels, index=reporter_index, key=f"{corr_token}_reporter")
                        reporter_id = _user_id_from_label(users, reporter_label)
                    edit_title = st.text_input("Título do ticket", value=clean_input(selected_corr.get("title")), key=f"{corr_token}_title")
                    occurrence_date_default, occurrence_time_default = _datetime_input_defaults(selected_corr.get("occurrence_datetime"))
                    edit_occurrence_date = st.date_input("Data da ocorrência", value=occurrence_date_default, key=f"{corr_token}_occurrence_date")
                    edit_occurrence_time = st.time_input("Hora da ocorrência", value=occurrence_time_default, step=timedelta(minutes=15), key=f"{corr_token}_occurrence_time")
                with c2:
                    edit_description = st.text_area("Descrição detalhada", value=clean_input(selected_corr.get("description")), key=f"{corr_token}_desc")
                    edit_impact = st.selectbox("Impacto", ["crítico", "moderado", "baixo"], index=_option_index(["crítico", "moderado", "baixo"], selected_corr.get("impact"), default=2), key=f"{corr_token}_impact")
                    edit_priority = st.selectbox("Prioridade sugerida", ["alta", "média", "baixa"], index=_option_index(["alta", "média", "baixa"], selected_corr.get("priority"), default=2), key=f"{corr_token}_priority")
                    edit_attachment = st.file_uploader("Novo anexo (foto, vídeo, print)", type=["png", "jpg", "jpeg", "pdf", "mp4", "mov"], key=f"{corr_token}_attach")

                st.markdown("#### Peças de reposição associadas ao equipamento")
                render_equipment_spare_parts(list_spare_parts_for_equipment(conn, edit_equipment_id))

                st.markdown("#### Diagnóstico e ações")
                d1, d2, d3 = st.columns(3)
                with d1:
                    edit_assigned_to = st.text_input("Responsável pelo atendimento", value=clean_input(selected_corr.get("assigned_to")), key=f"{corr_token}_assigned_to")
                    trained_options = ["não informado", "sim", "não"]
                    edit_operator_trained = st.selectbox("Operador era treinado?", trained_options, index=_option_index(trained_options, selected_corr.get("operator_trained")), key=f"{corr_token}_operator_trained")
                    edit_external_supplier_needed = st.checkbox("Necessita fornecedor externo", value=truthy(selected_corr.get("external_supplier_needed")), key=f"{corr_token}_external_supplier_needed")
                with d2:
                    edit_initial_diagnosis = st.text_area("Diagnóstico inicial", value=clean_input(selected_corr.get("initial_diagnosis")), key=f"{corr_token}_diag")
                    edit_probable_cause = st.text_area("Causa provável", value=clean_input(selected_corr.get("probable_cause")), key=f"{corr_token}_cause")
                with d3:
                    edit_corrective_action = st.text_area("Ação corretiva realizada", value=clean_input(selected_corr.get("corrective_action")), key=f"{corr_token}_action")
                    edit_replaced_parts = st.text_input("Peças substituídas", value=clean_input(selected_corr.get("replaced_parts")), key=f"{corr_token}_replaced_parts")
                    edit_costs = st.number_input("Custos envolvidos (R$)", min_value=0.0, value=float(selected_corr.get("costs") or 0), step=50.0, key=f"{corr_token}_costs")
                    edit_downtime_hours = st.number_input("Downtime (h)", min_value=0.0, value=float(selected_corr.get("downtime_hours") or 0), step=0.5, key=f"{corr_token}_downtime_hours")

                previous_corr_status = clean_input(selected_corr.get("status")) or "aberto"
                edit_status = st.selectbox(
                    "Status do ticket",
                    CORRECTIVE_STATUSES,
                    index=_option_index(CORRECTIVE_STATUSES, previous_corr_status),
                    key=f"{corr_token}_status",
                )
                edit_status_reason = ""
                edit_requires_reason = _maintenance_status_requires_justification(previous_corr_status, edit_status, creating=False)
                if edit_requires_reason:
                    edit_status_reason = st.text_area(
                        "Justificativa da alteração de status *",
                        placeholder="Explique brevemente o motivo do status selecionado.",
                        key=f"{corr_token}_status_reason",
                    )
                edit_conclusion_date = st.date_input("Data de conclusão", value=_date_input_value(selected_corr.get("conclusion_date")), key=f"{corr_token}_conclusion")

                n1, n2, n3, n4 = st.columns(4)
                edit_notify_technical = n1.checkbox("Responsável técnico", value=truthy(selected_corr.get("notify_technical")), key=f"{corr_token}_notify_technical")
                edit_notify_manager = n2.checkbox("Gestor do laboratório", value=truthy(selected_corr.get("notify_manager")), key=f"{corr_token}_notify_manager")
                edit_notify_supplier = n3.checkbox("Fornecedor", value=truthy(selected_corr.get("notify_supplier")), key=f"{corr_token}_notify_supplier")
                edit_notify_reporter = n4.checkbox("Usuário que abriu", value=truthy(selected_corr.get("notify_reporter")), key=f"{corr_token}_notify_reporter")

                if st.button("Atualizar ticket corretivo", type="primary", key=f"{corr_token}_save"):
                    if not edit_title.strip() or not edit_description.strip():
                        st.error("Informe o título e a descrição do ticket.")
                    elif edit_requires_reason and not edit_status_reason.strip():
                        st.error("Informe a justificativa para este status.")
                    elif not _ensure_storage_ready_for_upload(edit_attachment):
                        pass
                    else:
                        attachment_final = clean_input(selected_corr.get("attachment_path")) or None
                        if edit_attachment is not None:
                            attachment_final = _save_upload(
                                conn,
                                edit_attachment,
                                entity_type="maintenance_corrective",
                                entity_id=int(selected_corr["id"]),
                                attachment_role="corrective_attachment",
                            )
                        edit_occurrence_dt = datetime.combine(edit_occurrence_date, edit_occurrence_time).isoformat(timespec="minutes")
                        ok, msg = update_corrective_ticket(
                            conn,
                            int(selected_corr["id"]),
                            equipment_id=edit_equipment_id,
                            reporter_id=reporter_id,
                            title=edit_title.strip(),
                            description=edit_description.strip(),
                            occurrence_datetime=edit_occurrence_dt,
                            impact=edit_impact,
                            priority=edit_priority,
                            attachment_path=attachment_final,
                            assigned_to=edit_assigned_to.strip() or None,
                            initial_diagnosis=edit_initial_diagnosis.strip() or None,
                            probable_cause=edit_probable_cause.strip() or None,
                            operator_trained=edit_operator_trained,
                            external_supplier_needed=int(edit_external_supplier_needed),
                            corrective_action=edit_corrective_action.strip() or None,
                            replaced_parts=edit_replaced_parts.strip() or None,
                            costs=float(edit_costs) if edit_costs else None,
                            downtime_hours=float(edit_downtime_hours) if edit_downtime_hours else None,
                            conclusion_date=edit_conclusion_date.isoformat() if edit_conclusion_date else None,
                            status=edit_status,
                            notify_technical=int(edit_notify_technical),
                            notify_manager=int(edit_notify_manager),
                            notify_supplier=int(edit_notify_supplier),
                            notify_reporter=int(edit_notify_reporter),
                            status_justification=edit_status_reason.strip() or None,
                            changed_by_id=_current_user_id(),
                        )
                        if ok:
                            st.success(msg)
                            clear_app_caches()
                            st.rerun()
                        else:
                            st.error(msg)

            with st.expander("Anexo e histórico do ticket", expanded=False):
                render_attachment_list(
                    conn,
                    entity_type="maintenance_corrective",
                    entity_id=int(selected_corr["id"]),
                    attachment_role="corrective_attachment",
                    legacy_path=selected_corr.get("attachment_path"),
                    key_prefix=f"{corr_token}_current_attachment",
                    title="Anexo",
                    empty_message="Nenhum anexo cadastrado.",
                )
                st.markdown("##### Histórico de status")
                render_maintenance_status_history(conn, entity_type="corrective", entity_id=int(selected_corr["id"]))

            with st.expander("Inativar lançamento incorreto", expanded=False):
                inactive_reason = st.text_area("Motivo da inativação *", key=f"{corr_token}_inactive_reason")
                if st.button("Inativar ticket corretivo", key=f"{corr_token}_inactivate"):
                    ok, msg = inactivate_maintenance_record(
                        conn,
                        entity_type="corrective",
                        entity_id=int(selected_corr["id"]),
                        inactive_reason=inactive_reason,
                        inactive_by_id=_current_user_id(),
                    )
                    if ok:
                        st.success(msg)
                        clear_app_caches()
                        st.rerun()
                    else:
                        st.error(msg)

    elif maintenance_section == "Indicadores e histórico":
        st.markdown("### Indicadores de manutenção")
        corr = query_df(conn, "SELECT * FROM maintenance_corrective WHERE COALESCE(is_active, 1) = 1")
        prev = query_df(conn, "SELECT * FROM maintenance_preventive WHERE COALESCE(is_active, 1) = 1")
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Preventivas/calibrações", len(prev))
        k2.metric("Tickets corretivos", len(corr))
        if not corr.empty and "downtime_hours" in corr.columns:
            k3.metric("Downtime total (h)", f"{corr['downtime_hours'].fillna(0).sum():.1f}")
            k4.metric("MTTR preliminar (h)", f"{corr['downtime_hours'].fillna(0).mean():.1f}")
        else:
            k3.metric("Downtime total (h)", "0")
            k4.metric("MTTR preliminar (h)", "0")

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("#### Tickets por status")
            if not corr.empty:
                status_df = corr.groupby("status", dropna=False).size().reset_index(name="total")
                fig = px.bar(status_df, x="status", y="total", color="status", color_discrete_sequence=[LAB_BLUE, LAB_CYAN, "#6BAED6", "#9ECAE1"])
                fig.update_layout(height=330, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Sem tickets corretivos para calcular indicadores.")
        with c2:
            st.markdown("#### Preventivas por status")
            if not prev.empty:
                status_df = prev.groupby("status", dropna=False).size().reset_index(name="total")
                fig = px.bar(status_df, x="status", y="total", color="status", color_discrete_sequence=[LAB_BLUE, LAB_CYAN, "#6BAED6", "#9ECAE1"])
                fig.update_layout(height=330, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Sem preventivas/calibrações para calcular indicadores.")


def _semester_bounds(year: int, semester: int) -> tuple[date, date]:
    if semester == 1:
        return date(year, 1, 1), date(year, 6, 30)
    return date(year, 7, 1), date(year, 12, 31)


def _previous_semester_bounds(today: date) -> tuple[date, date]:
    if today.month <= 6:
        return _semester_bounds(today.year - 1, 2)
    return _semester_bounds(today.year, 1)


def _period_label(start_date: date, end_date: date) -> str:
    return f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"


def _movement_type_is_consumption(value) -> bool:
    return clean_input(value).lower() in SUPPLY_CONSUMPTION_MOVEMENT_TYPES


def _report_label(row: pd.Series, code_col: str, name_col: str, fallback: str) -> str:
    code = clean_input(row.get(code_col))
    name = clean_input(row.get(name_col))
    if code and name:
        return f"{code} - {name}"
    return name or code or fallback


def _attachment_count_map(conn, *, entity_type: str, attachment_role: str) -> dict[int, int]:
    rows = query_df(
        conn,
        """
        SELECT entity_id, COUNT(*) AS attachment_count
        FROM attachments
        WHERE entity_type = ?
          AND attachment_role = ?
          AND COALESCE(is_active, 1) = 1
        GROUP BY entity_id
        """,
        [entity_type, attachment_role],
    )
    if rows.empty:
        return {}
    return {
        int(row["entity_id"]): int(row["attachment_count"] or 0)
        for _, row in rows.iterrows()
        if not is_blank(row.get("entity_id"))
    }


def _attachment_rows_by_entity(
    conn,
    *,
    entity_type: str,
    attachment_role: str,
    entity_ids: list[int] | tuple[int, ...] | set[int],
) -> dict[int, list[dict]]:
    ids = sorted({int(entity_id) for entity_id in entity_ids if not is_blank(entity_id)})
    if not ids:
        return {}
    attachment_rows: list[dict] = []
    for offset in range(0, len(ids), 800):
        id_chunk = ids[offset : offset + 800]
        placeholders = ", ".join(["?"] * len(id_chunk))
        rows = query_df(
            conn,
            f"""
            SELECT *
            FROM attachments
            WHERE entity_type = ?
              AND attachment_role = ?
              AND COALESCE(is_active, 1) = 1
              AND entity_id IN ({placeholders})
            ORDER BY entity_id, uploaded_at DESC, id DESC
            """,
            [entity_type, attachment_role, *id_chunk],
        )
        if not rows.empty:
            attachment_rows.extend(rows.to_dict("records"))
    grouped: dict[int, list[dict]] = {}
    for row in attachment_rows:
        grouped.setdefault(int(row["entity_id"]), []).append(row)
    return grouped


def _metadata_status_from_attachment(path_value, attachment_count, *, present_label: str, missing_label: str) -> str:
    try:
        count = int(attachment_count or 0)
    except Exception:
        count = 0
    if count > 0:
        return present_label
    if not is_blank(path_value):
        return "Anexo legado"
    return missing_label


def _enrich_supply_report_metadata(
    conn,
    supply_lots: pd.DataFrame,
    supply_movements: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    lots = supply_lots.copy()
    movements = supply_movements.copy()

    certificate_counts = _attachment_count_map(
        conn,
        entity_type="supply_lot",
        attachment_role="analysis_certificate",
    )
    movement_attachment_counts = _attachment_count_map(
        conn,
        entity_type="supply_movement",
        attachment_role="movement_document",
    )

    if not lots.empty and "id" in lots.columns:
        lots["certificate_attachment_count"] = lots["id"].map(
            lambda value: certificate_counts.get(int(value), 0) if not is_blank(value) else 0
        )
        lots["certificate_status"] = lots.apply(
            lambda row: _metadata_status_from_attachment(
                row.get("certificate_path"),
                row.get("certificate_attachment_count"),
                present_label="Com certificado",
                missing_label="Sem certificado",
            ),
            axis=1,
        )

    lot_certificate_status = {}
    if not lots.empty and {"id", "certificate_status"}.issubset(lots.columns):
        lot_certificate_status = {
            int(row["id"]): clean_value(row.get("certificate_status"), "Sem certificado")
            for _, row in lots.iterrows()
            if not is_blank(row.get("id"))
        }

    if not movements.empty:
        movements["project_label"] = movements.apply(
            lambda row: _report_label(row, "project_code", "project_name", "Sem projeto"),
            axis=1,
        )
        movements["service_label"] = movements.apply(
            lambda row: _report_label(row, "service_code", "service_title", "Sem serviço/análise"),
            axis=1,
        )
        movements["supply_lot_label"] = movements["supply_lot_code"].map(lambda value: clean_value(value, "Sem lote"))
        movements["lot_supplier_name"] = movements.get("lot_supplier_name", pd.Series(index=movements.index)).map(clean_value)
        movements["lot_location"] = movements.get("lot_location", pd.Series(index=movements.index)).map(clean_value)
        movements["movement_attachment_count"] = movements["id"].map(
            lambda value: movement_attachment_counts.get(int(value), 0) if not is_blank(value) else 0
        )
        movements["movement_document_status"] = movements.apply(
            lambda row: _metadata_status_from_attachment(
                row.get("document_path"),
                row.get("movement_attachment_count"),
                present_label="Com anexo",
                missing_label="Sem anexo",
            ),
            axis=1,
        )
        movements["certificate_status"] = movements["supply_lot_id"].map(
            lambda value: "Sem lote" if is_blank(value) else lot_certificate_status.get(int(value), "Sem certificado")
        )

    return lots, movements


def _stock_minimum_alerts_df(supplies: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "alerta",
        "supply_type",
        "supply_name",
        "supply_code",
        "current_quantity",
        "unit",
        "minimum_quantity",
        "quantity_to_minimum",
        "location",
        "responsible_name",
        "category",
    ]
    if supplies.empty:
        return pd.DataFrame(columns=columns)
    out = supplies.copy()
    if "active" in out.columns:
        out = out[out["active"].map(lambda value: True if is_blank(value) else truthy(value))]
    if out.empty:
        return pd.DataFrame(columns=columns)
    out["current_quantity"] = pd.to_numeric(out["current_quantity"], errors="coerce").fillna(0)
    out["minimum_quantity"] = pd.to_numeric(out["minimum_quantity"], errors="coerce").fillna(0)
    out["alerta"] = out.apply(_supply_alert_status, axis=1)
    out = out[(out["minimum_quantity"] > 0) & (out["alerta"] == "Estoque baixo")].copy()
    if out.empty:
        return pd.DataFrame(columns=columns)
    out["quantity_to_minimum"] = (out["minimum_quantity"] - out["current_quantity"]).clip(lower=0).round(4)
    out = out.sort_values(["quantity_to_minimum", "supply_name"], ascending=[False, True])
    return out[[c for c in columns if c in out.columns]]


def _lot_validity_alerts_df(supply_lots: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "lot_status",
        "supply_name",
        "supply_code",
        "lot_code",
        "expiration_date",
        "days_until_expiration",
        "current_quantity",
        "unit",
        "supplier_name",
        "location",
        "certificate_status",
        "is_active",
    ]
    if supply_lots.empty:
        return pd.DataFrame(columns=columns)
    out = supply_lots.copy()
    if "is_active" in out.columns:
        out = out[out["is_active"].map(lambda value: True if is_blank(value) else truthy(value))]
    if out.empty:
        return pd.DataFrame(columns=columns)
    if "lot_status" not in out.columns:
        out["lot_status"] = out["expiration_date"].map(_lot_expiration_status)
    out["days_until_expiration"] = out["expiration_date"].map(_days_until_date)
    out = out[out["lot_status"].isin(LOT_EXPIRATION_ALERT_STATUSES)].copy()
    if out.empty:
        return pd.DataFrame(columns=columns)
    out = out.sort_values(["days_until_expiration", "supply_name", "lot_code"], ascending=[True, True, True])
    return out[[c for c in columns if c in out.columns]]


def _supply_consumption_detail_df(supply_movements: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "project_label",
        "service_label",
        "supply_name",
        "supply_lot_label",
        "unit",
        "movement_count",
        "consumed_quantity",
    ]
    if supply_movements.empty:
        return pd.DataFrame(columns=columns)
    consumed = supply_movements[supply_movements["movement_type"].map(_movement_type_is_consumption)].copy()
    if consumed.empty:
        return pd.DataFrame(columns=columns)
    if "project_label" not in consumed.columns:
        consumed["project_label"] = consumed.apply(lambda row: _report_label(row, "project_code", "project_name", "Sem projeto"), axis=1)
    if "service_label" not in consumed.columns:
        consumed["service_label"] = consumed.apply(lambda row: _report_label(row, "service_code", "service_title", "Sem serviço/análise"), axis=1)
    if "supply_lot_label" not in consumed.columns:
        consumed["supply_lot_label"] = consumed["supply_lot_code"].map(lambda value: clean_value(value, "Sem lote"))
    consumed["quantity"] = pd.to_numeric(consumed["quantity"], errors="coerce").fillna(0)
    grouped = (
        consumed.groupby(["project_label", "service_label", "supply_name", "supply_lot_label", "unit"], dropna=False)
        .agg(movement_count=("id", "count"), consumed_quantity=("quantity", "sum"))
        .reset_index()
    )
    grouped["consumed_quantity"] = grouped["consumed_quantity"].round(4)
    return grouped.sort_values(["project_label", "service_label", "supply_name", "supply_lot_label"])


def _supply_traceability_df(supply_movements: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "id",
        "supply_name",
        "supply_lot_label",
        "supply_lot_expiration_date",
        "lot_supplier_name",
        "lot_location",
        "certificate_status",
        "movement_type",
        "movement_date",
        "quantity",
        "unit",
        "responsavel",
        "project_label",
        "service_label",
        "movement_document_status",
        "purpose",
        "created_at",
    ]
    if supply_movements.empty:
        return pd.DataFrame(columns=columns)
    out = supply_movements.copy()
    if "project_label" not in out.columns:
        out["project_label"] = out.apply(lambda row: _report_label(row, "project_code", "project_name", "Sem projeto"), axis=1)
    if "service_label" not in out.columns:
        out["service_label"] = out.apply(lambda row: _report_label(row, "service_code", "service_title", "Sem serviço/análise"), axis=1)
    if "supply_lot_label" not in out.columns:
        out["supply_lot_label"] = out["supply_lot_code"].map(lambda value: clean_value(value, "Sem lote"))
    for col in ["lot_supplier_name", "lot_location", "responsavel"]:
        if col in out.columns:
            out[col] = out[col].map(clean_value)
    if "certificate_status" not in out.columns:
        out["certificate_status"] = out["supply_lot_label"].map(lambda value: "Sem lote" if value == "Sem lote" else "Sem certificado")
    if "movement_document_status" not in out.columns:
        out["movement_document_status"] = out["document_path"].map(lambda value: "Anexo legado" if not is_blank(value) else "Sem anexo")
    out["quantity"] = pd.to_numeric(out["quantity"], errors="coerce").fillna(0)
    out = out.sort_values(["supply_name", "supply_lot_label", "movement_date", "id"], ascending=[True, True, True, True])
    return out[[c for c in columns if c in out.columns]]


def _filtered_reports_data(conn, start_date: date, end_date: date) -> dict[str, pd.DataFrame]:
    start_datetime_iso = datetime.combine(start_date, time.min).isoformat(timespec="minutes")
    end_exclusive_datetime_iso = datetime.combine(end_date + timedelta(days=1), time.min).isoformat(timespec="minutes")
    bookings = query_df(
        conn,
        """
        SELECT b.id, e.equipment_code, e.equipment_name, e.lab_unit, e.location,
               u.full_name AS solicitante, u.department AS departamento,
               b.project_id, p.project_code, p.project_name, p.funding_source,
               b.service_id, ps.service_code, ps.title AS service_title, ps.service_type,
               op.full_name AS operador, perf.full_name AS executante,
               COALESCE(perf.full_name, op.full_name, u.full_name) AS responsavel_execucao,
               b.start_datetime, b.end_datetime, b.sample_count, b.purpose, b.status,
               b.created_at, b.updated_at
        FROM bookings b
        JOIN equipment e ON e.id=b.equipment_id
        JOIN users u ON u.id=b.user_id
        LEFT JOIN users op ON op.id=b.operator_id
        LEFT JOIN users perf ON perf.id=b.performed_by_id
        LEFT JOIN projects p ON p.id=b.project_id
        LEFT JOIN project_services ps ON ps.id=b.service_id
        WHERE b.start_datetime >= ?
          AND b.start_datetime < ?
        ORDER BY b.start_datetime
        """,
        [start_datetime_iso, end_exclusive_datetime_iso],
    )

    preventive = query_df(
        conn,
        """
        SELECT mp.id, e.equipment_code, e.equipment_name, e.lab_unit, e.location,
               mp.activity_type, mp.description, mp.periodicity, mp.planned_date,
               mp.planned_end_date, mp.performed_date, mp.execution_time,
               mp.internal_responsible, mp.external_supplier, mp.service_order,
               mp.status, mp.next_date, mp.blocks_booking, mp.observations,
               mp.created_at, mp.updated_at
        FROM maintenance_preventive mp
        JOIN equipment e ON e.id=mp.equipment_id
        WHERE COALESCE(mp.is_active, 1) = 1
          AND SUBSTR(COALESCE(mp.performed_date, mp.planned_date, mp.created_at), 1, 10) BETWEEN ? AND ?
        ORDER BY COALESCE(mp.performed_date, mp.planned_date, mp.created_at)
        """,
        [start_date.isoformat(), end_date.isoformat()],
    )

    corrective = query_df(
        conn,
        """
        SELECT mc.id, e.equipment_code, e.equipment_name, e.lab_unit, e.location,
               u.full_name AS reporter, mc.title, mc.description, mc.occurrence_datetime,
               mc.impact, mc.priority, mc.assigned_to, mc.initial_diagnosis,
               mc.probable_cause, mc.operator_trained, mc.external_supplier_needed,
               mc.corrective_action, mc.replaced_parts, mc.costs, mc.downtime_hours,
               mc.conclusion_date, mc.status, mc.created_at, mc.updated_at
        FROM maintenance_corrective mc
        JOIN equipment e ON e.id=mc.equipment_id
        LEFT JOIN users u ON u.id=mc.reporter_id
        WHERE COALESCE(mc.is_active, 1) = 1
          AND SUBSTR(COALESCE(mc.conclusion_date, mc.occurrence_datetime, mc.created_at), 1, 10) BETWEEN ? AND ?
        ORDER BY COALESCE(mc.conclusion_date, mc.occurrence_datetime, mc.created_at)
        """,
        [start_date.isoformat(), end_date.isoformat()],
    )

    supply_movements = query_df(
        conn,
        """
        SELECT sm.id, sm.supply_id, sm.supply_lot_id,
               s.supply_type, s.supply_code, s.supply_name,
               s.commercial_name, s.manufacturer, s.category,
               s.physical_state, sm.movement_type, sm.movement_date, sm.quantity,
               COALESCE(sm.unit, s.unit) AS unit,
               sl.lot_code AS supply_lot_code,
               sl.expiration_date AS supply_lot_expiration_date,
               sl.supplier_name AS lot_supplier_name,
               sl.location AS lot_location,
               sl.certificate_path AS lot_certificate_path,
               sl.is_active AS supply_lot_active,
               u.full_name AS responsavel,
               sm.project_id, p.project_code, p.project_name,
               sm.service_id, ps.service_code, ps.title AS service_title, ps.service_type,
               sm.purpose, sm.document_path, sm.created_at
        FROM supply_movements sm
        JOIN supplies s ON s.id=sm.supply_id
        LEFT JOIN supply_lots sl ON sl.id=sm.supply_lot_id
        LEFT JOIN users u ON u.id=sm.user_id
        LEFT JOIN projects p ON p.id=sm.project_id
        LEFT JOIN project_services ps ON ps.id=sm.service_id
        WHERE sm.movement_date >= ?
          AND sm.movement_date <= ?
        ORDER BY sm.movement_date
        """,
        [start_date.isoformat(), end_date.isoformat()],
    )

    supplies = query_df(
        conn,
        """
        SELECT id, supply_type, supply_code, commercial_name, manufacturer,
               manufacturer_code, supply_name, category, physical_state,
               application_function, addition_mode, current_quantity, minimum_quantity,
               unit, lot, expiration_date, location, responsible_name, active,
               safety_doc_path, technical_doc_path, notes
        FROM supplies
        ORDER BY active DESC, supply_name
        """,
    )
    if not supplies.empty:
        supplies["alerta"] = supplies.apply(_supply_alert_status, axis=1)

    supply_lots = query_df(
        conn,
        """
        SELECT sl.id, sl.supply_id, s.supply_type, s.supply_code, s.supply_name,
               sl.lot_code, sl.expiration_date,
               sl.received_date, sl.supplier_name, sl.initial_quantity,
               sl.current_quantity, sl.unit, sl.location, sl.certificate_path,
               sl.notes, sl.is_active, sl.created_at, sl.updated_at
        FROM supply_lots sl
        JOIN supplies s ON s.id=sl.supply_id
        ORDER BY s.supply_name, sl.expiration_date, sl.lot_code
        """,
    )
    if not supply_lots.empty:
        supply_lots["lot_status"] = supply_lots["expiration_date"].map(_lot_expiration_status)

    equipment = query_df(
        conn,
        """
        SELECT equipment_code, equipment_name, lab_unit, location, operational_status,
               unavailable_functions, max_sample_capacity, capacity_unit,
               technical_manager, responsible_name, pop_title, pop_version, active
        FROM equipment
        ORDER BY active DESC, equipment_code
        """,
    )

    services = query_df(
        conn,
        """
        SELECT ps.id, p.project_code, p.project_name, ps.service_code,
               ps.title AS service_title, ps.service_type,
               requester.full_name AS requester_name,
               responsible.full_name AS responsible_name,
               ps.status, ps.requested_date, ps.expected_date, ps.completed_date,
               ps.active, ps.notes, ps.created_at, ps.updated_at
        FROM project_services ps
        JOIN projects p ON p.id=ps.project_id
        LEFT JOIN users requester ON requester.id=ps.requester_id
        LEFT JOIN users responsible ON responsible.id=ps.responsible_id
        WHERE ps.active = 1
          AND (
              SUBSTR(COALESCE(ps.completed_date, ps.expected_date, ps.requested_date, ps.created_at), 1, 10) BETWEEN ? AND ?
              OR LOWER(COALESCE(ps.status, '')) NOT IN ('concluído', 'concluido', 'cancelado', 'arquivado')
          )
        ORDER BY p.project_name, ps.requested_date DESC, ps.title
        """,
        [start_date.isoformat(), end_date.isoformat()],
    )

    supply_lots, supply_movements = _enrich_supply_report_metadata(conn, supply_lots, supply_movements)
    stock_alerts = _stock_minimum_alerts_df(supplies)
    validity_alerts = _lot_validity_alerts_df(supply_lots)
    consumption_detailed = _supply_consumption_detail_df(supply_movements)
    traceability = _supply_traceability_df(supply_movements)

    return {
        "bookings": bookings,
        "preventive": preventive,
        "corrective": corrective,
        "supply_movements": supply_movements,
        "supplies": supplies,
        "supply_lots": supply_lots,
        "stock_alerts": stock_alerts,
        "validity_alerts": validity_alerts,
        "consumption_detailed": consumption_detailed,
        "traceability": traceability,
        "equipment": equipment,
        "services": services,
    }


def _reports_data_fingerprint(conn) -> str:
    counts = cached_table_counts(conn)
    relevant_counts = tuple(
        (table_name, int(counts.get(table_name, 0) or 0))
        for table_name in REPORT_CACHE_TABLES
    )
    return hashlib.sha256(repr(relevant_counts).encode("utf-8")).hexdigest()[:16]


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def _cached_filtered_reports_data(
    database_fingerprint: str,
    start_iso: str,
    end_iso: str,
    data_fingerprint: str,
    _database_url_value: str | None = None,
) -> dict[str, pd.DataFrame]:
    conn = connect(DB_PATH, database_url=_database_url_value)
    try:
        return _filtered_reports_data(
            conn,
            date.fromisoformat(start_iso),
            date.fromisoformat(end_iso),
        )
    finally:
        conn.close()


def filtered_reports_data(
    conn,
    start_date: date,
    end_date: date,
    *,
    data_fingerprint: str | None = None,
) -> dict[str, pd.DataFrame]:
    database_url = _database_url()
    data_fingerprint = data_fingerprint or _reports_data_fingerprint(conn)
    with perf_timer("Dados de relatórios"):
        return _cached_filtered_reports_data(
            _database_fingerprint(database_url),
            start_date.isoformat(),
            end_date.isoformat(),
            data_fingerprint,
            _database_url_value=database_url,
        )


def _reports_excel_signature(
    start_date: date,
    end_date: date,
    database_fingerprint: str,
    data_fingerprint: str,
    cache_generation: int,
) -> str:
    raw = "|".join([
        start_date.isoformat(),
        end_date.isoformat(),
        database_fingerprint,
        data_fingerprint,
        str(cache_generation),
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def _with_booking_duration(bookings: pd.DataFrame) -> pd.DataFrame:
    out = bookings.copy()
    if out.empty:
        return out
    start = pd.to_datetime(out["start_datetime"], errors="coerce")
    end = pd.to_datetime(out["end_datetime"], errors="coerce")
    out["duracao_h"] = ((end - start).dt.total_seconds() / 3600).fillna(0).clip(lower=0)
    out["mes"] = start.dt.to_period("M").astype(str)
    out["status_legivel"] = out["status"].map(STATUS_LABELS).fillna(out["status"])
    return out


def _report_summary(data: dict[str, pd.DataFrame]) -> pd.DataFrame:
    bookings = _with_booking_duration(data["bookings"])
    corrective = data["corrective"]
    preventive = data["preventive"]
    supply_movements = data["supply_movements"]
    supplies = data["supplies"]
    services = data.get("services", pd.DataFrame())
    stock_alerts = data.get("stock_alerts", pd.DataFrame())
    validity_alerts = data.get("validity_alerts", pd.DataFrame())

    total_bookings = len(bookings)
    completed = int((bookings["status"] == "done").sum()) if not bookings.empty else 0
    cancelled = int((bookings["status"] == "cancelled").sum()) if not bookings.empty else 0
    total_samples = int(bookings["sample_count"].fillna(0).sum()) if not bookings.empty and "sample_count" in bookings else 0
    total_hours = float(bookings["duracao_h"].sum()) if not bookings.empty else 0.0
    unique_equipment = int(bookings["equipment_code"].nunique()) if not bookings.empty else 0
    unique_users = int(bookings["solicitante"].nunique()) if not bookings.empty else 0
    downtime = float(corrective["downtime_hours"].fillna(0).sum()) if not corrective.empty and "downtime_hours" in corrective else 0.0
    if stock_alerts.empty and validity_alerts.empty:
        low_stock = int(supplies["alerta"].isin(["Estoque baixo", "Vencido", "Vence em até 60 dias"]).sum()) if not supplies.empty else 0
    else:
        low_stock = len(stock_alerts) + len(validity_alerts)
    outputs = supply_movements[supply_movements["movement_type"].map(_movement_type_is_consumption)] if not supply_movements.empty else supply_movements

    rows = [
        ("Reservas registradas", total_bookings),
        ("Reservas concluídas", completed),
        ("Reservas canceladas", cancelled),
        ("Amostras previstas/registradas", total_samples),
        ("Horas reservadas", round(total_hours, 1)),
        ("Equipamentos utilizados", unique_equipment),
        ("Usuários solicitantes", unique_users),
        ("Preventivas/calibrações", len(preventive)),
        ("Tickets corretivos", len(corrective)),
        ("Serviços/análises", len(services)),
        ("Downtime corretivo (h)", round(downtime, 1)),
        ("Movimentações de insumos", len(supply_movements)),
        ("Saídas/consumo de insumos", len(outputs)),
        ("Insumos em alerta", low_stock),
    ]
    return pd.DataFrame(rows, columns=["Indicador", "Valor"])


def _top_counts(df: pd.DataFrame, group_col: str, value_col: str = "total", limit: int = 10) -> pd.DataFrame:
    if df.empty or group_col not in df.columns:
        return pd.DataFrame(columns=[group_col, value_col])
    return (
        df[group_col]
        .fillna("Não informado")
        .replace("", "Não informado")
        .value_counts()
        .head(limit)
        .reset_index()
        .rename(columns={group_col: value_col, "index": group_col})
    )


def _excel_color(hex_color: str) -> str:
    value = hex_color.strip().lstrip("#").upper()
    return value if len(value) == 8 else f"FF{value}"


EXCEL_BLUE = _excel_color(LAB_BLUE)
EXCEL_CYAN = _excel_color(LAB_CYAN)
EXCEL_TEXT = "FF1F2937"
EXCEL_MUTED_TEXT = "FF64748B"
EXCEL_LIGHT_BG = "FFF8FAFC"
EXCEL_ALT_ROW = "FFF3F7FB"
EXCEL_BORDER = "FFD9E2EC"
EXCEL_WARNING = "FFFFF3CD"
EXCEL_CRITICAL = "FFFDE2E2"
EXCEL_OK = "FFDFF3E4"


def _excel_thin_border() -> Border:
    side = Side(style="thin", color=EXCEL_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _style_cover_sheet(worksheet, period_text: str, generated_at: str) -> None:
    worksheet.sheet_view.showGridLines = False
    for row_idx in range(1, 12):
        worksheet.row_dimensions[row_idx].height = 22
    worksheet.row_dimensions[1].height = 34
    worksheet.row_dimensions[2].height = 26
    worksheet.column_dimensions["A"].width = 4
    worksheet.column_dimensions["B"].width = 24
    worksheet.column_dimensions["C"].width = 32
    worksheet.column_dimensions["D"].width = 24
    worksheet.column_dimensions["E"].width = 24
    worksheet.column_dimensions["F"].width = 4

    if worksheet.max_row:
        worksheet.delete_rows(1, worksheet.max_row)
    worksheet.merge_cells("A1:F1")
    worksheet.merge_cells("A2:F2")
    worksheet.merge_cells("B5:E5")
    worksheet.merge_cells("B6:E6")
    worksheet["A1"] = "LabCim Manager"
    worksheet["A2"] = "Relatório operacional do laboratório"
    worksheet["B5"] = "Período analisado"
    worksheet["B6"] = period_text
    worksheet["B8"] = "Gerado em"
    worksheet["C8"] = generated_at

    title_fill = PatternFill("solid", fgColor=EXCEL_BLUE)
    subtitle_fill = PatternFill("solid", fgColor=EXCEL_CYAN)
    for cell in worksheet[1]:
        cell.fill = title_fill
    for cell in worksheet[2]:
        cell.fill = subtitle_fill
    worksheet["A1"].font = Font(color="FFFFFFFF", bold=True, size=18)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A2"].font = Font(color="FFFFFFFF", bold=True, size=12)
    worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    label_fill = PatternFill("solid", fgColor=EXCEL_LIGHT_BG)
    for cell_ref in ["B5", "B8"]:
        cell = worksheet[cell_ref]
        cell.font = Font(color=EXCEL_TEXT, bold=True)
        cell.fill = label_fill
        cell.border = _excel_thin_border()
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for cell_ref in ["B6", "C8"]:
        cell = worksheet[cell_ref]
        cell.font = Font(color=EXCEL_TEXT, bold=True if cell_ref == "B6" else False)
        cell.border = _excel_thin_border()
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows(min_row=5, max_row=8, min_col=2, max_col=5):
        for cell in row:
            cell.border = _excel_thin_border()


def _excel_status_fill(value) -> PatternFill | None:
    text = clean_input(value).lower()
    if not text:
        return None
    if "vencido" in text or "crítico" in text or "critico" in text:
        return PatternFill("solid", fgColor=EXCEL_CRITICAL)
    if "vence em até 60 dias" in text or "vence em ate 60 dias" in text or "estoque baixo" in text:
        return PatternFill("solid", fgColor=EXCEL_WARNING)
    if text in {"ok", "com certificado", "com anexo"}:
        return PatternFill("solid", fgColor=EXCEL_OK)
    return None


def _style_table_sheet(worksheet, *, emphasize_values: bool = False) -> None:
    worksheet.sheet_view.showGridLines = False
    max_row = worksheet.max_row or 1
    max_col = worksheet.max_column or 1
    if max_row >= 1 and max_col >= 1:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    header_fill = PatternFill("solid", fgColor=EXCEL_BLUE)
    header_font = Font(color="FFFFFFFF", bold=True)
    border = _excel_thin_border()
    alt_fill = PatternFill("solid", fgColor=EXCEL_ALT_ROW)
    base_fill = PatternFill("solid", fgColor="FFFFFFFF")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 26

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=max_row), start=2):
        row_fill = alt_fill if row_idx % 2 == 0 else base_fill
        for cell in row:
            cell.fill = row_fill
            status_fill = _excel_status_fill(cell.value)
            if status_fill is not None:
                cell.fill = status_fill
            cell.font = Font(color=EXCEL_TEXT)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if emphasize_values and cell.column == 2:
                cell.font = Font(color=EXCEL_BLUE, bold=True)

    for column_cells in worksheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        width = min(max(max_len + 2, 12), 55)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def _reports_excel_bytes(period_text: str, data: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    summary = _report_summary(data)
    bookings = _with_booking_duration(data["bookings"])
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    sheets = {
        "Resumo": summary,
        "Reservas": _display_df(bookings),
        "Preventivas": _display_df(data["preventive"]),
        "Corretivas": _display_df(data["corrective"]),
        "Serviços": _display_df(data["services"]),
        "Movimentos insumos": _display_df(data["supply_movements"]),
        "Estoque atual": _display_df(data["supplies"]),
        "Estoque por lote": _display_df(data["supply_lots"]),
        "Alertas estoque": _display_df(data.get("stock_alerts", pd.DataFrame())),
        "Alertas validade": _display_df(data.get("validity_alerts", pd.DataFrame())),
        "Consumo detalhado": _display_df(data.get("consumption_detailed", pd.DataFrame())),
        "Rastreabilidade": _display_df(data.get("traceability", pd.DataFrame())),
        "Equipamentos": _display_df(data["equipment"]),
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame({"Relatório": ["LabCim Manager"], "Período": [period_text], "Gerado em": [generated_at]}).to_excel(
            writer,
            sheet_name="Capa",
            index=False,
        )
        _style_cover_sheet(writer.sheets["Capa"], period_text, generated_at)
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
            worksheet = writer.sheets[safe_name]
            _style_table_sheet(worksheet, emphasize_values=(sheet_name == "Resumo"))
    return output.getvalue()


def _download_table_button(df: pd.DataFrame, file_name: str, label: str, *, allow_empty: bool = False) -> None:
    if not can_export_reports():
        st.caption("Exportações são restritas a Gerente ou Administrador.")
        return
    if df.empty and not allow_empty:
        st.caption("Sem dados para exportar nesta tabela.")
        return
    st.download_button(
        label,
        data=df.to_csv(index=False).encode("utf-8-sig"),
        file_name=file_name,
        mime="text/csv",
        key=f"download_{file_name}",
    )


def page_relatorios(conn):
    hero()
    st.subheader("Relatórios semestrais e anuais")
    st.caption("Consolide uso de equipamentos, responsáveis, manutenções e insumos para acompanhamento interno, reuniões e registros da qualidade.")

    if not can_view_reports():
        st.info("Relatórios e exportações são restritos a Gerente ou Administrador.")
        return

    can_export = can_export_reports()
    today = date.today()
    current_year = today.year
    min_year = 2024
    max_year = current_year + 1

    with st.container(border=True):
        c1, c2, c3 = st.columns([1.4, 1, 1])
        with c1:
            period_mode = st.selectbox(
                "Tipo de relatório",
                ["Semestre atual", "Semestre anterior", "Ano atual", "Ano anterior", "Semestre específico", "Ano específico", "Intervalo personalizado"],
                key="report_period_mode",
            )
        with c2:
            selected_year = st.number_input("Ano", min_value=min_year, max_value=max_year, value=current_year, step=1, key="report_year")
        with c3:
            selected_semester = st.selectbox("Semestre", ["1º semestre", "2º semestre"], key="report_semester")

        if period_mode == "Semestre atual":
            start_date, end_date = _semester_bounds(today.year, 1 if today.month <= 6 else 2)
        elif period_mode == "Semestre anterior":
            start_date, end_date = _previous_semester_bounds(today)
        elif period_mode == "Ano atual":
            start_date, end_date = date(today.year, 1, 1), date(today.year, 12, 31)
        elif period_mode == "Ano anterior":
            start_date, end_date = date(today.year - 1, 1, 1), date(today.year - 1, 12, 31)
        elif period_mode == "Semestre específico":
            start_date, end_date = _semester_bounds(int(selected_year), 1 if selected_semester.startswith("1") else 2)
        elif period_mode == "Ano específico":
            start_date, end_date = date(int(selected_year), 1, 1), date(int(selected_year), 12, 31)
        else:
            d1, d2 = st.columns(2)
            start_date = d1.date_input("Data inicial", value=date(today.year, 1, 1), key="report_custom_start")
            end_date = d2.date_input("Data final", value=today, key="report_custom_end")

        if start_date > end_date:
            st.error("A data inicial não pode ser posterior à data final.")
            return

    period_text = _period_label(start_date, end_date)
    database_url = _database_url()
    database_fingerprint = _database_fingerprint(database_url)
    reports_data_fingerprint = _reports_data_fingerprint(conn)
    cache_generation = int(st.session_state.get(APP_CACHE_GENERATION_KEY, 0) or 0)
    excel_signature = _reports_excel_signature(
        start_date,
        end_date,
        database_fingerprint,
        reports_data_fingerprint,
        cache_generation,
    )
    data = filtered_reports_data(conn, start_date, end_date, data_fingerprint=reports_data_fingerprint)
    bookings = _with_booking_duration(data["bookings"])
    preventive = data["preventive"]
    corrective = data["corrective"]
    supply_movements = data["supply_movements"]
    supplies = data["supplies"]
    supply_lots = data["supply_lots"]
    stock_alerts = data["stock_alerts"]
    validity_alerts = data["validity_alerts"]
    consumption_detailed = data["consumption_detailed"]
    traceability = data["traceability"]
    services = data["services"]
    summary = _report_summary(data)

    st.markdown(f"#### Período analisado: {period_text}")
    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Reservas", len(bookings))
    k2.metric("Concluídas", int((bookings["status"] == "done").sum()) if not bookings.empty else 0)
    k3.metric("Amostras", int(bookings["sample_count"].fillna(0).sum()) if not bookings.empty and "sample_count" in bookings else 0)
    k4.metric("Manutenções", len(preventive) + len(corrective))
    k5.metric("Serviços", len(services))
    k6.metric("Mov. insumos", len(supply_movements))

    if can_export and st.button("Preparar Excel completo", type="primary", key="prepare_full_report_xlsx"):
        st.session_state[REPORT_EXCEL_BYTES_KEY] = _reports_excel_bytes(period_text, data)
        st.session_state[REPORT_EXCEL_SIGNATURE_KEY] = excel_signature

    if (
        can_export
        and st.session_state.get(REPORT_EXCEL_SIGNATURE_KEY) == excel_signature
        and REPORT_EXCEL_BYTES_KEY in st.session_state
    ):
        st.download_button(
            "Baixar relatório completo em Excel",
            data=st.session_state[REPORT_EXCEL_BYTES_KEY],
            file_name=f"LabCim_Relatorio_{start_date.isoformat()}_{end_date.isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_full_report_xlsx",
            type="primary",
        )
    else:
        st.caption("Prepare novamente o Excel para o período selecionado.")

    tab_overview, tab_bookings, tab_projects, tab_maintenance, tab_supplies, tab_tables = st.tabs(
        ["Resumo executivo", "Reservas e uso", "Projetos e serviços", "Manutenção", "Insumos", "Tabelas auditáveis"]
    )

    with tab_overview:
        c1, c2 = st.columns([1, 1])
        with c1:
            st.markdown("### Indicadores consolidados")
            st.dataframe(summary, use_container_width=True, hide_index=True)
        with c2:
            st.markdown("### Leitura rápida")
            if bookings.empty and preventive.empty and corrective.empty and supply_movements.empty and services.empty:
                st.info("Não há registros operacionais no período selecionado.")
            else:
                top_equipment = clean_value(bookings["equipment_name"].value_counts().idxmax()) if not bookings.empty else "-"
                top_user = clean_value(bookings["responsavel_execucao"].value_counts().idxmax()) if not bookings.empty else "-"
                open_services = int((~services["status"].isin(["concluído", "cancelado", "arquivado"])).sum()) if not services.empty else 0
                st.markdown(
                    f"""
                    <div class="soft-card">
                    <b>Equipamento mais demandado:</b> {top_equipment}<br>
                    <b>Responsável/executante mais frequente:</b> {top_user}<br>
                    <b>Serviços/análises em andamento:</b> {open_services}<br>
                    <b>Registros de manutenção:</b> {len(preventive) + len(corrective)}<br>
                    <b>Insumos/lotes em alerta:</b> {len(stock_alerts) + len(validity_alerts)}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        if not bookings.empty:
            st.markdown("### Reservas por mês")
            monthly = bookings.groupby("mes", dropna=False).size().reset_index(name="Reservas")
            fig = px.bar(monthly, x="mes", y="Reservas", color_discrete_sequence=[LAB_BLUE])
            fig.update_layout(height=320, margin=dict(l=20, r=20, t=20, b=20), xaxis_title="Mês", yaxis_title="Reservas")
            st.plotly_chart(fig, use_container_width=True)

    with tab_bookings:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("### Equipamentos mais utilizados")
            if not bookings.empty:
                top_eq = bookings.groupby(["equipment_code", "equipment_name"], dropna=False).size().reset_index(name="reservas").sort_values("reservas", ascending=False).head(10)
                top_eq["equipamento"] = top_eq["equipment_code"].astype(str) + " — " + top_eq["equipment_name"].astype(str)
                fig = px.bar(top_eq, y="equipamento", x="reservas", orientation="h", color_discrete_sequence=[LAB_BLUE])
                fig.update_layout(height=390, margin=dict(l=20, r=20, t=20, b=20), yaxis_title="", xaxis_title="Reservas")
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Sem reservas no período.")
        with c2:
            st.markdown("### Reservas por status")
            if not bookings.empty:
                status_df = bookings.groupby("status_legivel", dropna=False).size().reset_index(name="reservas")
                fig = px.bar(status_df, x="status_legivel", y="reservas", color="status_legivel", color_discrete_sequence=[LAB_BLUE, LAB_CYAN, "#94A3B8", "#F97316"])
                fig.update_layout(height=390, margin=dict(l=20, r=20, t=20, b=20), showlegend=False, xaxis_title="Status", yaxis_title="Reservas")
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Sem reservas no período.")

        st.markdown("### Responsáveis/executantes")
        if not bookings.empty:
            performers = bookings.groupby("responsavel_execucao", dropna=False).agg(
                reservas=("id", "count"),
                amostras=("sample_count", "sum"),
                horas=("duracao_h", "sum"),
            ).reset_index().sort_values("reservas", ascending=False)
            performers["horas"] = performers["horas"].round(1)
            st.dataframe(performers.rename(columns={"responsavel_execucao": "Responsável/executante", "reservas": "Reservas", "amostras": "Amostras", "horas": "Horas"}), use_container_width=True, hide_index=True)
        else:
            st.info("Sem executantes para listar.")

    with tab_projects:
        st.markdown("### Serviços/análises")
        if services.empty:
            st.info("Sem serviços/análises para o período ou em andamento.")
        else:
            st.dataframe(_display_df(services), use_container_width=True, hide_index=True)

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("### Reservas por projeto")
            if bookings.empty:
                st.info("Sem reservas no período.")
            else:
                by_project = bookings.copy()
                by_project["projeto"] = by_project["project_name"].fillna("Não informado")
                project_bookings = by_project.groupby("projeto", dropna=False).agg(
                    reservas=("id", "count"),
                    amostras=("sample_count", "sum"),
                    horas=("duracao_h", "sum"),
                ).reset_index().sort_values("reservas", ascending=False)
                project_bookings["horas"] = project_bookings["horas"].round(1)
                st.dataframe(project_bookings.rename(columns={"projeto": "Projeto", "reservas": "Reservas", "amostras": "Amostras", "horas": "Horas"}), use_container_width=True, hide_index=True)
        with c2:
            st.markdown("### Reservas por serviço/análise")
            if bookings.empty or "service_title" not in bookings.columns or bookings["service_title"].dropna().empty:
                st.info("Sem reservas vinculadas a serviços/análises no período.")
            else:
                by_service = bookings.copy()
                by_service["serviço/análise"] = by_service["service_title"].fillna("Não informado")
                service_bookings = by_service.groupby("serviço/análise", dropna=False).agg(
                    reservas=("id", "count"),
                    amostras=("sample_count", "sum"),
                    horas=("duracao_h", "sum"),
                ).reset_index().sort_values("reservas", ascending=False)
                service_bookings["horas"] = service_bookings["horas"].round(1)
                st.dataframe(service_bookings.rename(columns={"serviço/análise": "Serviço/análise", "reservas": "Reservas", "amostras": "Amostras", "horas": "Horas"}), use_container_width=True, hide_index=True)

        c3, c4 = st.columns(2)
        with c3:
            st.markdown("### Consumo por projeto")
            consumed = supply_movements[supply_movements["movement_type"].map(_movement_type_is_consumption)].copy() if not supply_movements.empty else supply_movements
            if consumed.empty:
                st.info("Sem consumo de insumos no período.")
            else:
                consumed["projeto"] = consumed["project_name"].fillna("Não informado")
                project_consumption = consumed.groupby("projeto", dropna=False).agg(
                    movimentações=("id", "count"),
                    quantidade=("quantity", "sum"),
                ).reset_index().sort_values("movimentações", ascending=False)
                st.dataframe(project_consumption.rename(columns={"projeto": "Projeto", "movimentações": "Movimentações", "quantidade": "Quantidade"}), use_container_width=True, hide_index=True)
        with c4:
            st.markdown("### Consumo por serviço/análise")
            if supply_movements.empty or "service_title" not in supply_movements.columns:
                st.info("Sem consumo vinculado a serviços/análises no período.")
            else:
                consumed_by_service = supply_movements[
                    supply_movements["movement_type"].map(_movement_type_is_consumption)
                    & supply_movements["service_title"].notna()
                ].copy()
                if consumed_by_service.empty:
                    st.info("Sem consumo vinculado a serviços/análises no período.")
                else:
                    consumed_by_service["serviço/análise"] = consumed_by_service["service_title"].fillna("Não informado")
                    service_consumption = consumed_by_service.groupby("serviço/análise", dropna=False).agg(
                        movimentações=("id", "count"),
                        quantidade=("quantity", "sum"),
                    ).reset_index().sort_values("movimentações", ascending=False)
                    st.dataframe(service_consumption.rename(columns={"serviço/análise": "Serviço/análise", "movimentações": "Movimentações", "quantidade": "Quantidade"}), use_container_width=True, hide_index=True)

    with tab_maintenance:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("### Preventivas/calibrações")
            if preventive.empty:
                st.info("Sem preventivas/calibrações no período.")
            else:
                prev_status = preventive.groupby("status", dropna=False).size().reset_index(name="total")
                fig = px.bar(prev_status, x="status", y="total", color="status", color_discrete_sequence=[LAB_BLUE, LAB_CYAN, "#94A3B8"])
                fig.update_layout(height=330, margin=dict(l=20, r=20, t=20, b=20), showlegend=False, xaxis_title="Status", yaxis_title="Total")
                st.plotly_chart(fig, use_container_width=True)
        with c2:
            st.markdown("### Corretivas")
            if corrective.empty:
                st.info("Sem tickets corretivos no período.")
            else:
                corr_status = corrective.groupby("status", dropna=False).agg(
                    tickets=("id", "count"),
                    downtime_h=("downtime_hours", "sum"),
                    custos=("costs", "sum"),
                ).reset_index()
                corr_status["downtime_h"] = corr_status["downtime_h"].fillna(0).round(1)
                corr_status["custos"] = corr_status["custos"].fillna(0).round(2)
                st.dataframe(corr_status.rename(columns={"status": "Status", "tickets": "Tickets", "downtime_h": "Downtime (h)", "custos": "Custos (R$)"}), use_container_width=True, hide_index=True)

        if not corrective.empty:
            st.markdown("### Equipamentos com tickets corretivos")
            corr_eq = corrective.groupby(["equipment_code", "equipment_name"], dropna=False).agg(
                tickets=("id", "count"),
                downtime_h=("downtime_hours", "sum"),
            ).reset_index().sort_values("tickets", ascending=False)
            corr_eq["downtime_h"] = corr_eq["downtime_h"].fillna(0).round(1)
            st.dataframe(corr_eq.rename(columns={"equipment_code": "Código", "equipment_name": "Equipamento", "tickets": "Tickets", "downtime_h": "Downtime (h)"}), use_container_width=True, hide_index=True)

    with tab_supplies:
        validity_display_cols = [
            "lot_status",
            "supply_name",
            "supply_code",
            "lot_code",
            "expiration_date",
            "days_until_expiration",
            "current_quantity",
            "unit",
            "supplier_name",
            "location",
            "certificate_status",
        ]
        validity_display = validity_alerts[[c for c in validity_display_cols if c in validity_alerts.columns]].copy()
        traceability_display_cols = [
            "supply_name",
            "supply_lot_label",
            "supply_lot_expiration_date",
            "lot_supplier_name",
            "lot_location",
            "certificate_status",
            "movement_type",
            "movement_date",
            "quantity",
            "unit",
            "responsavel",
            "project_label",
            "service_label",
            "movement_document_status",
            "purpose",
        ]
        traceability_display = traceability[[c for c in traceability_display_cols if c in traceability.columns]].copy()

        with st.expander("Alertas operacionais", expanded=True):
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("### Alertas de estoque mínimo")
                if stock_alerts.empty:
                    st.success("Sem itens ativos abaixo do estoque mínimo.")
                else:
                    st.dataframe(_display_df(stock_alerts), use_container_width=True, hide_index=True)
                _download_table_button(
                    _display_df(stock_alerts),
                    f"LabCim_alertas_estoque_minimo_{start_date.isoformat()}_{end_date.isoformat()}.csv",
                    "Baixar alertas de estoque em CSV",
                    allow_empty=True,
                )
            with c2:
                st.markdown("### Alertas de validade por lote")
                st.caption("Critério: lotes vencidos ou com vencimento em até 60 dias.")
                if validity_alerts.empty:
                    st.success("Sem lotes ativos vencidos ou próximos do vencimento.")
                else:
                    st.dataframe(_display_df(validity_display), use_container_width=True, hide_index=True)
                _download_table_button(
                    _display_df(validity_alerts),
                    f"LabCim_alertas_validade_lotes_{start_date.isoformat()}_{end_date.isoformat()}.csv",
                    "Baixar alertas de validade em CSV",
                    allow_empty=True,
                )

        with st.expander("Consumo", expanded=True):
            st.markdown("### Movimentações por tipo")
            if supply_movements.empty:
                st.info("Sem movimentações de insumos no período.")
            else:
                move_type = supply_movements.groupby("movement_type", dropna=False).size().reset_index(name="total")
                fig = px.bar(move_type, x="movement_type", y="total", color="movement_type", color_discrete_sequence=[LAB_BLUE, LAB_CYAN, "#F97316", "#94A3B8"])
                fig.update_layout(height=330, margin=dict(l=20, r=20, t=20, b=20), showlegend=False, xaxis_title="Tipo", yaxis_title="Movimentações")
                st.plotly_chart(fig, use_container_width=True)

            st.markdown("### Consumo detalhado por projeto, serviço, insumo e lote")
            if consumption_detailed.empty:
                st.info("Não houve saídas, descartes ou ajustes negativos no período.")
            else:
                st.dataframe(_display_df(consumption_detailed), use_container_width=True, hide_index=True)
            _download_table_button(
                _display_df(consumption_detailed),
                f"LabCim_consumo_detalhado_{start_date.isoformat()}_{end_date.isoformat()}.csv",
                "Baixar consumo detalhado em CSV",
                allow_empty=True,
            )

        with st.expander("Rastreabilidade", expanded=False):
            st.markdown("### Rastreabilidade de insumos e lotes")
            st.caption(
                "Movimentações sem lote aparecem como 'Sem lote'. "
                "Isso pode representar registros antigos ou movimentações feitas sem rastreio por lote."
            )
            if traceability.empty:
                st.info("Sem movimentações de insumos no período selecionado.")
            else:
                traceability_screen = _display_df(traceability_display).rename(
                    columns={"Documento/anexo da movimentação": "Anexo da movimentação"}
                )
                st.dataframe(traceability_screen, use_container_width=True, hide_index=True)
            _download_table_button(
                _display_df(traceability),
                f"LabCim_rastreabilidade_insumos_lotes_{start_date.isoformat()}_{end_date.isoformat()}.csv",
                "Baixar rastreabilidade em CSV",
                allow_empty=True,
            )

        with st.expander("Estoque por lote", expanded=False):
            st.markdown("### Estoque atual por lote")
            st.caption("A validade do item é um campo legado. Para rastreabilidade operacional, priorize a validade do lote.")
            active_lots = supply_lots[supply_lots["is_active"].fillna(1).astype(int) == 1].copy() if not supply_lots.empty else supply_lots
            if active_lots.empty:
                st.info("Nenhum lote ativo cadastrado.")
            else:
                st.dataframe(_display_df(active_lots), use_container_width=True, hide_index=True)

            if not active_lots.empty:
                c3, c4 = st.columns(2)
                with c3:
                    st.markdown("### Lotes vencidos")
                    expired_lots = active_lots[active_lots["lot_status"] == "Vencido"].copy()
                    if expired_lots.empty:
                        st.success("Sem lotes vencidos.")
                    else:
                        st.dataframe(_display_df(expired_lots), use_container_width=True, hide_index=True)
                with c4:
                    st.markdown("### Lotes próximos do vencimento")
                    near_lots = active_lots[active_lots["lot_status"] == "Vence em até 60 dias"].copy()
                    if near_lots.empty:
                        st.success("Sem lotes próximos do vencimento.")
                    else:
                        st.dataframe(_display_df(near_lots), use_container_width=True, hide_index=True)

    with tab_tables:
        st.markdown("### Tabelas auditáveis")
        table_choice = st.selectbox(
            "Tabela",
            [
                "Reservas",
                "Serviços/análises",
                "Preventivas/calibrações",
                "Corretivas",
                "Movimentações de insumos",
                "Estoque atual",
                "Estoque por lote",
                "Alertas estoque",
                "Alertas validade",
                "Consumo detalhado",
                "Rastreabilidade",
                "Equipamentos",
            ],
            key="report_table_choice",
        )
        table_map = {
            "Reservas": bookings,
            "Serviços/análises": services,
            "Preventivas/calibrações": preventive,
            "Corretivas": corrective,
            "Movimentações de insumos": supply_movements,
            "Estoque atual": supplies,
            "Estoque por lote": supply_lots,
            "Alertas estoque": stock_alerts,
            "Alertas validade": validity_alerts,
            "Consumo detalhado": consumption_detailed,
            "Rastreabilidade": traceability,
            "Equipamentos": data["equipment"],
        }
        selected_df = table_map[table_choice]
        if selected_df.empty:
            st.info("Sem registros para a tabela selecionada.")
        else:
            st.dataframe(_display_df(selected_df), use_container_width=True, hide_index=True)
            safe = re.sub(r"[^A-Za-z0-9_-]+", "_", table_choice.lower())
            _download_table_button(selected_df, f"LabCim_{safe}_{start_date.isoformat()}_{end_date.isoformat()}.csv", "Baixar esta tabela em CSV")


def _make_qr_png(url: str) -> bytes:
    img = qrcode.make(url)
    buf = BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _qr_dataframe_signature(df: pd.DataFrame, base_url: str, columns: list[str]) -> str:
    available_columns = [column for column in columns if column in df.columns]
    if df.empty or not available_columns:
        payload = ""
    else:
        payload = df[available_columns].fillna("").astype(str).to_csv(index=False)
    return hashlib.sha256(f"{base_url}\0{payload}".encode("utf-8")).hexdigest()[:16]


def _equipment_has_legacy_pop(equipment_row) -> bool:
    return not is_blank(equipment_row.get("pop_path"))


def _equipment_ids_with_pop_attachments(conn, equipment_ids) -> set[int]:
    normalized_ids: list[int] = []
    for equipment_id in equipment_ids:
        if is_blank(equipment_id):
            continue
        try:
            normalized_ids.append(int(equipment_id))
        except (TypeError, ValueError):
            continue
    if not normalized_ids:
        return set()
    return set(
        _attachment_rows_by_entity(
            conn,
            entity_type="equipment",
            attachment_role="pop",
            entity_ids=normalized_ids,
        ).keys()
    )


def _equipment_has_pop_document(equipment_row, pop_attachment_equipment_ids: set[int] | None = None) -> bool:
    if _equipment_has_legacy_pop(equipment_row):
        return True
    if pop_attachment_equipment_ids is None:
        return False
    equipment_id = equipment_row.get("id")
    if is_blank(equipment_id):
        return False
    try:
        return int(equipment_id) in pop_attachment_equipment_ids
    except (TypeError, ValueError):
        return False


def _equipment_qr_signature(equipment: pd.DataFrame, base_url: str, pop_attachment_equipment_ids: set[int]) -> str:
    base_signature = _qr_dataframe_signature(equipment, base_url, ["id", "equipment_code", "pop_path"])
    attachment_payload = ",".join(str(equipment_id) for equipment_id in sorted(pop_attachment_equipment_ids))
    return hashlib.sha256(f"{base_signature}\0{attachment_payload}".encode("utf-8")).hexdigest()[:16]


def _equipment_qr_zip_bytes(
    equipment: pd.DataFrame,
    base_url: str,
    pop_attachment_equipment_ids: set[int] | None = None,
) -> bytes:
    pop_attachment_equipment_ids = pop_attachment_equipment_ids or set()
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for _, row in equipment.iterrows():
            code = row["equipment_code"]
            suffixes = ["reserva", "manutencao"]
            if _equipment_has_pop_document(row, pop_attachment_equipment_ids):
                suffixes.append("pop")
            for suffix in suffixes:
                url = f"{base_url}?eq={code}&view={suffix}"
                zf.writestr(f"equipamentos/{code}_{suffix}.png", _make_qr_png(url))
    return zip_buf.getvalue()


def _supply_qr_zip_bytes(supplies: pd.DataFrame, base_url: str) -> bytes:
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for _, row in supplies.iterrows():
            supply_id = int(row["id"])
            safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", clean_value(row.get("supply_name")))
            url = f"{base_url}?view=insumo&sid={supply_id}"
            zf.writestr(f"insumos/INSUMO_{supply_id}_{safe_name}.png", _make_qr_png(url))
    return zip_buf.getvalue()


def page_qrcodes(conn):
    hero()
    st.subheader("QR Codes físicos")
    st.caption("Gere QR Codes para fixar em equipamentos, embalagens, prateleiras ou armários. A ideia é reduzir atrito: escaneou, achou a ação certa.")
    equipment = query_df(
        conn,
        """
        SELECT *
        FROM equipment
        WHERE active=1
          AND (operational_status IS NULL OR LOWER(TRIM(operational_status)) <> 'inactive')
        ORDER BY equipment_code
        """,
    )
    supplies = query_df(conn, "SELECT * FROM supplies WHERE active=1 ORDER BY supply_name")
    base_url = st.text_input("URL pública do aplicativo", value="https://labcim-manager.streamlit.app", key="qr_base_url")
    base_url = base_url.rstrip("/")

    tab_eq, tab_sup = st.tabs(["Equipamentos", "Insumos"])

    with tab_eq:
        st.markdown("### QR Codes por equipamento")
        if equipment.empty:
            st.info("Nenhum equipamento ativo encontrado.")
        else:
            equipment_ids = equipment["id"].tolist() if "id" in equipment.columns else []
            pop_attachment_equipment_ids = _equipment_ids_with_pop_attachments(conn, equipment_ids)
            eq = st.selectbox("Equipamento", equipment["equipment_code"].tolist(), key="qr_equipment")
            selected = equipment[equipment["equipment_code"] == eq].iloc[0]
            selected_has_pop_document = _equipment_has_pop_document(selected, pop_attachment_equipment_ids)
            pop_summary = clean_value(
                selected.get("pop_title"),
                "documento operacional cadastrado" if selected_has_pop_document else "não cadastrado",
            )
            st.markdown(
                f"""
                <div class="soft-card">
                <b>{clean_value(selected.get('equipment_code'))} — {clean_value(selected.get('equipment_name'))}</b><br>
                Local: {clean_value(selected.get('location'))} · Responsável: {clean_value(selected.get('responsible_name'))}<br>
                POP: {pop_summary}
                </div>
                """,
                unsafe_allow_html=True,
            )

            cards = [
                ("Reservar / Ver agenda", "reserva", "Aponte a câmera para reservar ou consultar a agenda deste equipamento."),
                ("Reportar problema / Manutenção", "manutencao", "Aponte a câmera para abrir um ticket de suporte/manutenção."),
            ]
            if selected_has_pop_document:
                cards.append(("POP / documentação operacional", "pop", "Aponte a câmera para consultar ou baixar o POP/documentação operacional."))

            cols = st.columns(len(cards))
            for (label, suffix, instruction), col in zip(cards, cols):
                url = f"{base_url}?eq={eq}&view={suffix}"
                png = _make_qr_png(url)
                with col:
                    st.markdown(f"#### {label}")
                    st.image(png, width=230)
                    st.caption(instruction)
                    st.code(url)
                    st.download_button(
                        f"Baixar QR - {label}",
                        data=png,
                        file_name=f"{eq}_{suffix}.png",
                        mime="image/png",
                        key=f"download_{eq}_{suffix}",
                    )

            st.markdown("### Baixar todos os QR Codes de equipamentos")
            if can_export_qr_bulk():
                st.caption("Gera um pacote ZIP com QR Codes de reserva, manutenção e POP quando houver documentação cadastrada.")
                equipment_zip_signature = _equipment_qr_signature(equipment, base_url, pop_attachment_equipment_ids)
                if st.button("Preparar ZIP - QR Codes de equipamentos", key="prepare_all_equipment_qr"):
                    st.session_state["equipment_qr_zip_bytes"] = _equipment_qr_zip_bytes(
                        equipment,
                        base_url,
                        pop_attachment_equipment_ids,
                    )
                    st.session_state["equipment_qr_zip_signature"] = equipment_zip_signature
                if st.session_state.get("equipment_qr_zip_signature") == equipment_zip_signature:
                    st.download_button(
                        "Baixar ZIP - QR Codes de equipamentos",
                        data=st.session_state["equipment_qr_zip_bytes"],
                        file_name="LabCim_QRCodes_Equipamentos.zip",
                        mime="application/zip",
                        key="download_all_equipment_qr",
                    )
                else:
                    st.caption("Prepare o ZIP para habilitar o download do pacote completo.")
            else:
                st.caption("Geração de pacotes ZIP de QR Codes é restrita a Gerente ou Administrador.")

    with tab_sup:
        st.markdown("### QR Codes por insumo")
        if supplies.empty:
            st.info("Nenhum insumo ativo cadastrado. Cadastre insumos antes de gerar QR Codes do almoxarifado.")
        else:
            supply_label = st.selectbox("Insumo", _supply_options(supplies), key="qr_supply")
            supply_id = _supply_id_from_label(supplies, supply_label)
            selected_supply = supplies[supplies["id"] == supply_id].iloc[0]
            render_supply_quick_card(conn, selected_supply)

            url = f"{base_url}?view=insumo&sid={supply_id}"
            png = _make_qr_png(url)
            c1, c2 = st.columns([1, 2])
            with c1:
                st.image(png, width=260)
                st.download_button(
                    "Baixar QR do insumo",
                    data=png,
                    file_name=f"INSUMO_{supply_id}_{re.sub(r'[^A-Za-z0-9_-]+', '_', clean_value(selected_supply.get('supply_name')))}.png",
                    mime="image/png",
                    key=f"download_supply_qr_{supply_id}",
                )
            with c2:
                st.markdown("#### Ficha rápida / movimentação")
                st.caption("Cole este QR Code na embalagem, prateleira ou armário. Ao escanear, o usuário verá saldo, lote, validade, localização, responsável e documentos do insumo.")
                st.code(url)

            st.markdown("### Baixar todos os QR Codes de insumos")
            if can_export_qr_bulk():
                supply_zip_signature = _qr_dataframe_signature(supplies, base_url, ["id", "supply_name"])
                if st.button("Preparar ZIP - QR Codes de insumos", key="prepare_all_supply_qr"):
                    st.session_state["supply_qr_zip_bytes"] = _supply_qr_zip_bytes(supplies, base_url)
                    st.session_state["supply_qr_zip_signature"] = supply_zip_signature
                if st.session_state.get("supply_qr_zip_signature") == supply_zip_signature:
                    st.download_button(
                        "Baixar ZIP - QR Codes de insumos",
                        data=st.session_state["supply_qr_zip_bytes"],
                        file_name="LabCim_QRCodes_Insumos.zip",
                        mime="application/zip",
                        key="download_all_supply_qr",
                    )
                else:
                    st.caption("Prepare o ZIP para habilitar o download do pacote completo.")
            else:
                st.caption("Geração de pacotes ZIP de QR Codes é restrita a Gerente ou Administrador.")


def page_importar(conn):
    hero()
    st.subheader("Importar base inicial")
    if not can_import_base():
        st.info("Importação de base é restrita a Administrador.")
        return
    st.write("Use esta página para atualizar a base a partir do arquivo `LabCim_Base.xlsx`.")
    uploaded = st.file_uploader("Enviar arquivo Excel", type=["xlsx"])
    if uploaded is not None:
        tmp = Path("data/_uploaded_base.xlsx")
        tmp.write_bytes(uploaded.getvalue())
        if st.button("Importar arquivo enviado", type="primary"):
            try:
                counts = import_base_xlsx(conn, tmp)
                st.success(f"Importação concluída: {counts}")
                clear_app_caches()
                st.rerun()
            except Exception as exc:
                st.error(f"Erro na importação: {exc}")

    if BASE_XLSX.exists():
        if st.button("Reimportar arquivo local data/LabCim_Base.xlsx"):
            counts = import_base_xlsx(conn, BASE_XLSX)
            st.success(f"Importação concluída: {counts}")
            clear_app_caches()
            st.rerun()

    st.markdown("### Contagem atual")
    st.json(cached_table_counts(conn))


def apply_url_params_hint():
    params = st.query_params
    if "eq" in params:
        st.sidebar.success(f"Equipamento via QR: {params.get('eq')}")
    if params.get("view") == "manutencao":
        st.sidebar.info("QR de manutenção detectado. Use a aba Manutenção.")
    if params.get("view") == "pop":
        st.sidebar.info("QR de POP/documentação detectado.")
    if params.get("view") == "insumo":
        st.sidebar.success(f"Insumo via QR: {params.get('sid', '-')}")


def main():
    setup_page()
    _reset_perf_events()
    with perf_timer("get_conn"):
        conn = get_conn()
    if not is_authenticated():
        with perf_timer("Página: Login"):
            page_login(conn)
        _render_perf_debug()
        return
    if revalidate_authenticated_user(conn) is None:
        st.warning("Sua sessão não está mais válida. Faça login novamente.")
        with perf_timer("Página: Login"):
            page_login(conn)
        _render_perf_debug()
        return
    apply_url_params_hint()
    page = sidebar()
    page = render_mobile_menu_navigation(page)
    scroll_to_top_on_page_change(page)
    with perf_timer(f"Página: {page}"):
        if page == "Painel inicial":
            page_dashboard(conn)
        elif page == "Reservas":
            page_reservas(conn)
        elif page == "Equipamentos":
            page_equipamentos(conn)
        elif page == "Insumos":
            page_insumos(conn)
        elif page == "Usuários":
            page_usuarios(conn)
        elif page == "Projetos":
            page_projetos(conn)
        elif page == "Manutenção":
            page_manutencao(conn)
        elif page == "QR Codes":
            page_qrcodes(conn)
        elif page == "Relatórios":
            page_relatorios(conn)
        elif page == "Importar base":
            page_importar(conn)
    _render_perf_debug()


if __name__ == "__main__":
    main()
